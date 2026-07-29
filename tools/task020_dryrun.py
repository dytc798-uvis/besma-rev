from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook


def norm(value: str | None) -> str:
    return (value or "").replace("\\", "/").lstrip("/")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=Path, required=True)
    parser.add_argument("--inventory", type=Path, required=True)
    parser.add_argument("--hashes", type=Path, required=True)
    parser.add_argument("--forms-reference", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)

    inventory = [json.loads(line) for line in args.inventory.read_text(encoding="utf-8").splitlines()]
    hashes = [json.loads(line) for line in args.hashes.read_text(encoding="utf-8").splitlines()]
    storage_hashes = {
        f"documents/{norm(x['path'])}": x
        for x in hashes
        if x["root"] == "/srv/besma/storage/documents"
    }
    form_hashes = {
        norm(x["path"]): x for x in hashes if x["root"] == "/srv/besma/docs/base"
    }
    storage = {norm(x["path"]): x for x in inventory if x["root"] == "/srv/besma/storage"}
    by_instance: dict[int, list[dict]] = defaultdict(list)
    for rel, item in storage.items():
        parts = rel.split("/")
        if len(parts) >= 4 and parts[:2] == ["documents", "by_instance"] and parts[2].isdigit():
            if "__original" not in parts[-1].lower() and "__optimized" not in parts[-1].lower():
                by_instance[int(parts[2])].append(item)
        elif rel.startswith("documents/instance_"):
            try:
                by_instance[int(parts[-1].split("_", 2)[1])].append(item)
            except (IndexError, ValueError):
                pass

    db = sqlite3.connect(f"file:{args.db.as_posix()}?mode=ro", uri=True)
    db.row_factory = sqlite3.Row
    docs = db.execute(
        """select d.*, i.document_type_code, i.period_start as instance_period_start,
                  i.period_end as instance_period_end, s.site_name
             from documents d
             left join document_instances i on i.id=d.instance_id
             left join sites s on s.id=d.site_id
            order by d.id"""
    ).fetchall()

    rows = []
    for d in docs:
        before = norm(d["file_path"])
        direct = storage.get(before)
        candidates = by_instance.get(d["instance_id"], [])
        if direct:
            grade, reason, proposed = "CONNECTED", "DB 경로와 실파일 경로가 정확히 일치", before
        else:
            same_ext = [x for x in candidates if Path(x["path"]).suffix.lower() == Path(d["file_name"] or before).suffix.lower()]
            same_size = [x for x in same_ext if d["file_size"] is not None and x["size"] == d["file_size"]]
            if len(candidates) == 1 and len(same_size) == 1:
                grade = "A"
                proposed = norm(same_size[0]["path"])
                reason = "instance_id 전용 디렉터리·단일 파일·확장자·DB 크기 일치"
            elif len(same_size) == 1:
                grade = "B"
                proposed = norm(same_size[0]["path"])
                reason = "instance_id와 크기·확장자 일치 후보는 1개이나 동일 디렉터리에 복수 파일 존재"
            elif candidates:
                grade = "C"
                proposed = ""
                reason = "instance_id 후보는 있으나 크기/확장자 또는 후보 유일성 불일치"
            else:
                grade = "D"
                proposed = ""
                reason = "현재 서버 인벤토리에 동일 instance_id 후보 없음"
            direct = storage.get(proposed)
        rows.append(
            {
                "document_id": d["id"],
                "instance_id": d["instance_id"],
                "site_id": d["site_id"],
                "site_name": d["site_name"],
                "document_type": d["document_type_code"] or d["document_type"],
                "period_start": str(d["instance_period_start"] or d["period_start"] or ""),
                "period_end": str(d["instance_period_end"] or d["period_end"] or ""),
                "db_file_name": d["file_name"],
                "db_file_path_before": before,
                "db_file_size": d["file_size"],
                "candidate_count": len(candidates),
                "candidate_paths": " | ".join(norm(x["path"]) for x in candidates),
                "proposed_file_path_after": proposed,
                "candidate_size": direct["size"] if direct else None,
                "candidate_sha256": storage_hashes.get(proposed, {}).get("sha256"),
                "candidate_head16": storage_hashes.get(proposed, {}).get("head16"),
                "extension_match": bool(proposed and Path(proposed).suffix.lower() == Path(d["file_name"] or before).suffix.lower()),
                "size_match": bool(direct and d["file_size"] == direct["size"]),
                "grade": grade,
                "evidence": reason,
                "planned_action": "변경 없음" if grade == "CONNECTED" else ("DB file_path 후보 제시" if grade in {"A", "B"} else "자동 적용 금지"),
                "dry_run": True,
            }
        )

    reference_by_size: dict[int, list[tuple[Path, str]]] = defaultdict(list)
    for p in args.forms_reference.rglob("*"):
        if p.is_file():
            reference_by_size[p.stat().st_size].append(
                (p, hashlib.sha256(p.read_bytes()).hexdigest())
            )
    form_rows = []
    for item in (x for x in inventory if x["root"] == "/srv/besma/docs/base"):
        candidates = reference_by_size[item["size"]]
        server_hash = form_hashes.get(norm(item["path"]), {}).get("sha256")
        matches = [(p, digest) for p, digest in candidates if digest == server_hash]
        if len(matches) == 1:
            rel = matches[0][0].relative_to(args.forms_reference).as_posix()
            grade = "A"
            evidence = "서버 SHA-256·파일 크기와 로컬 기준 양식이 유일하게 일치"
            proposed = rel
        else:
            grade = "C"
            evidence = "기준 양식 크기 일치 후보가 없거나 복수"
            proposed = ""
        form_rows.append(
            {
                "source_path": norm(item["path"]),
                "source_size": item["size"],
                "proposed_relative_path": proposed,
                "reference_candidate_count": len(matches),
                "server_sha256": server_hash,
                "reference_sha256": matches[0][1] if len(matches) == 1 else None,
                "grade": grade,
                "evidence": evidence,
                "planned_action": "이름/경로 변경 금지; 해시 재검증 후 별도 적용안 검토",
                "dry_run": True,
            }
        )

    payload = {
        "generated_for": "TASK020",
        "mode": "DRY_RUN",
        "db_updates": 0,
        "file_changes": 0,
        "summary": {
            "documents_total": len(rows),
            **{f"grade_{g}": sum(r["grade"] == g for r in rows) for g in ["CONNECTED", "A", "B", "C", "D"]},
            "forms_total": len(form_rows),
            **{f"forms_grade_{g}": sum(r["grade"] == g for r in form_rows) for g in ["A", "B", "C", "D"]},
        },
        "documents": rows,
        "forms": form_rows,
    }
    (args.out / "020_파일연결복구_DRYRUN_매핑.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "문서매핑"
    headers = list(rows[0])
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    fs = wb.create_sheet("본사양식")
    fheaders = list(form_rows[0])
    fs.append(fheaders)
    for row in form_rows:
        fs.append([row[h] for h in fheaders])
    summary = wb.create_sheet("요약")
    summary.append(["항목", "건수"])
    for key, value in payload["summary"].items():
        summary.append([key, value])
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    wb.save(args.out / "020_파일연결복구_DRYRUN_매핑.xlsx")

    sql = [
        "-- TASK020 REVIEW ONLY / APPLY DISABLED",
        "-- 이 파일의 UPDATE 문은 모두 주석이며 실행해도 DB를 변경하지 않습니다.",
        "BEGIN;",
    ]
    for row in rows:
        if row["grade"] in {"A", "B"}:
            before = row["db_file_path_before"].replace("'", "''")
            after = row["proposed_file_path_after"].replace("'", "''")
            sql.append(
                f"-- UPDATE documents SET file_path='{after}' WHERE id={row['document_id']} "
                f"AND file_path='{before}'; -- grade {row['grade']}"
            )
    sql.extend(["ROLLBACK;", "-- DB 변경 0건 보장: 실제 적용 스위치/실행문 없음"])
    (args.out / "020_파일연결복구_검토전용.sql").write_text("\n".join(sql) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
