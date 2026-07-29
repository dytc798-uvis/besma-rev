#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")


def val(row, key, default=None):
    if row is None:
        return default
    return row[key] if key in row.keys() and row[key] is not None else default


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, col).value or "") for row in range(1, min(ws.max_row, 200) + 1)]
        width = min(max(max((len(v) for v in values), default=8) + 2, 10), 42)
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def save_book(path: Path, sheets: list[tuple[str, list[str], list[list]]]):
    wb = Workbook()
    wb.remove(wb.active)
    for name, headers, rows in sheets:
        ws = wb.create_sheet(name[:31])
        ws.append(headers)
        for row in rows:
            ws.append(row)
        style_sheet(ws)
    wb.save(path)


def table_exists(db, name):
    return db.execute(
        "select 1 from sqlite_master where type='table' and name=?", (name,)
    ).fetchone() is not None


def counts_by_user(db, table, column):
    if not table_exists(db, table):
        return {}
    return dict(
        db.execute(
            f"select {column}, count(*) from {table} where {column} is not null group by {column}"
        ).fetchall()
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True, type=Path)
    parser.add_argument("--erp", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args()
    args.out.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(args.db)
    db.row_factory = sqlite3.Row

    erp_by_name = defaultdict(list)
    with args.erp.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            if (row.get("erp_login_id") or "").strip():
                erp_by_name[(row.get("name") or "").strip()].append(row)

    persons = {r["id"]: r for r in db.execute("select * from persons")}
    sites = {r["id"]: r for r in db.execute("select * from sites")}
    log_stats = {
        r["user_id"]: r
        for r in db.execute(
            """
            select user_id,
                   max(case when succeeded=1 then created_at end) last_success,
                   sum(case when succeeded=1 then 1 else 0 end) success_count,
                   sum(case when succeeded=0 then 1 else 0 end) failure_count
            from auth_login_events where user_id is not null group by user_id
            """
        )
    }
    usage_sources = {
        "문서업로드": counts_by_user(db, "documents", "uploaded_by_user_id"),
        "문서제출": counts_by_user(db, "documents", "submitter_user_id"),
        "문서검토": counts_by_user(db, "document_review_histories", "action_by_user_id"),
        "문서의견": counts_by_user(db, "document_comments", "user_id"),
        "평가": counts_by_user(db, "functional_eval_assessments", "updated_by_user_id"),
        "전자동의": counts_by_user(db, "functional_eval_consents", "user_id"),
        "전자서명": counts_by_user(db, "functional_eval_signatures", "signer_user_id"),
        "현장소통": counts_by_user(db, "communications", "sender_user_id"),
    }

    active = list(db.execute("select * from users where is_active=1 order by id"))
    same_name = defaultdict(list)
    same_person = defaultdict(list)
    for user in active:
        same_name[user["name"]].append(user)
        if user["person_id"] is not None:
            same_person[user["person_id"]].append(user)
    shared_person_groups = {
        pid: rows for pid, rows in same_person.items() if len(rows) > 1
    }
    duplicate_group_names = {name for name, rows in same_name.items() if len(rows) > 1}
    common_ids = {
        u["id"]
        for u in active
        if u["login_id"].lower() in {
            "hqsafe1", "hqsafe2", "six", "work", "year2026", "testtest", "cost"
        }
        or u["login_id"].lower().startswith("public-")
        or "공용" in (u["name"] or "")
        or "관리용" in (u["name"] or "")
    }

    account_rows = []
    for u in active:
        person = persons.get(u["person_id"])
        site = sites.get(u["site_id"])
        erp = erp_by_name.get((u["name"] or "").strip(), [])
        logs = log_stats.get(u["id"])
        usage = {label: source.get(u["id"], 0) for label, source in usage_sources.items()}
        usage_total = sum(usage.values())
        if u["person_id"] in shared_person_groups:
            duplicate = "중복 가능성 높음(동일 person_id, 업무범위 확인 필요)"
        elif u["name"] in duplicate_group_names:
            duplicate = "추가 확인 필요(동명이인·복수현장·별도업무 가능)"
        else:
            duplicate = ""
        if u["id"] in common_ids:
            duplicate = (duplicate + "; " if duplicate else "") + "공용 계정 후보"
        issued_by = u["account_issued_by"] or "기록 없음"
        account_rows.append(
            [
                u["id"],
                u["login_id"],
                u["name"],
                "있음" if u["birth_date"] else "없음",
                "있음" if person and person["phone_mobile"] else "없음",
                u["department"],
                "별도 직무/직책 컬럼 없음",
                val(site, "site_code") if site else None,
                val(site, "site_name") if site else None,
                u["role"],
                "활성",
                issued_by,
                "예" if issued_by == "self_service_site" else "아니오",
                "예" if issued_by == "self_service_hq" else "아니오",
                ", ".join(sorted({r["erp_login_id"] for r in erp if r.get("erp_login_id")})),
                val(logs, "last_success"),
                val(logs, "success_count", 0),
                val(logs, "failure_count", 0),
                "변경 이력 있음" if u["password_changed_at"] else "변경 이력 없음",
                duplicate,
                "예" if u["id"] in common_ids else "아니오",
                "관찰됨" if val(logs, "success_count", 0) or usage_total else "관찰 기간 내 근거 없음",
                ", ".join(f"{k}:{v}" for k, v in usage.items() if v) or "-",
                "개별 업무·승인자 확인" if u["role"] in {
                    "HQ_OTHER", "HQ_OUTSOURCING_PURCHASE", "FUNCTIONAL_EVAL_VIEWER"
                } else "역할 정책 대조",
                "삭제·병합 금지; person/employee 식별자 확인 후 처리"
                if duplicate
                else "현행 유지",
                "예" if duplicate or u["id"] in common_ids else "아니오",
            ]
        )
    account_headers = [
        "사용자ID","login_id","이름","생년월일 저장","휴대전화 저장","부서","직무/직책",
        "현장코드","현장명","역할","활성","생성방식","self_service_site","self_service_hq",
        "ERP ID/별칭","마지막 성공 로그인","성공 로그인 건수","실패 로그인 건수",
        "비밀번호 변경","중복 분류","공용 계정 후보","실제 사용 관찰","업무 이력",
        "권한 과다·부족 진단","권장 처리","사용자 확인 필요",
    ]
    role_counts = Counter(u["role"] for u in active)
    role_summary = [
        [role, count, sum(1 for u in active if u["role"] == role and val(log_stats.get(u["id"]), "success_count", 0))]
        for role, count in sorted(role_counts.items())
    ]
    duplicate_summary = [
        ["확정 중복", 0, "대상자·관리자 확인 전 확정하지 않음"],
        ["동일 person_id 복수 활성 계정 그룹", len(shared_person_groups), "동일인 강한 근거이나 복수 현장/업무 계정 가능"],
        ["동일 이름 복수 활성 계정 그룹", len(duplicate_group_names), "동명이인·복수현장 포함"],
        ["공용 계정 후보", len(common_ids), "실제 사용자·용도 확인 전 비활성화 금지"],
    ]
    save_book(
        args.out / "021_활성계정_역할_중복분석표.xlsx",
        [
            ("활성계정404", account_headers, account_rows),
            ("역할요약", ["역할","활성계정","성공로그인 관찰계정"], role_summary),
            ("중복공용요약", ["분류","그룹/계정수","판정 기준"], duplicate_summary),
        ],
    )

    mapping_rows = [
        ["안전","안전보건실","HQ_SAFE","없음(단일 역할 스키마)","안전보건실 지정 승인자","전사","안전·문서","안전/문서 API","정기 재검토","퇴사·부서이동·업무종료"],
        ["공사","공사관리 계열","미확정","현재 역할 조합 검토","해당 부서 책임자","업무범위 확인","미확정","미확정","승인 시 지정","부서이동·업무종료"],
        ["공무","DB에 정확한 공무팀 없음","미확정","현재 역할 조합 검토","해당 부서 책임자","업무범위 확인","미확정","미확정","승인 시 지정","부서이동·업무종료"],
        ["예산·견적","예산견적팀","HQ_BUDGET_ESTIMATE","없음","해당 부서 책임자","본사 업무범위","예산·견적","예산·견적 API","정기 재검토","부서이동·업무종료"],
        ["외주·구매","외주구매팀","HQ_OUTSOURCING_PURCHASE","없음","해당 부서 책임자","본사 업무범위","외주·구매","외주·구매 API","정기 재검토","부서이동·업무종료"],
        ["현장","현장","SITE","SITE_FUNCTIONAL_EVAL은 별도 검토","현장소장 또는 본사 지정 승인자","승인 현장 1개","현장 문서","자기 현장 API","현장 종료일까지","현장 종료·이동"],
        ["기능인 평가 조회","본사 조회업무","FUNCTIONAL_EVAL_VIEWER","없음","기능인평가 지정 승인자","전사 평가 조회","기능인평가 조회","기능인평가 읽기 API","정기 재검토","업무 종료"],
        ["기타","확인 필요","미확정","없음","상위 관리자","확인 후 지정","미확정","미확정","단기 권장","업무 종료"],
        ["고권한 관리자","안전보건 관리자","HQ_SAFE_ADMIN/SUPER_ADMIN","없음","별도 상위 승인","전사","관리자","관리 API","단기·정기 재검토","즉시 회수"],
    ]
    save_book(
        args.out / "021_업무구분_역할_승인자_매핑표.xlsx",
        [("매핑표", ["사용자 업무구분","실제 부서","기본 역할","추가 가능 역할","승인자","현장 범위","허용 메뉴","허용 API","유효기간","회수 조건"], mapping_rows)],
    )

    role_policy = {
        "HQ_SAFE": ("전사", "허용", "허용", "안전 문서", "허용"),
        "HQ_SAFE_ADMIN": ("전사", "허용", "허용", "안전 문서+계정승인", "허용"),
        "SUPER_ADMIN": ("전사", "허용", "허용", "전체 관리자", "허용"),
        "ACCIDENT_ADMIN": ("전사", "허용", "허용", "사고+HQ 안전 UI", "허용"),
        "SITE": ("자기 현장", "허용", "타 현장 403", "현장", "자기 현장만"),
        "SITE_FUNCTIONAL_EVAL": ("자기 현장", "허용", "타 현장 403", "현장 기능인평가", "자기 현장만"),
        "HQ_OTHER": ("조회 정책", "차단 403", "차단 403", "본사 조회", "파일 차단"),
        "HQ_OUTSOURCING_PURCHASE": ("업무 메뉴", "차단 403", "차단 403", "외주·구매", "파일 차단"),
        "HQ_BUDGET_ESTIMATE": ("업무 메뉴", "차단 403", "차단 403", "예산·견적", "파일 차단"),
        "FUNCTIONAL_EVAL_VIEWER": ("평가 전사조회", "차단 403", "차단 403", "기능인평가 조회", "파일 차단"),
        "WORKER": ("개인 범위", "차단 403", "차단 403", "근로자", "파일 차단"),
    }
    selected = {}
    for u in active:
        selected.setdefault(u["role"], u)
    permission_rows = []
    for role in sorted(role_policy):
        scope, own, other, menu, api = role_policy[role]
        user = selected.get(role)
        permission_rows.append([
            role, role_counts.get(role, 0), user["id"] if user else None, user["login_id"] if user else None,
            "가능(활성 계정 존재)" if user else "활성 계정 없음", menu, scope,
            "역할별 라우트 정책", "역할별 라우트 정책", own, own, "역할별 업무 정책",
            "HQ 안전/관리자만", "역할별 정책", "HQ_SAFE_ADMIN/SUPER_ADMIN만",
            "HQ_SAFE_ADMIN/SUPER_ADMIN만", api,
            "실계정 비밀번호·세션 미제공으로 HTTP 로그인 테스트 불가; 코드+자동테스트 판정",
        ])
    live_rows = [
        ["미로그인","GET /documents",401,"운영 HTTP","통과"],
        ["미로그인","GET /documents/29/file",401,"운영 HTTP","통과"],
        ["미로그인","GET /document-submissions/collection?...site_id=8",401,"운영 HTTP","통과"],
        ["미로그인","GET /documents/999999/file",401,"인증 선행","통과"],
        ["CORS","OPTIONS /documents/29/file",200,"allow-origin https://www.besma.co.kr","통과"],
        ["HQ_SAFE","대표 실계정",None,"비밀번호/기존 세션 미제공","테스트 불가"],
        ["SITE","대표 실계정",None,"비밀번호/기존 세션 미제공","테스트 불가"],
        ["HQ_OTHER","대표 실계정",None,"비밀번호/기존 세션 미제공","테스트 불가"],
        ["외주구매·조회","대표 실계정",None,"비밀번호/기존 세션 미제공","테스트 불가"],
    ]
    save_book(
        args.out / "021_실계정_파일권한_검증표.xlsx",
        [
            ("역할별대응", ["역할","활성수","대표 사용자ID","대표 login_id","로그인 가능","노출 메뉴","조회 현장","문서 목록","문서 상세","다운로드","미리보기","업로드","승인·반려","기능인평가","사용자관리","권한승인","API 직접접근","검증 한계"], permission_rows),
            ("운영HTTP", ["유형","요청","HTTP","근거","결과"], live_rows),
        ],
    )

    comparison = [
        ["1. DB를 깨진 실파일명에 맞춤","A 274 즉시 연결 가능","낮음","낮음","DB 되돌림 쉬움","매우 낮음","경로 제어문자·인코딩 위험","낮음","낮음","가능하나 비권장","양식도 비권장","깨진 경로 영구화로 제외"],
        ["2. 실파일명을 정상화하고 DB 정상 경로 유지","DB 원본명·크기·해시로 A 274 검증","중간","중간(원자적 rename 필요)","적용 로그로 역 rename","높음","경로 이탈·충돌 사전검사","중간~높음","중간~높음","274건 적용 가능","A 86건은 별도 도구 필요","단기 A등급 권장"],
        ["3. UUID 저장키+원본 표시명 분리","가장 높음","높음","전면 구조 변경 시 높음","마이그레이션 설계 필요","가장 높음","가장 높음","가장 높음","가장 높음","장기 적용","장기 적용","장기 표준"],
    ]
    save_book(
        args.out / "021_A등급_복구방식_비교표.xlsx",
        [("복구방식비교", ["방식","복구 정확성","구현 난이도","운영 중단 위험","롤백","한글 호환성","보안","유지보수","CMS 이관","문서274","기준양식93","결론"], comparison)],
    )

    summary = {
        "generated_at": datetime.now().isoformat(),
        "active_accounts": len(active),
        "inactive_accounts": db.execute("select count(*) from users where is_active=0").fetchone()[0],
        "role_counts": dict(sorted(role_counts.items())),
        "successful_login_observed_accounts": sum(
            1 for u in active if val(log_stats.get(u["id"]), "success_count", 0)
        ),
        "confirmed_duplicate_groups": 0,
        "same_person_id_multi_account_groups": len(shared_person_groups),
        "same_name_multi_account_groups": len(duplicate_group_names),
        "common_account_candidates": len(common_ids),
        "birth_date_present": sum(1 for u in active if u["birth_date"]),
        "phone_present_via_person": sum(
            1 for u in active if persons.get(u["person_id"]) and persons[u["person_id"]]["phone_mobile"]
        ),
        "password_changed_at_present": sum(1 for u in active if u["password_changed_at"]),
        "must_change_password": sum(1 for u in active if u["must_change_password"]),
        "erp_alias_unique_names": len(erp_by_name),
        "hq_like_active": sum(1 for u in active if u["site_id"] is None),
        "hq_name_erp_match": sum(1 for u in active if u["site_id"] is None and u["name"] in erp_by_name),
        "login_id_direct_erp": sum(
            1
            for u in active
            if any(u["login_id"] == row.get("erp_login_id") for row in erp_by_name.get(u["name"], []))
        ),
        "hq_login_id_direct_erp": sum(
            1
            for u in active
            if u["site_id"] is None
            and any(u["login_id"] == row.get("erp_login_id") for row in erp_by_name.get(u["name"], []))
        ),
    }
    (args.out / "021_계정분석_요약.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
