"""현장별 기능인등급 엑셀 — docs/01.*기능인등급*.xlsx 템플릿 기반 출력."""

from __future__ import annotations

import io
import re
import shutil
import tempfile
from datetime import date
from pathlib import Path
from typing import Any, Literal

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.worksheet import Worksheet

from app.config.settings import BASE_DIR
from app.modules.functional_eval.eval_catalog import get_criteria

TEMPLATE_GLOB = "01.*기능인등급*.xlsx"
MODULE_TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"
SHEET1_DATA_START_ROW = 6
SHEET_EVAL_START_ROW = 7
GRADE_KEYS = ("TOP", "MID", "LOW", "BOTTOM")

# 1. 현장별 인원현황 — 데이터 열 (템플릿 6행 기준)
COL_SITE_CODE = 2
COL_SEQ = 3
COL_SITE_NAME = 4
COL_NAME = 5
COL_AGE = 6
COL_RRN = 7
COL_POSITION = 9
COL_JOB = 10
COL_PHONE = 12
COL_GRADE_FUNCTIONAL = 17
COL_GRADE_SAFETY = 18


def resolve_site_grade_template_path() -> Path:
    bundled = sorted(MODULE_TEMPLATE_DIR.glob(TEMPLATE_GLOB))
    if bundled:
        return bundled[0]
    docs_dir = BASE_DIR / "docs"
    matches = sorted(docs_dir.glob(TEMPLATE_GLOB))
    if not matches:
        raise FileNotFoundError(
            f"기능인등급 템플릿을 찾을 수 없습니다: {MODULE_TEMPLATE_DIR / TEMPLATE_GLOB} 또는 {docs_dir / TEMPLATE_GLOB}"
        )
    return matches[0]


def site_grade_export_filename(for_day: date | None = None) -> str:
    d = for_day or date.today()
    return f"현장별 기능인등급-{d:%Y%m%d}.xlsx"


def _find_sheet_name(wb: openpyxl.Workbook, needle: str) -> str:
    for name in wb.sheetnames:
        if needle in name:
            return name
    raise KeyError(f"시트를 찾을 수 없습니다: {needle!r} (sheets={wb.sheetnames})")


def _grade_code(assessment: dict[str, Any] | None) -> str | None:
    if not assessment or not assessment.get("is_complete"):
        return None
    from app.modules.functional_eval.eval_catalog import normalize_grade_code

    code = str(assessment.get("grade_code") or "").strip()
    return normalize_grade_code(code)


def _sheet1_age_or_birth(worker: dict[str, Any]) -> str:
    """2-2/2-1 VLOOKUP(…,4) → 1.인원현황 F열(생년월일·연령 표시)."""
    age = (worker.get("age_label") or "").strip()
    if age:
        return age
    rrn = (worker.get("rrn_masked") or "").replace("-", "").strip()
    if len(rrn) >= 6 and rrn[:6].isdigit():
        yy, mm, dd = rrn[0:2], rrn[2:4], rrn[4:6]
        century = "19" if int(yy) >= 30 else "20"
        return f"{century}{yy}.{mm}.{dd}"
    return ""


def _clear_sheet1_data(ws: Worksheet) -> None:
    max_row = ws.max_row or SHEET1_DATA_START_ROW
    for row in range(SHEET1_DATA_START_ROW, max_row + 1):
        for col in range(COL_SITE_CODE, COL_GRADE_SAFETY + 1):
            ws.cell(row=row, column=col, value=None)


def _write_sheet1_row(ws: Worksheet, row: int, seq: int, worker: dict[str, Any]) -> None:
    ws.cell(row=row, column=COL_SITE_CODE, value=worker.get("site_code") or "")
    ws.cell(row=row, column=COL_SEQ, value=seq)
    ws.cell(row=row, column=COL_SITE_NAME, value=worker.get("site_name") or "")
    ws.cell(row=row, column=COL_NAME, value=worker.get("name") or "")
    ws.cell(row=row, column=COL_AGE, value=_sheet1_age_or_birth(worker))
    ws.cell(row=row, column=COL_RRN, value=worker.get("rrn_masked") or "")
    ws.cell(row=row, column=COL_POSITION, value=worker.get("position_name") or "")
    ws.cell(row=row, column=COL_JOB, value=worker.get("job_name") or "")
    ws.cell(row=row, column=COL_PHONE, value=worker.get("phone_mobile") or "")
    ws.cell(row=row, column=COL_GRADE_FUNCTIONAL, value=_grade_code(worker.get("functional_assessment")))
    ws.cell(row=row, column=COL_GRADE_SAFETY, value=_grade_code(worker.get("safety_assessment")))


def _criteria_col_count(eval_type: Literal["FUNCTIONAL", "SAFETY"]) -> int:
    return len(get_criteria(eval_type))


def _eval_mark_end_column(ws: Worksheet, template_row: int, eval_type: Literal["FUNCTIONAL", "SAFETY"]) -> int:
    """평가 O 표시 열 끝 — 카탈로그 항목 수와 템플릿 샘플 행 중 넓은 쪽."""
    end = 8 + _criteria_col_count(eval_type) * 4 - 1
    max_col = ws.max_column or end
    for col in range(8, max_col + 1):
        if ws.cell(template_row, col).value == "O":
            end = max(end, col + 3)
    return end


def _clear_eval_marks(
    ws: Worksheet,
    row: int,
    eval_type: Literal["FUNCTIONAL", "SAFETY"],
    *,
    end_col: int | None = None,
) -> None:
    last_col = end_col if end_col is not None else 8 + _criteria_col_count(eval_type) * 4 - 1
    for col in range(8, last_col + 1):
        # openpyxl: cell(..., value=None)은 기존 값을 지우지 않음
        ws.cell(row=row, column=col).value = None


def _write_eval_marks(
    ws: Worksheet,
    row: int,
    assessment: dict[str, Any] | None,
    eval_type: Literal["FUNCTIONAL", "SAFETY"],
) -> None:
    criteria = get_criteria(eval_type)
    scores = (assessment or {}).get("scores") or {}
    complete = bool((assessment or {}).get("is_complete"))
    for i, crit in enumerate(criteria):
        base = 8 + i * 4
        grade_key = scores.get(crit["id"]) if complete else None
        for off, key in enumerate(GRADE_KEYS):
            ws.cell(row=row, column=base + off).value = "O" if grade_key == key else None


def _copy_eval_formula_row(ws: Worksheet, template_row: int, target_row: int) -> None:
    """2-1 / 2-2 행의 VLOOKUP·합계 수식을 템플릿 7행에서 행 단위로 복사한다."""
    if target_row == template_row:
        return
    max_col = ws.max_column or 80
    for col in range(1, max_col + 1):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.data_type == "f" and src.value:
            cell_origin = f"{src.column_letter}{template_row}"
            cell_target = f"{dst.column_letter}{target_row}"
            try:
                dst.value = Translator(str(src.value), origin=cell_origin).translate_formula(cell_target)
            except ValueError:
                # 구형/비표준 수식 — 행 참조만 치환 (VLOOKUP 열 인덱스 보호)
                dst.value = re.sub(
                    rf"\$?{re.escape(src.column_letter)}{template_row}\b",
                    f"{src.column_letter}{target_row}",
                    str(src.value),
                )
        elif col <= 7 and src.data_type != "f":
            dst.value = None


def generate_site_grade_workbook_bytes(workers: list[dict[str, Any]]) -> bytes:
    """템플릿 복사 → 1.인원현황 · 2-1 · 2-2 채움 → xlsx bytes."""
    template = resolve_site_grade_template_path()
    sorted_workers = sorted(
        workers,
        key=lambda w: (
            str(w.get("site_code") or ""),
            int(w.get("row_no") or 0),
            int(w.get("id") or 0),
        ),
    )

    with tempfile.TemporaryDirectory() as tmp:
        out_path = Path(tmp) / "export.xlsx"
        shutil.copy2(template, out_path)
        wb = openpyxl.load_workbook(out_path)

        sheet1_name = _find_sheet_name(wb, "1.")
        sheet_fn_name = _find_sheet_name(wb, "2-1")
        sheet_sf_name = _find_sheet_name(wb, "2-2")
        ws1 = wb[sheet1_name]
        ws_fn = wb[sheet_fn_name]
        ws_sf = wb[sheet_sf_name]

        _clear_sheet1_data(ws1)
        n = len(sorted_workers)
        for idx, worker in enumerate(sorted_workers):
            row1 = SHEET1_DATA_START_ROW + idx
            _write_sheet1_row(ws1, row1, idx + 1, worker)

        fn_cols = _criteria_col_count("FUNCTIONAL")
        sf_cols = _criteria_col_count("SAFETY")
        template_eval_row = SHEET_EVAL_START_ROW
        max_template_row = ws_fn.max_row or template_eval_row

        fn_mark_end = _eval_mark_end_column(ws_fn, template_eval_row, "FUNCTIONAL")
        sf_mark_end = _eval_mark_end_column(ws_sf, template_eval_row, "SAFETY")

        for idx in range(n):
            eval_row = SHEET_EVAL_START_ROW + idx
            if eval_row != template_eval_row:
                _copy_eval_formula_row(ws_fn, template_eval_row, eval_row)
                _copy_eval_formula_row(ws_sf, template_eval_row, eval_row)
            worker = sorted_workers[idx]
            _clear_eval_marks(ws_fn, eval_row, "FUNCTIONAL", end_col=fn_mark_end)
            _clear_eval_marks(ws_sf, eval_row, "SAFETY", end_col=sf_mark_end)
            _write_eval_marks(ws_fn, eval_row, worker.get("functional_assessment"), "FUNCTIONAL")
            _write_eval_marks(ws_sf, eval_row, worker.get("safety_assessment"), "SAFETY")

        for row in range(SHEET_EVAL_START_ROW + n, max_template_row + 1):
            _clear_eval_marks(ws_fn, row, "FUNCTIONAL", end_col=fn_mark_end)
            _clear_eval_marks(ws_sf, row, "SAFETY", end_col=sf_mark_end)

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        buf.seek(0)
        return buf.getvalue()
