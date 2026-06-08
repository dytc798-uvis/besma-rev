from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import openpyxl

RosterDiffType = Literal["NEW", "UPDATED", "UNCHANGED", "REMOVED"]


@dataclass
class ParsedRosterRow:
    name: str
    site_code: str
    rrn_raw: str
    rrn_hash: str
    rrn_masked: str | None
    job_code: str | None
    is_site_manager: bool
    phone: str | None = None


@dataclass
class RosterDiffItem:
    type: RosterDiffType
    rrn_hash: str
    name: str
    site_code: str
    worker_id: int | None = None
    changes: dict[str, tuple[Any, Any]] | None = None


@dataclass
class RosterDiffResult:
    items: list[RosterDiffItem] = field(default_factory=list)

    @property
    def new_count(self) -> int:
        return sum(1 for i in self.items if i.type == "NEW")

    @property
    def updated_count(self) -> int:
        return sum(1 for i in self.items if i.type == "UPDATED")

    @property
    def unchanged_count(self) -> int:
        return sum(1 for i in self.items if i.type == "UNCHANGED")

    @property
    def removed_count(self) -> int:
        return sum(1 for i in self.items if i.type == "REMOVED")


def hash_rrn(rrn_raw: str) -> str:
    return hashlib.sha256(rrn_raw.encode("utf-8")).hexdigest()


def mask_rrn(rrn_raw: str) -> str | None:
    digits = re.sub(r"\D", "", rrn_raw)
    if len(digits) >= 7:
        return f"{digits[:6]}-{digits[6]}"
    return rrn_raw or None


def normalize_site_code(value: Any) -> str:
    site_code = str(value).strip() if value is not None else ""
    if site_code.endswith(".0"):
        site_code = str(int(float(site_code)))
    return site_code.strip()


def normalize_job_code(value: Any) -> str | None:
    if value is None or value == "":
        return None
    job_code = str(value).strip()
    if job_code.endswith(".0"):
        job_code = str(int(float(job_code)))
    return job_code or None


def _parse_daily_roster_rows(rows: list[tuple[Any, ...]]) -> list[ParsedRosterRow]:
    if not rows:
        raise ValueError("EMPTY_FILE")

    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    name_idx = next((i for i, h in enumerate(headers) if "성" in h and "명" in h), 0)
    rrn_idx = next((i for i, h in enumerate(headers) if "주민" in h), 1)
    job_idx = next((i for i, h in enumerate(headers) if h == "직종" or "직종" in h), 6)
    site_idx = next((i for i, h in enumerate(headers) if "소속현장" in h or "현장코드" in h), 7)
    phone_idx = next((i for i, h in enumerate(headers) if "휴대" in h or "핸드" in h), 10)

    parsed: list[ParsedRosterRow] = []
    seen_hashes: set[str] = set()

    for raw in rows[1:]:
        if not raw or not any(raw):
            continue
        name = str(raw[name_idx]).strip() if raw[name_idx] is not None else ""
        if not name:
            continue
        site_code = normalize_site_code(raw[site_idx])
        if not site_code:
            continue
        rrn_raw = str(raw[rrn_idx]).strip() if raw[rrn_idx] is not None else ""
        digits = re.sub(r"\D", "", rrn_raw)
        if len(digits) < 13:
            continue
        rrn_hash = hash_rrn(digits)
        if rrn_hash in seen_hashes:
            continue
        seen_hashes.add(rrn_hash)
        job_code = normalize_job_code(raw[job_idx])
        phone = str(raw[phone_idx]).strip() if phone_idx < len(raw) and raw[phone_idx] else None
        parsed.append(
            ParsedRosterRow(
                name=name,
                site_code=site_code,
                rrn_raw=digits,
                rrn_hash=rrn_hash,
                rrn_masked=mask_rrn(digits),
                job_code=job_code,
                is_site_manager=job_code == "1",
                phone=phone or None,
            )
        )
    return parsed


def parse_daily_roster_xlsx(file_path: Path) -> list[ParsedRosterRow]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _parse_daily_roster_rows(rows)


@dataclass
class ParsedEmployeeMasterRow:
    """본사/ERP 사원리스트 (현장코드 없음, 이름·주민·아이디)."""

    name: str
    rrn_raw: str
    job_code: str | None
    login_id: str | None


def _combine_rrn_parts(front: Any, back: Any) -> str:
    a = re.sub(r"\D", "", str(front or ""))
    b = re.sub(r"\D", "", str(back or ""))
    digits = a + b
    if len(digits) >= 13:
        return digits[:13]
    if len(a) == 13:
        return a
    return digits


def _is_daily_roster_header(headers: list[str]) -> bool:
    joined = " ".join(headers)
    return "소속현장" in joined or "현장코드" in joined


def parse_employee_master(file_path: Path) -> list[ParsedEmployeeMasterRow]:
    """사원리스트_*.xls — 입사일·부서·아이디 형식 (현장코드 없음)."""
    path = Path(file_path)
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
    else:
        from app.modules.functional_eval.xls_io import iter_sheet_rows

        rows = iter_sheet_rows(path)
    if not rows:
        raise ValueError("EMPTY_FILE")
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    if _is_daily_roster_header(headers):
        return []

    name_idx = next((i for i, h in enumerate(headers) if "성" in h and "명" in h), 0)
    job_idx = next((i for i, h in enumerate(headers) if "직종" in h), 5)
    rrn_idx = next((i for i, h in enumerate(headers) if "주민" in h), 6)
    login_idx = next((i for i, h in enumerate(headers) if h in {"아이디", "ID", "login_id"} or "아이디" in h), 9)

    parsed: list[ParsedEmployeeMasterRow] = []
    for raw in rows[1:]:
        if not raw or not any(raw):
            continue
        name = str(raw[name_idx]).strip() if name_idx < len(raw) and raw[name_idx] is not None else ""
        if not name or not is_person_name_simple(name):
            continue
        back_idx = rrn_idx + 1 if rrn_idx + 1 < len(raw) else rrn_idx
        digits = _combine_rrn_parts(
            raw[rrn_idx] if rrn_idx < len(raw) else "",
            raw[back_idx] if back_idx < len(raw) else "",
        )
        if len(digits) < 6:
            continue
        job_code = normalize_job_code(raw[job_idx] if job_idx < len(raw) else None)
        login_id = (
            str(raw[login_idx]).strip()
            if login_idx < len(raw) and raw[login_idx] not in (None, "")
            else None
        )
        parsed.append(
            ParsedEmployeeMasterRow(
                name=name,
                rrn_raw=digits[:13] if len(digits) >= 13 else digits,
                job_code=job_code,
                login_id=login_id,
            )
        )
    return parsed


def is_person_name_simple(name: str) -> bool:
    text = (name or "").strip().replace(" ", "")
    if not text:
        return False
    if not re.fullmatch(r"[가-힣]{2,4}", text):
        return False
    blocked = {"직영", "외주", "합계", "소계", "미배정", "없음", "공무", "소장", "팀장"}
    return text not in blocked


def parse_daily_roster(file_path: Path) -> list[ParsedRosterRow]:
    """일용직 사원리스트 xls/xlsx (소속현장코드·주민번호 13자리)."""
    path = Path(file_path)
    if path.suffix.lower() == ".xlsx":
        return parse_daily_roster_xlsx(path)
    from app.modules.functional_eval.xls_io import iter_sheet_rows

    rows = iter_sheet_rows(path)
    if rows and not _is_daily_roster_header([str(x).strip() if x is not None else "" for x in rows[0]]):
        return []
    return _parse_daily_roster_rows(rows)
