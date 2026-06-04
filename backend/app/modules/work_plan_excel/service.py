from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

import openpyxl

from app.config.settings import BASE_DIR, settings
from app.schemas.work_plan_forklift import ForkliftWorkPlanInput

SHEET_NAME = "1. 지게차(일반)"
TEMPLATE_GLOB = "*지게차*.xlsx"


def resolve_forklift_template_path() -> Path:
    docs_dir = BASE_DIR / "docs"
    matches = sorted(docs_dir.glob(TEMPLATE_GLOB))
    if not matches:
        raise FileNotFoundError(
            f"작업계획서 템플릿을 찾을 수 없습니다: {docs_dir / TEMPLATE_GLOB}"
        )
    return matches[0]


def _work_plan_output_dir() -> Path:
    out_dir = settings.storage_root / "work_plans" / "forklift"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _set_if_value(ws, coord: str, value: str | int | float | None) -> None:
    if value is None:
        return
    if isinstance(value, str) and not value.strip():
        return
    ws[coord] = value


def _format_document_date(payload: ForkliftWorkPlanInput) -> str | None:
    parts = [
        payload.document_date_year.strip() if payload.document_date_year else "",
        payload.document_date_month.strip() if payload.document_date_month else "",
        payload.document_date_day.strip() if payload.document_date_day else "",
    ]
    if not any(parts):
        return None
    y, m, d = parts
    if y and m and d:
        return f"{y} {m} {d}".replace("  ", " ").strip()
    return " ".join(p for p in parts if p)


def _apply_forklift_values(ws, payload: ForkliftWorkPlanInput) -> None:
    """템플릿 서식·그림은 유지하고 입력 셀만 갱신한다."""
    _set_if_value(ws, "A1", payload.site_name)
    doc_date = _format_document_date(payload)
    if doc_date:
        _set_if_value(ws, "AD6", doc_date)
    _set_if_value(ws, "G6", payload.work_name)
    _set_if_value(ws, "H7", payload.period_start_year)
    _set_if_value(ws, "K7", payload.period_start_month)
    _set_if_value(ws, "M7", payload.period_start_day)
    _set_if_value(ws, "P7", payload.period_end_year)
    _set_if_value(ws, "S7", payload.period_end_month)
    if payload.period_end_day:
        _set_if_value(ws, "U7", payload.period_end_day)
    _set_if_value(ws, "K8", payload.company_name)
    if payload.participants:
        _set_if_value(ws, "AD8", payload.participants)
    _set_if_value(ws, "AD7", payload.work_location)
    _set_if_value(ws, "K11", payload.supervisor_name)
    _set_if_value(ws, "K12", payload.supervisor_phone)
    _set_if_value(ws, "M13", payload.supervisor_license_type)
    _set_if_value(ws, "T13", payload.supervisor_license_no)
    _set_if_value(ws, "K14", payload.signal_name)
    _set_if_value(ws, "K15", payload.signal_phone)
    _set_if_value(ws, "M16", payload.signal_license_type)
    _set_if_value(ws, "T16", payload.signal_license_no)
    _set_if_value(ws, "K17", payload.commander_name)
    _set_if_value(ws, "K18", payload.commander_role)
    _set_if_value(ws, "G21", payload.equipment_type)
    _set_if_value(ws, "U21", payload.equipment_model)
    _set_if_value(ws, "AI21", payload.registration_no)
    _set_if_value(ws, "G22", payload.manufacture_year)
    _set_if_value(ws, "U22", payload.rated_capacity)
    _set_if_value(ws, "AI22", payload.registered_company)
    _set_if_value(ws, "K25", payload.length_mm)
    _set_if_value(ws, "T25", payload.width_mm)
    _set_if_value(ws, "AC25", payload.height_mm)
    _set_if_value(ws, "A28", payload.max_lifting_kg)
    if payload.work_location_plan:
        _set_if_value(ws, "I36", f"작업장소 : {payload.work_location_plan}")
    if payload.work_content_plan:
        _set_if_value(ws, "I37", f"작업내용 : {payload.work_content_plan}")


def generate_forklift_work_plan(payload: ForkliftWorkPlanInput) -> tuple[Path, str]:
    template_path = resolve_forklift_template_path()
    out_dir = _work_plan_output_dir()
    today = date.today().strftime("%Y%m%d")
    filename = f"{today}_작업계획서.xlsx"
    out_path = out_dir / filename

    # 원본 바이너리 복사 후 해당 시트 입력값만 수정 (다른 시트·그림 구조 유지)
    shutil.copy2(template_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise KeyError(f"템플릿에 시트가 없습니다: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    _apply_forklift_values(ws, payload)
    wb.save(out_path)
    wb.close()
    return out_path, filename


def resolve_saved_work_plan(filename: str) -> Path:
    safe_name = Path(filename).name
    if safe_name != filename or ".." in filename:
        raise ValueError("invalid_filename")
    candidate = _work_plan_output_dir() / safe_name
    if not candidate.is_file():
        raise FileNotFoundError(safe_name)
    return candidate
