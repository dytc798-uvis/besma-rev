from __future__ import annotations

import csv
import io
import mimetypes
from datetime import date
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, Response
from openpyxl import load_workbook

from app.config.settings import settings
from app.core.auth import DbDep
from app.core.datetime_utils import utc_now
from app.core.enums import Role
from app.core.permissions import CurrentUserDep
from app.core.upload_processing import process_uploaded_image
from app.modules.documents.models import Document
from app.modules.safety_features import risk_gates as rg
from app.modules.safety_features.models import (
    NonconformityItem,
    NonconformityLedger,
    SafetyEducationMaterial,
    SafetyInspectionComment,
    WorkerVoiceComment,
    WorkerVoiceItem,
    WorkerVoiceLedger,
)
from app.modules.sites.models import Site
from app.modules.users.models import User

router = APIRouter(prefix="/safety-features", tags=["safety-features"])


def _ensure_documents_dir() -> Path:
    d = settings.storage_root / settings.documents_dir_name
    d.mkdir(parents=True, exist_ok=True)
    return d


def _assert_site_access(current_user, site_id: int | None) -> None:
    if current_user.role == Role.SITE and site_id is not None and current_user.site_id != site_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")


def _role_value(current_user) -> str:
    role = getattr(current_user, "role", None)
    return role.value if hasattr(role, "value") else str(role)


_HQ_ROLE_VALUES = {Role.HQ_SAFE.value, Role.HQ_SAFE_ADMIN.value, Role.SUPER_ADMIN.value}


def _require_site_only(current_user) -> None:
    if _role_value(current_user) != Role.SITE.value:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="SITE only")


def _require_hq_only(current_user) -> None:
    if _role_value(current_user) not in _HQ_ROLE_VALUES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="HQ only")


def _ledger_risk_db_site_edit_locked(item: WorkerVoiceItem | NonconformityItem) -> bool:
    return rg.risk_db_hq_status_effective(item) == rg.RISK_DB_HQ_APPROVED


def _sync_hq_checked_mirror(item: WorkerVoiceItem | NonconformityItem) -> None:
    if rg.risk_db_hq_status_effective(item) == rg.RISK_DB_HQ_APPROVED:
        item.hq_checked = True
        item.hq_checked_by_user_id = getattr(item, "risk_db_hq_decided_by_user_id", None)
        item.hq_checked_at = getattr(item, "risk_db_hq_decided_at", None) or utc_now()
    else:
        item.hq_checked = False
        item.hq_checked_by_user_id = None
        item.hq_checked_at = None


def _clear_risk_db_axis_on_site_receipt_reject(item: WorkerVoiceItem | NonconformityItem) -> None:
    item.risk_db_request_status = rg.RISK_DB_REQ_PENDING
    item.risk_db_hq_status = rg.RISK_DB_HQ_PENDING
    item.risk_db_requested_at = None
    item.risk_db_requested_by_user_id = None
    item.risk_db_hq_decided_at = None
    item.risk_db_hq_decided_by_user_id = None
    item.reward_candidate = False
    item.reward_candidate_by_user_id = None
    item.reward_candidate_at = None
    _sync_hq_checked_mirror(item)


def _parse_optional_date(value: str | None) -> date | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid date format") from exc


def _safe_relpath(path: str | None) -> Path | None:
    if not path:
        return None
    resolved = settings.storage_root / path
    if not resolved.exists():
        return None
    return resolved


def _store_resized_image(prefix: str, entity_id: int, content: bytes) -> str:
    try:
        asset = process_uploaded_image(
            content,
            source_name=f"{prefix}_{entity_id}.jpg",
            content_type="image/jpeg",
            generate_pdf=False,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(exc)) from exc
    photo_name = f"{prefix}_{entity_id}_{int(utc_now().timestamp())}{asset.optimized_ext}"
    photo_path = _ensure_documents_dir() / photo_name
    photo_path.write_bytes(asset.optimized_bytes)
    return str(photo_path.relative_to(settings.storage_root))


def _next_worker_voice_row_no(db, ledger_id: int) -> int:
    latest = (
        db.query(WorkerVoiceItem)
        .filter(WorkerVoiceItem.ledger_id == ledger_id)
        .order_by(WorkerVoiceItem.row_no.desc(), WorkerVoiceItem.id.desc())
        .first()
    )
    return (latest.row_no if latest else 0) + 1


def _next_nonconformity_row_no(db, ledger_id: int) -> int:
    latest = (
        db.query(NonconformityItem)
        .filter(NonconformityItem.ledger_id == ledger_id)
        .order_by(NonconformityItem.row_no.desc(), NonconformityItem.id.desc())
        .first()
    )
    return (latest.row_no if latest else 0) + 1


def _get_current_worker_voice_ledger(db, site_id: int | None):
    q = db.query(WorkerVoiceLedger)
    if site_id is None:
        return None
    return (
        q.filter(WorkerVoiceLedger.site_id == site_id)
        .order_by(
            (WorkerVoiceLedger.source_type == "MANUAL").desc(),
            WorkerVoiceLedger.uploaded_at.desc(),
            WorkerVoiceLedger.id.desc(),
        )
        .first()
    )


def _get_or_create_worker_voice_ledger(db, current_user, site_id: int | None):
    ledger = _get_current_worker_voice_ledger(db, site_id)
    if ledger is not None:
        return ledger
    ledger = WorkerVoiceLedger(
        site_id=site_id,
        title="근로자 의견청취 관리대장",
        file_path="",
        file_name="",
        file_size=0,
        uploaded_by_user_id=current_user.id,
        source_type="MANUAL",
        uploaded_at=utc_now(),
    )
    db.add(ledger)
    db.flush()
    return ledger


def _get_current_nonconformity_ledger(db, site_id: int | None):
    q = db.query(NonconformityLedger)
    if site_id is None:
        return None
    return (
        q.filter(NonconformityLedger.site_id == site_id)
        .order_by(
            (NonconformityLedger.source_type == "MANUAL").desc(),
            NonconformityLedger.uploaded_at.desc(),
            NonconformityLedger.id.desc(),
        )
        .first()
    )


def _get_or_create_nonconformity_ledger(db, current_user, site_id: int | None):
    ledger = _get_current_nonconformity_ledger(db, site_id)
    if ledger is not None:
        return ledger
    ledger = NonconformityLedger(
        site_id=site_id,
        title="부적합사항 관리대장",
        file_path="",
        file_name="",
        file_size=0,
        uploaded_by_user_id=current_user.id,
        source_type="MANUAL",
        uploaded_at=utc_now(),
    )
    db.add(ledger)
    db.flush()
    return ledger


def _norm_key_part(value: str | None) -> str:
    return (value or "").strip().lower()


def _worker_voice_import_key(row: dict) -> str:
    return "|".join(
        [
            _norm_key_part(row.get("worker_name")),
            _norm_key_part(row.get("worker_birth_date")),
            _norm_key_part(row.get("worker_phone_number")),
            _norm_key_part(row.get("opinion_text")),
        ]
    )


def _nonconformity_import_key(row: dict) -> str:
    return "|".join(
        [
            _norm_key_part(row.get("issue_text")),
            _norm_key_part(row.get("improvement_action")),
        ]
    )


def _serialize_worker_voice_item(item: WorkerVoiceItem, ledger: WorkerVoiceLedger, comment_map: dict[int, list[dict]] | None = None):
    receipt = rg.receipt_decision_effective(item)
    act = rg.normalize_action_status(item.action_status) or item.action_status
    return {
        "id": item.id,
        "ledger_id": ledger.id,
        "ledger_title": ledger.title,
        "ledger_source_type": getattr(ledger, "source_type", "IMPORT"),
        "row_no": item.row_no,
        "worker_name": item.worker_name,
        "worker_birth_date": item.worker_birth_date,
        "worker_phone_number": item.worker_phone_number,
        "opinion_kind": item.opinion_kind,
        "opinion_text": item.opinion_text,
        "action_before": item.action_before,
        "action_after": item.action_after,
        "action_status": act,
        "action_owner": item.action_owner,
        "before_photo_url": (f"/safety-features/worker-voice/items/{item.id}/before-photo" if item.before_photo_path else None),
        "after_photo_url": (f"/safety-features/worker-voice/items/{item.id}/after-photo" if item.after_photo_path else None),
        "receipt_decision": receipt,
        "site_approved": item.site_approved,
        "site_rejected": item.site_rejected,
        "site_reject_note": item.site_reject_note,
        "site_action_comment": item.site_action_comment,
        "site_comment": item.site_action_comment,
        "hq_review_comment": item.hq_review_comment,
        "hq_comment": item.hq_review_comment,
        "risk_db_request_status": rg.risk_db_request_status_effective(item),
        "risk_db_hq_status": rg.risk_db_hq_status_effective(item),
        "risk_db_requested_at": item.risk_db_requested_at,
        "risk_db_requested_by_user_id": item.risk_db_requested_by_user_id,
        "risk_db_hq_decided_at": item.risk_db_hq_decided_at,
        "risk_db_hq_decided_by_user_id": item.risk_db_hq_decided_by_user_id,
        "ready_for_risk_db": rg.worker_voice_item_ready_for_risk_db(item),
        "hq_checked": item.hq_checked,
        "hq_final_approved": rg.risk_db_hq_status_effective(item) == rg.RISK_DB_HQ_APPROVED,
        "reward_candidate": item.reward_candidate,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
        "comments": (comment_map or {}).get(item.id, []),
    }


def _serialize_nonconformity_item(item: NonconformityItem, ledger: NonconformityLedger | None = None):
    after_photo_url = f"/safety-features/nonconformities/items/{item.id}/after-photo" if (item.after_photo_path or item.improvement_photo_path) else None
    receipt = rg.receipt_decision_effective(item)
    act = rg.normalize_action_status(item.action_status) or item.action_status
    return {
        "id": item.id,
        "ledger_id": item.ledger_id,
        "ledger_title": (ledger.title if ledger else ""),
        "row_no": item.row_no,
        "issue_text": item.issue_text,
        "action_before": item.action_before,
        "action_after": item.improvement_action,
        "action_status": act,
        "action_due_date": item.action_due_date,
        "completed_at": item.improvement_date,
        "action_owner": item.improvement_owner,
        "before_photo_url": (f"/safety-features/nonconformities/items/{item.id}/before-photo" if item.before_photo_path else None),
        "after_photo_url": after_photo_url,
        "improvement_action": item.improvement_action,
        "improvement_date": item.improvement_date,
        "improvement_owner": item.improvement_owner,
        "improvement_photo_url": after_photo_url,
        "receipt_decision": receipt,
        "site_approved": item.site_approved,
        "site_rejected": item.site_rejected,
        "site_reject_note": item.site_reject_note,
        "site_action_comment": item.site_action_comment,
        "site_comment": item.site_action_comment,
        "hq_review_comment": item.hq_review_comment,
        "hq_comment": item.hq_review_comment,
        "risk_db_request_status": rg.risk_db_request_status_effective(item),
        "risk_db_hq_status": rg.risk_db_hq_status_effective(item),
        "risk_db_requested_at": item.risk_db_requested_at,
        "risk_db_requested_by_user_id": item.risk_db_requested_by_user_id,
        "risk_db_hq_decided_at": item.risk_db_hq_decided_at,
        "risk_db_hq_decided_by_user_id": item.risk_db_hq_decided_by_user_id,
        "ready_for_risk_db": rg.nonconformity_item_ready_for_risk_db(item),
        "hq_checked": item.hq_checked,
        "hq_final_approved": rg.risk_db_hq_status_effective(item) == rg.RISK_DB_HQ_APPROVED,
        "reward_candidate": item.reward_candidate,
        "updated_at": item.updated_at,
    }


@router.get("/education")
def list_education_materials(db: DbDep, current_user: CurrentUserDep):
    q = db.query(SafetyEducationMaterial)
    if current_user.role == Role.SITE:
        q = q.filter((SafetyEducationMaterial.site_id.is_(None)) | (SafetyEducationMaterial.site_id == current_user.site_id))
    items = q.order_by(SafetyEducationMaterial.uploaded_at.desc(), SafetyEducationMaterial.id.desc()).all()
    user_ids = {i.uploaded_by_user_id for i in items}
    users = db.query(User).filter(User.id.in_(list(user_ids))).all() if user_ids else []
    name_map = {u.id: u.name for u in users}
    return {
        "items": [
            {
                "id": i.id,
                "title": i.title,
                "site_id": i.site_id,
                "file_name": i.file_name,
                "uploaded_at": i.uploaded_at,
                "uploaded_by_name": name_map.get(i.uploaded_by_user_id),
                "file_url": f"/safety-features/file/education/{i.id}",
            }
            for i in items
        ]
    }


@router.post("/education/upload")
async def upload_education_material(
    db: DbDep,
    current_user: CurrentUserDep,
    title: str = Form(...),
    site_id: int | None = Form(None),
    file: UploadFile = File(...),
):
    if _role_value(current_user) not in {Role.HQ_SAFE.value, Role.HQ_SAFE_ADMIN.value, Role.SUPER_ADMIN.value}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")
    clean_title = (title or "").strip()
    if not clean_title:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="title is required")
    if site_id is not None and db.query(Site.id).filter(Site.id == site_id).first() is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Site not found")
    content = await file.read()
    if len(content) > settings.document_upload_max_bytes:
        raise HTTPException(status_code=status.HTTP_413_CONTENT_TOO_LARGE, detail="file too large")
    ext = Path(file.filename or "upload.bin").suffix or ".bin"
    filename = f"education_{int(utc_now().timestamp())}{ext}"
    stored = _ensure_documents_dir() / filename
    stored.write_bytes(content)
    row = SafetyEducationMaterial(
        title=clean_title,
        site_id=site_id,
        file_path=str(stored.relative_to(settings.storage_root)),
        file_name=file.filename or filename,
        file_size=len(content),
        uploaded_by_user_id=current_user.id,
        uploaded_at=utc_now(),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id}


@router.get("/inspections")
def list_safety_inspections(db: DbDep, current_user: CurrentUserDep):
    q = db.query(Document).filter(Document.document_type == "INSPECTION", Document.file_path.isnot(None))
    if current_user.role == Role.SITE:
        q = q.filter(Document.site_id == current_user.site_id)
    docs = q.order_by(Document.uploaded_at.asc().nullslast(), Document.id.asc()).all()
    doc_ids = [d.id for d in docs]
    comments = (
        db.query(SafetyInspectionComment)
        .filter(SafetyInspectionComment.document_id.in_(doc_ids))
        .order_by(SafetyInspectionComment.created_at.asc(), SafetyInspectionComment.id.asc())
        .all()
        if doc_ids
        else []
    )
    user_ids = {c.created_by_user_id for c in comments}
    users = db.query(User).filter(User.id.in_(list(user_ids))).all() if user_ids else []
    name_map = {u.id: u.name for u in users}
    comment_map: dict[int, list[dict]] = {}
    for c in comments:
        comment_map.setdefault(c.document_id, []).append(
            {
                "id": c.id,
                "body": c.body,
                "created_at": c.created_at,
                "created_by_name": name_map.get(c.created_by_user_id),
            }
        )
    return {
        "items": [
            {
                "document_id": d.id,
                "title": d.title,
                "file_name": d.file_name,
                "uploaded_at": d.uploaded_at,
                "file_url": f"/documents/{d.id}/file",
                "comments": comment_map.get(d.id, []),
            }
            for d in docs
        ]
    }


@router.post("/inspections/{document_id}/comments")
def add_safety_inspection_comment(
    document_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    body: str = Form(...),
):
    doc = db.query(Document).filter(Document.id == document_id, Document.document_type == "INSPECTION").first()
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Inspection document not found")
    _assert_site_access(current_user, doc.site_id)
    text = (body or "").strip()
    if not text:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="body is required")
    comment = SafetyInspectionComment(
        document_id=document_id,
        body=text,
        created_by_user_id=current_user.id,
        created_at=utc_now(),
    )
    db.add(comment)
    db.commit()
    db.refresh(comment)
    return {"id": comment.id}


def _extract_ledger_rows(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    def norm(v: object) -> str:
        if v is None:
            return ""
        return str(v).strip().replace(" ", "").replace("\n", "")

    header_idx = None
    for i, row in enumerate(rows[:30]):
        cells = [norm(c) for c in row]
        line = "|".join(cells)
        if ("부적합" in line or "지적" in line) and ("조치" in line or "담당" in line):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0

    headers = [norm(c) for c in rows[header_idx]]

    def find_col(candidates: list[str]) -> int | None:
        for i, h in enumerate(headers):
            if not h:
                continue
            if any(k in h for k in candidates):
                return i
        return None

    issue_col = find_col(["부적합사항", "부적합", "지적사항", "지적내용"])
    if issue_col is None:
        issue_col = 0
    action_col = find_col(["개선조치", "조치계획및내용", "조치계획", "조치내용", "조치"])
    date_col = find_col(["개선조치일자", "조치일자", "개선일자", "완료일자"])
    owner_col = find_col(["조치담당자", "담당자", "담당"])

    data: list[dict] = []
    for idx, row in enumerate(rows[header_idx + 1 :], start=1):
        issue = str(row[issue_col]).strip() if issue_col < len(row) and row[issue_col] is not None else ""
        if not issue:
            continue
        data.append(
            {
                "row_no": idx,
                "issue_text": issue,
                "improvement_action": (
                    str(row[action_col]).strip() if action_col is not None and action_col < len(row) and row[action_col] is not None else None
                ),
                "improvement_owner": (
                    str(row[owner_col]).strip() if owner_col is not None and owner_col < len(row) and row[owner_col] is not None else None
                ),
                "improvement_date_raw": (
                    row[date_col] if date_col is not None and date_col < len(row) else None
                ),
            }
        )
    return data


def _extract_worker_voice_rows(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    return _extract_worker_voice_rows_from_rows(rows)


def _extract_worker_voice_rows_from_rows(rows: list[tuple | list]) -> list[dict]:
    if not rows:
        return []

    def norm(v: object) -> str:
        if v is None:
            return ""
        return str(v).strip().replace(" ", "").replace("\n", "")

    header_idx = None
    for i, row in enumerate(rows[:30]):
        cells = [norm(c) for c in row]
        non_empty_count = sum(1 for c in cells if c)
        opinion_indices = [idx for idx, c in enumerate(cells) if ("의견" in c or "건의" in c or "제안" in c or c == "내용")]
        name_indices = [idx for idx, c in enumerate(cells) if ("근로자" in c or "성명" in c or "이름" in c)]
        has_split_columns = any(oi != ni for oi in opinion_indices for ni in name_indices)
        if non_empty_count >= 2 and opinion_indices and name_indices and has_split_columns:
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0

    def is_secondary_header(row: tuple | list) -> bool:
        cells = [norm(c) for c in row]
        markers = ["성명", "이름", "내용", "조치", "전화", "공종", "직무", "담당", "생년월일", "휴대전화"]
        return sum(1 for c in cells if any(marker in c for marker in markers)) >= 2

    def is_primary_header(row: tuple | list) -> bool:
        cells = [norm(c) for c in row]
        markers = ["NO", "일자", "현장명", "의견", "회신", "비고", "조치"]
        return sum(1 for c in cells if any(marker in c for marker in markers)) >= 2

    headers = [norm(c) for c in rows[header_idx]]
    data_start_idx = header_idx + 1
    if header_idx > 0 and is_secondary_header(rows[header_idx]) and is_primary_header(rows[header_idx - 1]):
        primary = [norm(c) for c in rows[header_idx - 1]]
        secondary = headers
        merged_headers: list[str] = []
        for idx in range(max(len(primary), len(secondary))):
            top = primary[idx] if idx < len(primary) else ""
            bottom = secondary[idx] if idx < len(secondary) else ""
            if top and bottom and bottom not in top:
                merged_headers.append(f"{top}{bottom}")
            else:
                merged_headers.append(top or bottom)
        headers = merged_headers
    elif header_idx + 1 < len(rows) and is_secondary_header(rows[header_idx + 1]):
        secondary = [norm(c) for c in rows[header_idx + 1]]
        merged_headers: list[str] = []
        for idx in range(max(len(headers), len(secondary))):
            top = headers[idx] if idx < len(headers) else ""
            bottom = secondary[idx] if idx < len(secondary) else ""
            if top and bottom and bottom not in top:
                merged_headers.append(f"{top}{bottom}")
            else:
                merged_headers.append(top or bottom)
        headers = merged_headers
        data_start_idx = header_idx + 2

    def find_col(candidates: list[str]) -> int | None:
        for i, h in enumerate(headers):
            if any(k in h for k in candidates):
                return i
        return None

    name_col = find_col(["근로자", "성명", "이름", "name"])
    birth_col = find_col(["생년월일", "출생", "dob", "birth"])
    phone_col = find_col(["휴대전화", "전화번호", "핸드폰", "연락처", "phone", "mobile"])
    kind_col = find_col(["의견종류", "의견유형", "제안유형", "의견수렴경로", "type"])
    opinion_col = find_col(["근로자의견", "의견내용", "건의내용", "제안내용", "의견", "건의", "제안", "내용"])
    if opinion_col is not None and kind_col is not None and opinion_col == kind_col:
        for i, h in enumerate(headers):
            if any(k in h for k in ["의견내용", "근로자의견", "건의내용", "제안내용"]):
                opinion_col = i
                break
    if opinion_col is None:
        opinion_col = 0

    def normalize_kind(value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        compact = text.replace(" ", "").lower()
        if "아차" in text or "near" in compact:
            return "아차사고"
        if "유해" in text or "위험" in text or "hazard" in compact:
            return "유해위험요인발굴"
        return text

    out: list[dict] = []
    for idx, row in enumerate(rows[data_start_idx:], start=1):
        opinion = str(row[opinion_col]).strip() if opinion_col < len(row) and row[opinion_col] is not None else ""
        if not opinion:
            continue
        worker_name = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] is not None else None
        worker_birth_date = (
            str(row[birth_col]).strip() if birth_col is not None and birth_col < len(row) and row[birth_col] is not None else None
        )
        worker_phone_number = (
            str(row[phone_col]).strip() if phone_col is not None and phone_col < len(row) and row[phone_col] is not None else None
        )
        opinion_kind = normalize_kind(row[kind_col]) if kind_col is not None and kind_col < len(row) else None
        out.append(
            {
                "row_no": idx,
                "worker_name": worker_name,
                "worker_birth_date": worker_birth_date,
                "worker_phone_number": worker_phone_number,
                "opinion_kind": opinion_kind,
                "opinion_text": opinion,
            }
        )
    return out


def _extract_worker_voice_rows_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader]
    if not rows:
        return []
    return _extract_worker_voice_rows_from_rows(rows)


def _is_worker_voice_ledger_shape(rows: list[tuple | list]) -> bool:
    if not rows:
        return False

    def norm(v: object) -> str:
        if v is None:
            return ""
        return str(v).strip().replace(" ", "").replace("\n", "")

    for row in rows[:30]:
        cells = [norm(c) for c in row]
        non_empty_count = sum(1 for c in cells if c)
        opinion_indices = [idx for idx, c in enumerate(cells) if ("의견" in c or "건의" in c or "제안" in c or c == "내용")]
        name_indices = [idx for idx, c in enumerate(cells) if ("근로자" in c or "성명" in c or "이름" in c)]
        has_split_columns = any(oi != ni for oi in opinion_indices for ni in name_indices)
        if non_empty_count >= 2 and opinion_indices and name_indices and has_split_columns:
            return True
    return False


def _is_nonconformity_ledger_shape(content: bytes) -> bool:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return False

    def norm(v: object) -> str:
        if v is None:
            return ""
        return str(v).strip().replace(" ", "").replace("\n", "")

    for row in rows[:30]:
        cells = [norm(c) for c in row]
        line = "|".join(cells)
        if ("부적합" in line or "지적" in line) and ("조치" in line or "담당" in line):
            return True
    return False


@router.post("/worker-voice/upload")
async def upload_worker_voice_ledger(
    db: DbDep,
    current_user: CurrentUserDep,
    title: str | None = Form(None),
    site_id: int | None = Form(None),
    file: UploadFile = File(...),
):
    if current_user.role == Role.SITE:
        site_id = current_user.site_id
    if site_id is not None and db.query(Site.id).filter(Site.id == site_id).first() is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Site not found")
    clean_title = (title or "").strip() or "근로자 의견청취 관리대장"
    content = await file.read()
    if len(content) > settings.document_upload_max_bytes:
        raise HTTPException(status_code=status.HTTP_413_CONTENT_TOO_LARGE, detail="file too large")
    ext = Path(file.filename or "ledger.xlsx").suffix or ".xlsx"
    filename = f"worker_voice_{int(utc_now().timestamp())}{ext}"
    stored = _ensure_documents_dir() / filename
    stored.write_bytes(content)
    import_ledger = WorkerVoiceLedger(
        site_id=site_id,
        title=clean_title,
        file_path=str(stored.relative_to(settings.storage_root)),
        file_name=file.filename or filename,
        file_size=len(content),
        uploaded_by_user_id=current_user.id,
        source_type="IMPORT",
        uploaded_at=utc_now(),
    )
    db.add(import_ledger)
    db.flush()

    manual_ledger = _get_or_create_worker_voice_ledger(db, current_user, site_id)
    manual_ledger.title = clean_title
    manual_ledger.file_path = str(stored.relative_to(settings.storage_root))
    manual_ledger.file_name = file.filename or filename
    manual_ledger.file_size = len(content)
    manual_ledger.uploaded_by_user_id = current_user.id
    manual_ledger.uploaded_at = utc_now()
    db.add(manual_ledger)
    suffix = ext.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        csv_rows = [row for row in csv.reader(io.StringIO(text))]
        if not _is_worker_voice_ledger_shape(csv_rows):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="근로자 의견청취 관리대장 형식 파일만 업로드할 수 있습니다.")
        parsed_rows = _extract_worker_voice_rows_from_rows(csv_rows)
    else:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.worksheets[0]
        excel_rows = list(ws.iter_rows(values_only=True))
        if not _is_worker_voice_ledger_shape(excel_rows):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="근로자 의견청취 관리대장 형식 파일만 업로드할 수 있습니다.")
        parsed_rows = _extract_worker_voice_rows_from_rows(excel_rows)
    if not parsed_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="근로자 의견청취 행을 찾지 못했습니다. 관리대장 형식을 확인해 주세요.")
    existing_items = (
        db.query(WorkerVoiceItem)
        .filter(WorkerVoiceItem.ledger_id == manual_ledger.id)
        .all()
    )
    existing_by_key = {_worker_voice_import_key(
        {
            "worker_name": item.worker_name,
            "worker_birth_date": item.worker_birth_date,
            "worker_phone_number": item.worker_phone_number,
            "opinion_text": item.opinion_text,
        }
    ): item for item in existing_items}

    for row in parsed_rows:
        key = _worker_voice_import_key(row)
        item = existing_by_key.get(key)
        if item is None:
            item = WorkerVoiceItem(
                ledger_id=manual_ledger.id,
                row_no=row["row_no"],
                worker_name=row["worker_name"],
                worker_birth_date=row["worker_birth_date"],
                worker_phone_number=row["worker_phone_number"],
                opinion_kind=row["opinion_kind"],
                opinion_text=row["opinion_text"],
            )
            db.add(item)
            existing_by_key[key] = item
            continue
        item.row_no = row["row_no"]
        item.worker_name = row["worker_name"]
        item.worker_birth_date = row["worker_birth_date"]
        item.worker_phone_number = row["worker_phone_number"]
        item.opinion_kind = row["opinion_kind"]
        item.opinion_text = row["opinion_text"]
        db.add(item)
    db.commit()
    db.refresh(import_ledger)
    return {"id": import_ledger.id}


@router.get("/worker-voice/items")
def list_worker_voice_items(db: DbDep, current_user: CurrentUserDep):
    q = db.query(WorkerVoiceItem, WorkerVoiceLedger).join(WorkerVoiceLedger, WorkerVoiceLedger.id == WorkerVoiceItem.ledger_id)
    if current_user.role == Role.SITE:
        q = q.filter(WorkerVoiceLedger.site_id == current_user.site_id)
    rows = q.order_by(WorkerVoiceLedger.uploaded_at.desc(), WorkerVoiceItem.row_no.asc(), WorkerVoiceItem.id.asc()).all()
    item_ids = [row[0].id for row in rows]
    comments = (
        db.query(WorkerVoiceComment)
        .filter(WorkerVoiceComment.item_id.in_(item_ids))
        .order_by(WorkerVoiceComment.created_at.asc(), WorkerVoiceComment.id.asc())
        .all()
        if item_ids
        else []
    )
    user_ids = {c.created_by_user_id for c in comments}
    users = db.query(User).filter(User.id.in_(list(user_ids))).all() if user_ids else []
    name_map = {u.id: u.name for u in users}
    comment_map: dict[int, list[dict]] = {}
    for c in comments:
        comment_map.setdefault(c.item_id, []).append(
            {"id": c.id, "body": c.body, "created_at": c.created_at, "created_by_name": name_map.get(c.created_by_user_id)}
        )
    return {
        "items": [
            _serialize_worker_voice_item(item, ledger, comment_map)
            for item, ledger in rows
        ]
    }


@router.get("/worker-voice/ledger")
def get_current_worker_voice_ledger_detail(db: DbDep, current_user: CurrentUserDep):
    site_id = current_user.site_id if current_user.role == Role.SITE else None
    ledger = _get_current_worker_voice_ledger(db, site_id)
    if ledger is None:
        return {"ledger": None, "items": [], "imports": []}
    _assert_site_access(current_user, ledger.site_id)
    items = (
        db.query(WorkerVoiceItem)
        .filter(WorkerVoiceItem.ledger_id == ledger.id)
        .order_by(WorkerVoiceItem.row_no.asc(), WorkerVoiceItem.id.asc())
        .all()
    )
    item_ids = [item.id for item in items]
    comments = (
        db.query(WorkerVoiceComment)
        .filter(WorkerVoiceComment.item_id.in_(item_ids))
        .order_by(WorkerVoiceComment.created_at.asc(), WorkerVoiceComment.id.asc())
        .all()
        if item_ids
        else []
    )
    user_ids = {c.created_by_user_id for c in comments}
    users = db.query(User).filter(User.id.in_(list(user_ids))).all() if user_ids else []
    name_map = {u.id: u.name for u in users}
    comment_map: dict[int, list[dict]] = {}
    for c in comments:
        comment_map.setdefault(c.item_id, []).append(
            {"id": c.id, "body": c.body, "created_at": c.created_at, "created_by_name": name_map.get(c.created_by_user_id)}
        )
    imports_q = db.query(WorkerVoiceLedger)
    if current_user.role == Role.SITE:
        imports_q = imports_q.filter(WorkerVoiceLedger.site_id == current_user.site_id)
    imports = (
        imports_q.filter(WorkerVoiceLedger.source_type == "IMPORT")
        .order_by(WorkerVoiceLedger.uploaded_at.desc(), WorkerVoiceLedger.id.desc())
        .limit(10)
        .all()
    )
    return {
        "ledger": {
            "id": ledger.id,
            "title": ledger.title,
            "source_type": getattr(ledger, "source_type", "IMPORT"),
            "uploaded_at": ledger.uploaded_at,
            "file_name": ledger.file_name,
        },
        "items": [_serialize_worker_voice_item(item, ledger, comment_map) for item in items],
        "imports": [
            {
                "id": row.id,
                "title": row.title,
                "uploaded_at": row.uploaded_at,
                "file_name": row.file_name,
            }
            for row in imports
        ],
    }


@router.post("/worker-voice/items")
def create_worker_voice_item(
    db: DbDep,
    current_user: CurrentUserDep,
    worker_name: str | None = Form(None),
    worker_birth_date: str | None = Form(None),
    worker_phone_number: str | None = Form(None),
    opinion_kind: str | None = Form(None),
    opinion_text: str = Form(...),
    action_before: str | None = Form(None),
    action_after: str | None = Form(None),
    action_status: str | None = Form(None),
    action_owner: str | None = Form(None),
):
    if current_user.role != Role.SITE or not current_user.site_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only SITE can add rows")
    clean_opinion = (opinion_text or "").strip()
    if not clean_opinion:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="opinion_text is required")
    ledger = _get_or_create_worker_voice_ledger(db, current_user, current_user.site_id)
    item = WorkerVoiceItem(
        ledger_id=ledger.id,
        row_no=_next_worker_voice_row_no(db, ledger.id),
        worker_name=(worker_name or "").strip() or None,
        worker_birth_date=(worker_birth_date or "").strip() or None,
        worker_phone_number=(worker_phone_number or "").strip() or None,
        opinion_kind=(opinion_kind or "").strip() or None,
        opinion_text=clean_opinion,
        action_before=(action_before or "").strip() or None,
        action_after=(action_after or "").strip() or None,
        action_status=(action_status or "").strip() or None,
        action_owner=(action_owner or "").strip() or None,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id}


@router.post("/worker-voice/items/{item_id}")
def update_worker_voice_item(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    worker_name: str | None = Form(None),
    worker_birth_date: str | None = Form(None),
    worker_phone_number: str | None = Form(None),
    opinion_kind: str | None = Form(None),
    opinion_text: str | None = Form(None),
    action_before: str | None = Form(None),
    action_after: str | None = Form(None),
    action_status: str | None = Form(None),
    action_owner: str | None = Form(None),
):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    if current_user.role != Role.SITE:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only SITE can edit rows")
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot edit row after HQ approved risk DB registration",
        )
    if opinion_text is not None:
        clean_opinion = opinion_text.strip()
        if not clean_opinion:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="opinion_text is required")
        item.opinion_text = clean_opinion
    item.worker_name = (worker_name or "").strip() or None
    item.worker_birth_date = (worker_birth_date or "").strip() or None
    item.worker_phone_number = (worker_phone_number or "").strip() or None
    item.opinion_kind = (opinion_kind or "").strip() or None
    item.action_before = (action_before or "").strip() or None
    item.action_after = (action_after or "").strip() or None
    norm_as = rg.normalize_action_status(action_status)
    item.action_status = norm_as if norm_as is not None else ((action_status or "").strip() or None)
    item.action_owner = (action_owner or "").strip() or None
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/before-photo")
async def upload_worker_voice_item_before_photo(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    file: UploadFile = File(...),
):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    if current_user.role != Role.SITE:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only SITE can upload photos")
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cannot upload photos after HQ approved risk DB registration")
    content = await file.read()
    item.before_photo_path = _store_resized_image("worker_voice_before", item.id, content)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/after-photo")
async def upload_worker_voice_item_after_photo(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    file: UploadFile = File(...),
):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    if current_user.role != Role.SITE:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only SITE can upload photos")
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cannot upload photos after HQ approved risk DB registration")
    content = await file.read()
    item.after_photo_path = _store_resized_image("worker_voice_after", item.id, content)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.get("/worker-voice/items/{item_id}/before-photo")
def view_worker_voice_item_before_photo(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    path = _safe_relpath(item.before_photo_path)
    if path is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Photo not found")
    return FileResponse(path=path, media_type="image/jpeg", filename=path.name)


@router.get("/worker-voice/items/{item_id}/after-photo")
def view_worker_voice_item_after_photo(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    path = _safe_relpath(item.after_photo_path)
    if path is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Photo not found")
    return FileResponse(path=path, media_type="image/jpeg", filename=path.name)


@router.post("/worker-voice/items/{item_id}/site-approve")
@router.post("/worker-voice/items/{item_id}/site-accept")
def site_approve_worker_voice(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot change receipt after HQ approved risk DB registration",
        )
    item.receipt_decision = rg.RECEIPT_ACCEPTED
    item.site_approved = True
    item.site_approved_by_user_id = current_user.id
    item.site_approved_at = utc_now()
    item.site_rejected = False
    item.site_reject_note = None
    item.site_rejected_at = None
    item.site_rejected_by_user_id = None
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/site-reject")
def site_reject_worker_voice(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    reject_note: str | None = Form(None),
):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cannot reject receipt after HQ approved risk DB registration")
    note = (reject_note or "").strip() or None
    item.receipt_decision = rg.RECEIPT_REJECTED
    item.site_rejected = True
    item.site_reject_note = note
    item.site_rejected_at = utc_now()
    item.site_rejected_by_user_id = current_user.id
    item.site_approved = False
    item.site_approved_by_user_id = None
    item.site_approved_at = None
    _clear_risk_db_axis_on_site_receipt_reject(item)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/action-status")
def worker_voice_set_action_status(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    action_status: str = Form(...),
):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cannot change action status after HQ approved risk DB registration")
    norm = rg.normalize_action_status(action_status)
    if norm not in {rg.ACTION_NOT_STARTED, rg.ACTION_IN_PROGRESS, rg.ACTION_COMPLETED}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid action_status")
    item.action_status = norm
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/request-risk-db-registration")
def worker_voice_request_risk_db_registration(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    if rg.receipt_decision_effective(item) != rg.RECEIPT_ACCEPTED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="SITE must accept receipt before risk DB request")
    if getattr(item, "site_rejected", False):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="SITE rejected receipt")
    if rg.risk_db_hq_status_effective(item) == rg.RISK_DB_HQ_APPROVED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Risk DB registration already approved")
    now = utc_now()
    if rg.risk_db_request_status_effective(item) == rg.RISK_DB_REQ_REQUESTED:
        if rg.risk_db_hq_status_effective(item) == rg.RISK_DB_HQ_REJECTED:
            item.risk_db_hq_status = rg.RISK_DB_HQ_PENDING
    else:
        item.risk_db_request_status = rg.RISK_DB_REQ_REQUESTED
        item.risk_db_requested_at = now
        item.risk_db_requested_by_user_id = current_user.id
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/site-action-comment")
def worker_voice_set_site_action_comment(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    comment: str | None = Form(None),
):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    item.site_action_comment = (comment or "").strip() or None
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/hq-review-comment")
def worker_voice_set_hq_review_comment(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    comment: str | None = Form(None),
):
    _require_hq_only(current_user)
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    item.hq_review_comment = (comment or "").strip() or None
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/approve-risk-db-registration")
@router.post("/worker-voice/items/{item_id}/hq-check")
def approve_risk_db_registration_worker_voice(item_id: int, db: DbDep, current_user: CurrentUserDep):
    _require_hq_only(current_user)
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    rg.assert_hq_can_approve_risk_db(item)
    now = utc_now()
    item.risk_db_hq_status = rg.RISK_DB_HQ_APPROVED
    item.risk_db_hq_decided_at = now
    item.risk_db_hq_decided_by_user_id = current_user.id
    _sync_hq_checked_mirror(item)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/reject-risk-db-registration")
def reject_risk_db_registration_worker_voice(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    reject_note: str | None = Form(None),
):
    _require_hq_only(current_user)
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    rg.assert_hq_can_reject_risk_db(item)
    now = utc_now()
    item.risk_db_hq_status = rg.RISK_DB_HQ_REJECTED
    item.risk_db_hq_decided_at = now
    item.risk_db_hq_decided_by_user_id = current_user.id
    note = (reject_note or "").strip()
    if note:
        item.hq_review_comment = note
    _sync_hq_checked_mirror(item)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/reward-candidate")
@router.post("/worker-voice/items/{item_id}/mark-reward-candidate")
def promote_worker_voice_reward_candidate(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    _require_hq_only(current_user)
    if not item.site_approved:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="SITE approval required first")
    if item.site_rejected:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="SITE rejected items cannot be reward candidates")
    if rg.risk_db_hq_status_effective(item) != rg.RISK_DB_HQ_APPROVED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="HQ risk DB registration approval required first")
    item.reward_candidate = True
    item.reward_candidate_by_user_id = current_user.id
    item.reward_candidate_at = utc_now()
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/worker-voice/items/{item_id}/comments")
def add_worker_voice_comment(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    body: str = Form(...),
):
    item = db.query(WorkerVoiceItem).filter(WorkerVoiceItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(WorkerVoiceLedger).filter(WorkerVoiceLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    text = (body or "").strip()
    if not text:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="body is required")
    comment = WorkerVoiceComment(
        item_id=item_id,
        body=text,
        created_by_user_id=current_user.id,
        created_at=utc_now(),
    )
    db.add(comment)
    db.commit()
    db.refresh(comment)
    return {"id": comment.id}


@router.post("/nonconformities/upload")
async def upload_nonconformity_ledger(
    db: DbDep,
    current_user: CurrentUserDep,
    title: str | None = Form(None),
    site_id: int | None = Form(None),
    file: UploadFile = File(...),
):
    if current_user.role == Role.SITE:
        site_id = current_user.site_id
    if site_id is not None and db.query(Site.id).filter(Site.id == site_id).first() is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Site not found")
    clean_title = (title or "").strip() or "부적합사항관리대장"
    content = await file.read()
    if len(content) > settings.document_upload_max_bytes:
        raise HTTPException(status_code=status.HTTP_413_CONTENT_TOO_LARGE, detail="file too large")
    ext = Path(file.filename or "ledger.xlsx").suffix or ".xlsx"
    filename = f"nonconformity_{int(utc_now().timestamp())}{ext}"
    stored = _ensure_documents_dir() / filename
    stored.write_bytes(content)

    import_ledger = NonconformityLedger(
        site_id=site_id,
        title=clean_title,
        file_path=str(stored.relative_to(settings.storage_root)),
        file_name=file.filename or filename,
        file_size=len(content),
        uploaded_by_user_id=current_user.id,
        source_type="IMPORT",
        uploaded_at=utc_now(),
    )
    db.add(import_ledger)
    db.flush()

    manual_ledger = _get_or_create_nonconformity_ledger(db, current_user, site_id)
    manual_ledger.title = clean_title
    manual_ledger.file_path = str(stored.relative_to(settings.storage_root))
    manual_ledger.file_name = file.filename or filename
    manual_ledger.file_size = len(content)
    manual_ledger.uploaded_by_user_id = current_user.id
    manual_ledger.uploaded_at = utc_now()
    db.add(manual_ledger)

    if not _is_nonconformity_ledger_shape(content):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="부적합사항 관리대장 형식 파일만 업로드할 수 있습니다.")
    rows = _extract_ledger_rows(content)
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="부적합사항 행을 찾지 못했습니다. 관리대장 형식을 확인해 주세요.")
    existing_items = (
        db.query(NonconformityItem)
        .filter(NonconformityItem.ledger_id == manual_ledger.id)
        .all()
    )
    existing_by_key = {_nonconformity_import_key(
        {
            "issue_text": item.issue_text,
            "improvement_action": item.improvement_action,
        }
    ): item for item in existing_items}

    for r in rows:
        key = _nonconformity_import_key(r)
        item = existing_by_key.get(key)
        if item is None:
            item = NonconformityItem(
                ledger_id=manual_ledger.id,
                row_no=r["row_no"],
                issue_text=r["issue_text"],
                improvement_action=r["improvement_action"],
                improvement_owner=r["improvement_owner"],
            )
            db.add(item)
            existing_by_key[key] = item
            continue
        item.row_no = r["row_no"]
        item.issue_text = r["issue_text"]
        item.improvement_action = r["improvement_action"]
        item.improvement_owner = r["improvement_owner"]
        db.add(item)
    db.commit()
    db.refresh(import_ledger)
    return {"id": import_ledger.id}


@router.get("/nonconformities")
def list_nonconformity_ledgers(db: DbDep, current_user: CurrentUserDep):
    q = db.query(NonconformityLedger)
    if current_user.role == Role.SITE:
        q = q.filter(NonconformityLedger.site_id == current_user.site_id)
    ledgers = q.order_by(NonconformityLedger.uploaded_at.desc(), NonconformityLedger.id.desc()).all()
    return {
        "items": [
            {
                "id": l.id,
                "title": l.title,
                "source_type": getattr(l, "source_type", "IMPORT"),
                "uploaded_at": l.uploaded_at,
                "file_name": l.file_name,
                "download_url": (f"/safety-features/file/nonconformity/{l.id}" if l.file_path else None),
                "pdf_view_url": f"/safety-features/nonconformities/{l.id}/pdf",
            }
            for l in ledgers
        ]
    }


@router.get("/nonconformities/items")
def list_nonconformity_items(db: DbDep, current_user: CurrentUserDep):
    q = db.query(NonconformityItem, NonconformityLedger).join(
        NonconformityLedger, NonconformityLedger.id == NonconformityItem.ledger_id
    )
    if current_user.role == Role.SITE:
        q = q.filter(NonconformityLedger.site_id == current_user.site_id)
    rows = q.order_by(NonconformityLedger.uploaded_at.desc(), NonconformityItem.row_no.asc(), NonconformityItem.id.asc()).all()
    return {
        "items": [_serialize_nonconformity_item(item, ledger) for item, ledger in rows],
    }


@router.get("/nonconformities/overview/current")
def get_current_nonconformity_ledger_detail(db: DbDep, current_user: CurrentUserDep):
    site_id = current_user.site_id if current_user.role == Role.SITE else None
    ledger = _get_current_nonconformity_ledger(db, site_id)
    if ledger is None:
        return {"ledger": None, "items": [], "imports": []}
    _assert_site_access(current_user, ledger.site_id)
    items = (
        db.query(NonconformityItem)
        .filter(NonconformityItem.ledger_id == ledger.id)
        .order_by(NonconformityItem.row_no.asc(), NonconformityItem.id.asc())
        .all()
    )
    imports_q = db.query(NonconformityLedger)
    if current_user.role == Role.SITE:
        imports_q = imports_q.filter(NonconformityLedger.site_id == current_user.site_id)
    imports = (
        imports_q.filter(NonconformityLedger.source_type == "IMPORT")
        .order_by(NonconformityLedger.uploaded_at.desc(), NonconformityLedger.id.desc())
        .limit(10)
        .all()
    )
    return {
        "ledger": {
            "id": ledger.id,
            "title": ledger.title,
            "source_type": getattr(ledger, "source_type", "IMPORT"),
            "uploaded_at": ledger.uploaded_at,
            "file_name": ledger.file_name,
        },
        "items": [_serialize_nonconformity_item(i, ledger) for i in items],
        "imports": [
            {
                "id": row.id,
                "title": row.title,
                "uploaded_at": row.uploaded_at,
                "file_name": row.file_name,
                "pdf_view_url": f"/safety-features/nonconformities/{row.id}/pdf",
                "download_url": f"/safety-features/file/nonconformity/{row.id}",
            }
            for row in imports
        ],
    }


@router.get("/nonconformities/{ledger_id}")
def get_nonconformity_ledger_detail(ledger_id: int, db: DbDep, current_user: CurrentUserDep):
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == ledger_id).first()
    if ledger is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ledger not found")
    _assert_site_access(current_user, ledger.site_id)
    items = (
        db.query(NonconformityItem)
        .filter(NonconformityItem.ledger_id == ledger_id)
        .order_by(NonconformityItem.row_no.asc(), NonconformityItem.id.asc())
        .all()
    )
    return {
        "ledger": {
            "id": ledger.id,
            "title": ledger.title,
            "source_type": getattr(ledger, "source_type", "IMPORT"),
            "uploaded_at": ledger.uploaded_at,
            "file_name": ledger.file_name,
        },
        "items": [_serialize_nonconformity_item(i, ledger) for i in items],
    }


@router.post("/nonconformities/items")
def create_nonconformity_item(
    db: DbDep,
    current_user: CurrentUserDep,
    issue_text: str = Form(...),
    action_before: str | None = Form(None),
    action_after: str | None = Form(None),
    action_status: str | None = Form(None),
    action_due_date: str | None = Form(None),
    completed_at: str | None = Form(None),
    action_owner: str | None = Form(None),
):
    if current_user.role != Role.SITE or not current_user.site_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only SITE can add rows")
    clean_issue = (issue_text or "").strip()
    if not clean_issue:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="issue_text is required")
    ledger = _get_or_create_nonconformity_ledger(db, current_user, current_user.site_id)
    item = NonconformityItem(
        ledger_id=ledger.id,
        row_no=_next_nonconformity_row_no(db, ledger.id),
        issue_text=clean_issue,
        action_before=(action_before or "").strip() or None,
        improvement_action=(action_after or "").strip() or None,
        action_status=(action_status or "").strip() or None,
        action_due_date=_parse_optional_date(action_due_date),
        improvement_date=_parse_optional_date(completed_at),
        improvement_owner=(action_owner or "").strip() or None,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id}


@router.post("/nonconformities/items/{item_id}")
def update_nonconformity_item(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    issue_text: str | None = Form(None),
    action_before: str | None = Form(None),
    action_after: str | None = Form(None),
    action_status: str | None = Form(None),
    action_due_date: str | None = Form(None),
    completed_at: str | None = Form(None),
    action_owner: str | None = Form(None),
):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    if current_user.role != Role.SITE:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only SITE can edit rows")
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot edit row after HQ approved risk DB registration",
        )
    if issue_text is not None:
        clean_issue = issue_text.strip()
        if not clean_issue:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="issue_text is required")
        item.issue_text = clean_issue
    item.action_before = (action_before or "").strip() or None
    item.improvement_action = (action_after or "").strip() or None
    norm_as = rg.normalize_action_status(action_status)
    item.action_status = norm_as if norm_as is not None else ((action_status or "").strip() or None)
    item.action_due_date = _parse_optional_date(action_due_date)
    item.improvement_date = _parse_optional_date(completed_at)
    item.improvement_owner = (action_owner or "").strip() or None
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/site-approve")
@router.post("/nonconformities/items/{item_id}/site-accept")
def site_approve_nonconformity(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Cannot change receipt after HQ approved risk DB registration",
        )
    item.receipt_decision = rg.RECEIPT_ACCEPTED
    item.site_approved = True
    item.site_approved_by_user_id = current_user.id
    item.site_approved_at = utc_now()
    item.site_rejected = False
    item.site_reject_note = None
    item.site_rejected_at = None
    item.site_rejected_by_user_id = None
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/site-reject")
def site_reject_nonconformity(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    reject_note: str | None = Form(None),
):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cannot reject receipt after HQ approved risk DB registration")
    note = (reject_note or "").strip() or None
    item.receipt_decision = rg.RECEIPT_REJECTED
    item.site_rejected = True
    item.site_reject_note = note
    item.site_rejected_at = utc_now()
    item.site_rejected_by_user_id = current_user.id
    item.site_approved = False
    item.site_approved_by_user_id = None
    item.site_approved_at = None
    _clear_risk_db_axis_on_site_receipt_reject(item)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/action-status")
def nonconformity_set_action_status(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    action_status: str = Form(...),
):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cannot change action status after HQ approved risk DB registration")
    norm = rg.normalize_action_status(action_status)
    if norm not in {rg.ACTION_NOT_STARTED, rg.ACTION_IN_PROGRESS, rg.ACTION_COMPLETED}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid action_status")
    item.action_status = norm
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/request-risk-db-registration")
def nonconformity_request_risk_db_registration(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    if rg.receipt_decision_effective(item) != rg.RECEIPT_ACCEPTED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="SITE must accept receipt before risk DB request")
    if getattr(item, "site_rejected", False):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="SITE rejected receipt")
    if rg.risk_db_hq_status_effective(item) == rg.RISK_DB_HQ_APPROVED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Risk DB registration already approved")
    now = utc_now()
    if rg.risk_db_request_status_effective(item) == rg.RISK_DB_REQ_REQUESTED:
        if rg.risk_db_hq_status_effective(item) == rg.RISK_DB_HQ_REJECTED:
            item.risk_db_hq_status = rg.RISK_DB_HQ_PENDING
    else:
        item.risk_db_request_status = rg.RISK_DB_REQ_REQUESTED
        item.risk_db_requested_at = now
        item.risk_db_requested_by_user_id = current_user.id
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/site-action-comment")
def nonconformity_set_site_action_comment(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    comment: str | None = Form(None),
):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    _require_site_only(current_user)
    item.site_action_comment = (comment or "").strip() or None
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/hq-review-comment")
def nonconformity_set_hq_review_comment(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    comment: str | None = Form(None),
):
    _require_hq_only(current_user)
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    item.hq_review_comment = (comment or "").strip() or None
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/approve-risk-db-registration")
@router.post("/nonconformities/items/{item_id}/hq-check")
def approve_risk_db_registration_nonconformity(item_id: int, db: DbDep, current_user: CurrentUserDep):
    _require_hq_only(current_user)
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    rg.assert_hq_can_approve_risk_db(item)
    now = utc_now()
    item.risk_db_hq_status = rg.RISK_DB_HQ_APPROVED
    item.risk_db_hq_decided_at = now
    item.risk_db_hq_decided_by_user_id = current_user.id
    _sync_hq_checked_mirror(item)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/reject-risk-db-registration")
def reject_risk_db_registration_nonconformity(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    reject_note: str | None = Form(None),
):
    _require_hq_only(current_user)
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    rg.assert_hq_can_reject_risk_db(item)
    now = utc_now()
    item.risk_db_hq_status = rg.RISK_DB_HQ_REJECTED
    item.risk_db_hq_decided_at = now
    item.risk_db_hq_decided_by_user_id = current_user.id
    note = (reject_note or "").strip()
    if note:
        item.hq_review_comment = note
    _sync_hq_checked_mirror(item)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/reward-candidate")
@router.post("/nonconformities/items/{item_id}/mark-reward-candidate")
def promote_nonconformity_reward_candidate(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    _require_hq_only(current_user)
    if not item.site_approved:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="SITE approval required first")
    if item.site_rejected:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="SITE rejected items cannot be reward candidates")
    if rg.risk_db_hq_status_effective(item) != rg.RISK_DB_HQ_APPROVED:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="HQ risk DB registration approval required first")
    item.reward_candidate = True
    item.reward_candidate_by_user_id = current_user.id
    item.reward_candidate_at = utc_now()
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/before-photo")
async def upload_nonconformity_item_before_photo(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    file: UploadFile = File(...),
):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    if current_user.role != Role.SITE:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only SITE can upload photos")
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cannot upload photos after HQ approved risk DB registration")
    content = await file.read()
    item.before_photo_path = _store_resized_image("nonconf_before", item.id, content)
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/after-photo")
async def upload_nonconformity_item_after_photo(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    file: UploadFile = File(...),
):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    if current_user.role != Role.SITE:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only SITE can upload photos")
    if _ledger_risk_db_site_edit_locked(item):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Cannot upload photos after HQ approved risk DB registration")
    content = await file.read()
    stored = _store_resized_image("nonconf_after", item.id, content)
    item.after_photo_path = stored
    item.improvement_photo_path = stored
    db.add(item)
    db.commit()
    return {"ok": True}


@router.post("/nonconformities/items/{item_id}/photo")
async def upload_nonconformity_item_photo(
    item_id: int,
    db: DbDep,
    current_user: CurrentUserDep,
    file: UploadFile = File(...),
):
    return await upload_nonconformity_item_after_photo(item_id, db, current_user, file)


@router.get("/nonconformities/items/{item_id}/before-photo")
def view_nonconformity_item_before_photo(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Photo not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    path = _safe_relpath(item.before_photo_path)
    if path is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Photo not found")
    return FileResponse(path=path, media_type="image/jpeg", filename=path.name)


@router.get("/nonconformities/items/{item_id}/after-photo")
def view_nonconformity_item_after_photo(item_id: int, db: DbDep, current_user: CurrentUserDep):
    item = db.query(NonconformityItem).filter(NonconformityItem.id == item_id).first()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Photo not found")
    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == item.ledger_id).first()
    _assert_site_access(current_user, ledger.site_id if ledger else None)
    path = _safe_relpath(item.after_photo_path or item.improvement_photo_path)
    if path is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Photo not found")
    return FileResponse(path=path, media_type="image/jpeg", filename=path.name)


@router.get("/nonconformities/items/{item_id}/photo")
def view_nonconformity_item_photo(item_id: int, db: DbDep, current_user: CurrentUserDep):
    return view_nonconformity_item_after_photo(item_id, db, current_user)


@router.get("/nonconformities/{ledger_id}/pdf")
def render_nonconformity_ledger_pdf(ledger_id: int, db: DbDep, current_user: CurrentUserDep):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfgen import canvas

    ledger = db.query(NonconformityLedger).filter(NonconformityLedger.id == ledger_id).first()
    if ledger is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ledger not found")
    _assert_site_access(current_user, ledger.site_id)
    items = (
        db.query(NonconformityItem)
        .filter(NonconformityItem.ledger_id == ledger_id)
        .order_by(NonconformityItem.row_no.asc(), NonconformityItem.id.asc())
        .all()
    )
    pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
    buff = io.BytesIO()
    p = canvas.Canvas(buff, pagesize=A4)
    p.setFont("HYGothic-Medium", 12)
    y = 800
    p.drawString(40, y, f"부적합사항관리대장: {ledger.title}")
    y -= 24
    for idx, i in enumerate(items, start=1):
        line = (
            f"{idx}. {i.issue_text} / 전:{i.action_before or '-'} / 후:{i.improvement_action or '-'} "
            f"/ 상태:{i.action_status or '-'} / 담당:{i.improvement_owner or '-'} / 완료:{i.improvement_date or '-'}"
        )
        p.drawString(40, y, line[:110])
        y -= 18
        if y < 60:
            p.showPage()
            p.setFont("HYGothic-Medium", 12)
            y = 800
    p.save()
    return Response(
        content=buff.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{quote((ledger.title or 'nonconformity') + '.pdf')}"},
    )


@router.get("/file/{file_type}/{entity_id}")
def download_feature_file(file_type: str, entity_id: int, db: DbDep, current_user: CurrentUserDep):
    if file_type == "education":
        row = db.query(SafetyEducationMaterial).filter(SafetyEducationMaterial.id == entity_id).first()
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
        _assert_site_access(current_user, row.site_id)
        file_path = settings.storage_root / row.file_path
        if not file_path.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
        media_type = mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
        resp = FileResponse(path=file_path, media_type=media_type, filename=row.file_name)
        resp.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(row.file_name)}"
        return resp
    if file_type == "nonconformity":
        row = db.query(NonconformityLedger).filter(NonconformityLedger.id == entity_id).first()
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
        _assert_site_access(current_user, row.site_id)
        file_path = settings.storage_root / row.file_path
        if not file_path.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found")
        media_type = mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
        resp = FileResponse(path=file_path, media_type=media_type, filename=row.file_name)
        resp.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(row.file_name)}"
        return resp
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unknown file type")
