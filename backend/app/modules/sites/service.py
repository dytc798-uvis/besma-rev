from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import re
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config.settings import settings
from app.core.datetime_utils import utc_now
from app.modules.sites.models import Site, SiteImportBatch
from app.modules.users.models import User


@dataclass
class ImportResult:
    total_rows: int
    created_rows: int
    updated_rows: int
    failed_rows: int
    errors: list[dict[str, Any]]
    batch: SiteImportBatch


def _ensure_site_import_dir() -> Path:
    d = settings.storage_root / "site_imports"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _normalize_str(value: str | None) -> str | None:
    if value is None:
        return None
    v = str(value).strip()
    if v == "" or v == "-":
        return None
    return v


def parse_construction_period(value: str | None) -> tuple[date | None, date | None]:
    if not value:
        return None, None
    text = str(value).strip()
    if not text or text == "-":
        return None, None
    parts = re.split(r"\s*~\s*", text)
    if len(parts) != 2:
        return None, None
    try:
        start = date.fromisoformat(parts[0])
        end = date.fromisoformat(parts[1])
    except ValueError:
        return None, None
    return start, end


def parse_int_from_text(value: str | int | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if text == "" or text == "-":
        return None
    digits = re.findall(r"\d+", text.replace(",", ""))
    if not digits:
        return None
    try:
        return int("".join(digits))
    except ValueError:
        return None


def parse_gross_area(value: str | None) -> tuple[int | None, str | None]:
    if value is None:
        return None, None
    text = str(value).strip()
    if text == "" or text == "-":
        return None, None
    m = re.match(r"([\d,]+)\s*([A-Za-z가-힣]+)?", text)
    if not m:
        return None, None
    num_part = m.group(1)
    unit = m.group(2) or None
    try:
        amount = int(num_part.replace(",", ""))
    except ValueError:
        return None, unit
    return amount, unit


def parse_amount(value: str | int | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if text == "" or text == "-":
        return None
    digits = re.findall(r"\d+", text.replace(",", ""))
    if not digits:
        return None
    try:
        return int("".join(digits))
    except ValueError:
        return None


RAW_MAPPING: dict[str, str] = {
    "현장코드": "site_code",
    "공사명": "site_name",
    "공사기간": "construction_period",
    "구분": "contract_type",
    "계약일자": "contract_date",
    "발주처명": "client_name",
    "도급사명": "contractor_name",
    "도급금액": "project_amount",
    "전화번호": "phone_number",
    "주소": "address",
    "동수": "building_count",
    "지하층": "floor_underground",
    "지상층": "floor_ground",
    "세대수": "household_count",
    "연면적": "gross_area_raw",
    "주용도": "main_usage",
    "공종": "work_types",
    "소장": "project_manager",
    "공무": "site_manager",
    "기타": "notes",
}


def upsert_site_from_row(db: Session, row: dict[str, Any], user: User) -> tuple[Site, bool]:
    site_code = _normalize_str(row.get("site_code"))
    site_name = _normalize_str(row.get("site_name"))
    if not site_code or not site_name:
        raise ValueError("site_code and site_name are required")

    existing = db.query(Site).filter(Site.site_code == site_code).first()

    construction_period = _normalize_str(row.get("construction_period"))
    start_date, end_date = parse_construction_period(construction_period)

    gross_area_raw = _normalize_str(row.get("gross_area_raw"))
    gross_area, gross_area_unit = parse_gross_area(gross_area_raw)

    project_amount = parse_amount(row.get("project_amount"))

    attrs: dict[str, Any] = {
        "site_code": site_code,
        "site_name": site_name,
        "start_date": start_date,
        "end_date": end_date,
        "contract_type": _normalize_str(row.get("contract_type")),
        "contract_date": None,
        "client_name": _normalize_str(row.get("client_name")),
        "contractor_name": _normalize_str(row.get("contractor_name")),
        "project_amount": project_amount,
        "phone_number": _normalize_str(row.get("phone_number")),
        "address": _normalize_str(row.get("address")),
        "building_count": parse_int_from_text(row.get("building_count")),
        "floor_underground": parse_int_from_text(row.get("floor_underground")),
        "floor_ground": parse_int_from_text(row.get("floor_ground")),
        "household_count": parse_int_from_text(row.get("household_count")),
        "gross_area": gross_area,
        "gross_area_unit": gross_area_unit,
        "main_usage": _normalize_str(row.get("main_usage")),
        "work_types": _normalize_str(row.get("work_types")),
        "project_manager": _normalize_str(row.get("project_manager")),
        "site_manager": _normalize_str(row.get("site_manager")),
        "notes": _normalize_str(row.get("notes")),
    }

    # 계약일자는 ISO 포맷 또는 YYYY-MM-DD 가정
    raw_contract_date = _normalize_str(row.get("contract_date"))
    if raw_contract_date:
        try:
            attrs["contract_date"] = date.fromisoformat(raw_contract_date)
        except ValueError:
            attrs["contract_date"] = None

    now = utc_now()

    if existing:
        for k, v in attrs.items():
            setattr(existing, k, v)
        existing.updated_by_user_id = user.id
        existing.updated_at = now
        db.add(existing)
        return existing, False

    site = Site(
        **attrs,
        created_by_user_id=user.id,
        updated_by_user_id=user.id,
        created_at=now,
        updated_at=now,
    )
    db.add(site)
    return site, True


async def import_sites_from_workbook(db: Session, file: UploadFile, user: User) -> ImportResult:
    storage_dir = _ensure_site_import_dir()
    safe_name = file.filename or "sites.xlsx"
    ts = int(utc_now().timestamp())
    stored_path = storage_dir / f"site_import_{ts}_{safe_name}"
    content = await file.read()
    stored_path.write_bytes(content)

    wb = load_workbook(filename=stored_path, data_only=True)
    ws = wb.active

    headers: list[str] = []
    total_rows = 0
    created_rows = 0
    updated_rows = 0
    failed_rows = 0
    errors: list[dict[str, Any]] = []

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            headers = [str(c).strip() if c is not None else "" for c in row]
            continue

        if not any(row):
            continue

        total_rows += 1
        raw: dict[str, Any] = {}
        for col_idx, cell_value in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else ""
            mapped = RAW_MAPPING.get(header)
            if not mapped:
                continue
            raw[mapped] = cell_value

        try:
            site, created = upsert_site_from_row(db, raw, user)
            if created:
                created_rows += 1
            else:
                updated_rows += 1
        except Exception as e:  # noqa: BLE001
            failed_rows += 1
            errors.append({"row_index": idx, "error": str(e)})

    batch = SiteImportBatch(
        original_filename=safe_name,
        stored_path=str(stored_path.relative_to(settings.storage_root)),
        uploaded_by_user_id=user.id,
        uploaded_at=utc_now(),
        total_rows=total_rows,
        created_rows=created_rows,
        updated_rows=updated_rows,
        failed_rows=failed_rows,
        error_summary="; ".join(f"row {e['row_index']}: {e['error']}" for e in errors) or None,
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)

    return ImportResult(
        total_rows=total_rows,
        created_rows=created_rows,
        updated_rows=updated_rows,
        failed_rows=failed_rows,
        errors=errors,
        batch=batch,
    )

