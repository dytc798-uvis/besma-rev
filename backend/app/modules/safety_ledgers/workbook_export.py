from __future__ import annotations

import calendar
import shutil
from io import BytesIO
from copy import copy
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps

from app.modules.safety_ledgers.models import SafetyCardExpense, SafetyVehicle, SafetyVehicleLog


CARD_FILENAME = "안전실_법인카드 정산서.xlsx"
VEHICLE_FILENAME = "안전실_업무용승용차 운행기록부.xlsx"
_BLUE = "1F4E78"
_LIGHT_BLUE = "D9EAF7"
_THIN = Side(style="thin", color="A6A6A6")


def _sheet_title(year: int, month: int) -> str:
    return f"{year % 100:02d}년 {month}월"


def _style_table(ws, start_row: int, end_row: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _longest_span(flags: list[bool]) -> tuple[int, int] | None:
    best: tuple[int, int] | None = None
    start: int | None = None
    for index, enabled in enumerate(flags + [False]):
        if enabled and start is None:
            start = index
        elif not enabled and start is not None:
            if best is None or index - start > best[1] - best[0]:
                best = (start, index)
            start = None
    return best


def _receipt_print_image(path: Path) -> BytesIO:
    with Image.open(path) as opened:
        image = ImageOps.exif_transpose(opened).convert("RGB")
    sample = image.copy()
    sample.thumbnail((500, 700), Image.Resampling.LANCZOS)
    pixels = sample.load()
    width, height = sample.size
    paper = [
        [
            min(pixels[x, y]) >= 185 and max(pixels[x, y]) - min(pixels[x, y]) <= 65
            for x in range(width)
        ]
        for y in range(height)
    ]
    columns = [sum(paper[y][x] for y in range(height)) / height >= 0.35 for x in range(width)]
    x_span = _longest_span(columns)
    if x_span and x_span[1] - x_span[0] >= width * 0.35:
        scale_x = image.width / width
        margin_x = max(8, int((x_span[1] - x_span[0]) * scale_x * 0.025))
        box = (
            max(0, int(x_span[0] * scale_x) - margin_x),
            0,
            min(image.width, int(x_span[1] * scale_x) + margin_x),
            image.height,
        )
        image = image.crop(box)

    canvas = Image.new("RGB", (1400, 1700), "white")
    fitted = image.copy()
    fitted.thumbnail((1300, 1600), Image.Resampling.LANCZOS)
    canvas.paste(fitted, ((canvas.width - fitted.width) // 2, (canvas.height - fitted.height) // 2))
    stream = BytesIO()
    canvas.save(stream, format="JPEG", quality=92, optimize=True)
    stream.seek(0)
    return stream


def _append_receipt_evidence(
    wb,
    expenses: list[SafetyCardExpense],
    receipt_storage_root: Path | None,
) -> list[BytesIO]:
    streams: list[BytesIO] = []
    evidence_rows = []
    for item in sorted(expenses, key=lambda row: (row.used_at or row.created_at, row.id)):
        raw_path = getattr(item, "receipt_image_path", None)
        if not raw_path:
            continue
        path = Path(raw_path)
        if not path.is_absolute() and receipt_storage_root is not None:
            path = receipt_storage_root / path
        if path.is_file():
            evidence_rows.append((item, path))

    for index, (item, path) in enumerate(evidence_rows, 1):
        when = item.used_at or item.created_at
        ws = wb.create_sheet(f"영수증{index:02d}")
        ws.merge_cells("A1:H1")
        ws["A1"] = f"법인카드 영수증 증빙 {index:02d}"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=_BLUE)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A2:H3")
        ws["A2"] = " | ".join(
            part
            for part in [
                f"{when:%Y-%m-%d}",
                item.site_name,
                item.merchant,
                f"{int(item.amount or 0):,}원",
            ]
            if part
        )
        ws["A2"].font = Font(size=11, bold=True, color="1F2937")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in range(1, 9):
            ws.column_dimensions[get_column_letter(column)].width = 12
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 12
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3
        ws.print_area = "A1:H55"
        try:
            stream = _receipt_print_image(path)
        except (OSError, ValueError):
            wb.remove(ws)
            continue
        streams.append(stream)
        receipt = ExcelImage(stream)
        receipt.width = 690
        receipt.height = 838
        ws.add_image(receipt, "A5")
    return streams


def build_card_workbook(
    expenses: Iterable[SafetyCardExpense],
    output_path: Path,
    *,
    template_path: Path | None = None,
    site_names_by_date: dict[date, str] | None = None,
    receipt_storage_root: Path | None = None,
    include_receipt_evidence: bool = False,
) -> Path:
    expense_rows = list(expenses)
    grouped: dict[tuple[int, int], list[SafetyCardExpense]] = defaultdict(list)
    for row in expense_rows:
        when = row.used_at or row.created_at
        grouped[(when.year, when.month)].append(row)
    if not grouped:
        now = datetime.now()
        grouped[(now.year, now.month)] = []

    if template_path and template_path.is_file():
        wb = load_workbook(template_path)
        source = wb["5월"] if "5월" in wb.sheetnames else wb.worksheets[0]
        target_sheets = []
        for year, month in sorted(grouped):
            ws = wb.copy_worksheet(source)
            ws.title = f"__safety_card_{year}_{month}"
            target_sheets.append((year, month, ws))
        for ws in list(wb.worksheets):
            if not ws.title.startswith("__safety_card_"):
                wb.remove(ws)
        for year, month, ws in target_sheets:
            ws.title = f"{month}월"
            ws["A2"] = (
                f"사용기간 : {year}년 {month}월 1일 ~ "
                f"{year}년 {month}월 {calendar.monthrange(year, month)[1]}일"
            )
            ws["G2"] = "사용자 : 안전보건실"
            ws["G46"] = "정산인 : 안전보건실 (인)"
            for row_index in range(4, 45):
                ws.cell(row_index, 1, "1" if row_index == 4 else "=ROW()-3")
                for col_index in range(2, 8):
                    ws.cell(row_index, col_index).value = None
            rows = sorted(
                grouped[(year, month)],
                key=lambda item: (item.used_at or item.created_at, item.id),
            )
            if len(rows) > 41:
                raise ValueError(f"{year}년 {month}월 법인카드 내역이 회사 양식의 41행을 초과했습니다.")
            for row_index, item in enumerate(rows, 4):
                when = item.used_at or item.created_at
                ws.cell(row_index, 2, when.date())
                ws.cell(row_index, 3, item.site_name or (site_names_by_date or {}).get(when.date(), ""))
                ws.cell(row_index, 4, item.merchant or "")
                ws.cell(row_index, 5, item.amount)
                ws.cell(row_index, 6, item.description or "")
                note_parts = [
                    item.note,
                ]
                ws.cell(row_index, 7, " / ".join(part for part in note_parts if part))
            ws["E45"] = "=SUM(E4:E44)"
            ws.print_area = "A1:G46"
        streams = (
            _append_receipt_evidence(wb, expense_rows, receipt_storage_root)
            if include_receipt_evidence
            else []
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        for stream in streams:
            stream.close()
        return output_path

    wb = Workbook()
    wb.remove(wb.active)
    for (year, month), rows in sorted(grouped.items()):
        ws = wb.create_sheet(_sheet_title(year, month))
        ws.merge_cells("A1:G1")
        ws["A1"] = "■ 법인카드 정산서"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=_BLUE)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:E2")
        ws["A2"] = f"사용기간 : {year}년 {month}월 1일 ~ {year}년 {month}월 {calendar.monthrange(year, month)[1]}일"
        ws["G2"] = "부서 : 안전보건실"
        headers = ["No", "사용일시", "현장명", "사용처", "금액", "내용", "비고(카드끝4자리)"]
        for col, value in enumerate(headers, 1):
            cell = ws.cell(3, col, value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_BLUE)
            cell.alignment = Alignment(horizontal="center")
        rows = sorted(rows, key=lambda item: (item.used_at or item.created_at, item.id))
        data_start = 4
        for index, item in enumerate(rows, 1):
            when = item.used_at or item.created_at
            values = [
                index,
                when,
                item.site_name or (site_names_by_date or {}).get(when.date(), ""),
                item.merchant or "",
                item.amount,
                item.description or "",
                " / ".join(x for x in [item.note, f"카드 ****-{item.card_last4}" if item.card_last4 else None] if x),
            ]
            for col, value in enumerate(values, 1):
                ws.cell(data_start + index - 1, col, value)
            ws.cell(data_start + index - 1, 2).number_format = "yyyy-mm-dd hh:mm"
            ws.cell(data_start + index - 1, 5).number_format = '#,##0"원"'
        total_row = max(data_start + len(rows), 5)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        ws.cell(total_row, 1, "소계")
        ws.cell(total_row, 5, f"=SUM(E{data_start}:E{total_row - 1})")
        ws.cell(total_row, 5).number_format = '#,##0"원"'
        ws.cell(total_row, 1).font = ws.cell(total_row, 5).font = Font(bold=True)
        _style_table(ws, 3, total_row, 7)
        widths = [7, 19, 24, 24, 14, 38, 24]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:G{max(3, total_row - 1)}"
        ws.sheet_view.showGridLines = False

    streams = (
        _append_receipt_evidence(wb, expense_rows, receipt_storage_root)
        if include_receipt_evidence
        else []
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    for stream in streams:
        stream.close()
    return output_path


def build_vehicle_workbook(
    vehicle: SafetyVehicle,
    logs: Iterable[SafetyVehicleLog],
    output_path: Path,
    *,
    template_path: Path | None = None,
) -> Path:
    grouped: dict[tuple[int, int], list[SafetyVehicleLog]] = defaultdict(list)
    for row in logs:
        grouped[(row.driven_on.year, row.driven_on.month)].append(row)
    if not grouped:
        now = datetime.now()
        grouped[(now.year, now.month)] = []
    latest_year = max(year for year, _month in grouped)
    latest_month = max(month for year, month in grouped if year == latest_year)
    for month in range(latest_month + 1, 13):
        grouped.setdefault((latest_year, month), [])

    if template_path and template_path.is_file():
        wb = load_workbook(template_path)
        source = wb["7월"] if "7월" in wb.sheetnames else wb.worksheets[0]
        target_sheets = []
        for year, month in sorted(grouped):
            title = f"{month}월"
            if title in wb.sheetnames:
                ws = wb[title]
            else:
                ws = wb.copy_worksheet(source)
                for validation in source.data_validations.dataValidation:
                    ws.add_data_validation(copy(validation))
                ws.title = title
            target_sheets.append((year, month, ws))
        keep = {id(ws) for _year, _month, ws in target_sheets}
        for ws in list(wb.worksheets):
            if id(ws) not in keep:
                wb.remove(ws)

        for year, month, ws in target_sheets:
            rows = sorted(grouped[(year, month)], key=lambda item: (item.driven_on, item.created_at, item.id))
            ws["A7"] = vehicle.vehicle_name
            ws["E7"] = vehicle.plate_number
            ws["G7"] = None
            ws["H7"] = vehicle.ownership_type
            last_day = calendar.monthrange(year, month)[1]
            for day in range(1, 32):
                row_index = 10 + day
                if day <= last_day:
                    ws.cell(row_index, 1, year)
                    ws.cell(row_index, 2, month)
                    ws.cell(row_index, 3, day)
                    ws.cell(row_index, 4, vehicle.department)
                    for col_index in range(5, 9):
                        ws.cell(row_index, col_index).value = None
                else:
                    for col_index in range(1, 9):
                        ws.cell(row_index, col_index).value = None
            rows_by_day: dict[int, list[SafetyVehicleLog]] = defaultdict(list)
            for item in rows:
                rows_by_day[item.driven_on.day].append(item)
            for day, day_rows in rows_by_day.items():
                row_index = 10 + day
                drivers = list(dict.fromkeys(item.driver_name for item in day_rows if item.driver_name))
                use_types = list(
                    dict.fromkeys(
                        item.use_type if str(item.use_type).startswith(tuple("1234567")) else "3.업무용"
                        for item in day_rows
                    )
                )
                purposes = list(dict.fromkeys(item.purpose for item in day_rows if item.purpose))
                ws.cell(row_index, 5, " / ".join(drivers))
                ws.cell(row_index, 6, use_types[0] if len(use_types) == 1 else "3.업무용")
                ws.cell(row_index, 7, sum(float(item.trip_km or 0) for item in day_rows))
                ws.cell(row_index, 8, " / ".join(purposes))
            ws["G42"] = "=SUM(G11:G41)"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    wb = Workbook()
    wb.remove(wb.active)
    for (year, month), rows in sorted(grouped.items()):
        rows = sorted(rows, key=lambda item: (item.driven_on, item.created_at, item.id))
        ws = wb.create_sheet(_sheet_title(year, month))
        ws.merge_cells("A1:H1")
        ws["A1"] = "운행기록부 (업무용승용차)"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor=_BLUE)
        ws["A1"].alignment = Alignment(horizontal="center")
        info = [
            ("A3", "①차종", "B3", vehicle.vehicle_name),
            ("D3", "②차량번호", "E3", vehicle.plate_number),
            ("G3", "③기초km", "H3", None),
            ("A4", "④명의구분", "B4", vehicle.ownership_type),
            ("D4", "부서", "E4", vehicle.department),
            (
                "G4",
                "운전자",
                "H4",
                " · ".join(driver.driver_name for driver in getattr(vehicle, "drivers", []) if driver.is_active),
            ),
        ]
        for label_cell, label, value_cell, value in info:
            ws[label_cell] = label
            ws[label_cell].font = Font(bold=True)
            ws[label_cell].fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
            ws[value_cell] = value
        headers = ["년도", "월", "일", "부서", "성명", "구분", "주행km", "비고"]
        for col, value in enumerate(headers, 1):
            cell = ws.cell(6, col, value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_BLUE)
            cell.alignment = Alignment(horizontal="center")
        for index, item in enumerate(rows, 7):
            values = [
                item.driven_on.year,
                item.driven_on.month,
                item.driven_on.day,
                vehicle.department,
                item.driver_name,
                item.use_type if str(item.use_type).startswith(tuple("1234567")) else "3.업무용",
                item.trip_km,
                item.purpose or "",
            ]
            for col, value in enumerate(values, 1):
                ws.cell(index, col, value)
        total_row = max(7 + len(rows), 8)
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
        ws.cell(total_row, 1, "월 주행거리 합계")
        ws.cell(total_row, 7, f"=SUM(G7:G{total_row - 1})")
        ws.cell(total_row, 1).font = ws.cell(total_row, 7).font = Font(bold=True)
        _style_table(ws, 3, 4, 8)
        _style_table(ws, 6, total_row, 8)
        widths = [9, 7, 7, 16, 13, 19, 12, 42]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A7"
        ws.sheet_view.showGridLines = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def copy_exports_to_nas(paths: Iterable[Path], nas_root: Path | None) -> None:
    if nas_root is None:
        return
    nas_root.mkdir(parents=True, exist_ok=True)
    for source in paths:
        temp = nas_root / f".{source.name}.tmp"
        shutil.copy2(source, temp)
        temp.replace(nas_root / source.name)
