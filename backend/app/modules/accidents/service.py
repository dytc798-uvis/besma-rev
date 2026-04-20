# -*- coding: utf-8 -*-
from __future__ import annotations

from copy import copy
from datetime import datetime
from io import BytesIO
import json
from pathlib import Path
import re
import shutil
import uuid

from fastapi import HTTPException, status
from fastapi.concurrency import run_in_threadpool
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.config.settings import settings
from app.modules.accidents import repository
from app.modules.accidents.models import Accident, AccidentAttachment
from app.modules.accidents.parser import (
    compose_initial_report_view,
    evaluate_parse_fields,
    parse_initial_report_message,
)

STATUS_OPTIONS = ["신규", "보완필요", "접수", "조사중", "치료중", "종결"]
MANAGEMENT_CATEGORY_OPTIONS = ["일반", "유지", "별도관리"]
MASTER_HEADERS = [
    "등록일",
    "수정일",
    "사고ID",
    "사고일시",
    "발생연도",
    "현장명",
    "작업명",
    "성명",
    "생년월일",
    "연령",
    "소속",
    "직종",
    "고용형태",
    "사고장소",
    "사고유형",
    "재해부위",
    "사고개요",
    "직접원인",
    "간접원인",
    "사고상태",
    "치료현황",
    "휴업예상일수",
    "산재표신고여부",
    "산재신청여부",
    "고객사보고일시",
    "본사보고일시",
    "보고자",
    "비고",
    "폴더키",
    "로컬폴더경로",
    "NAS폴더경로",
    "증빙폴더링크",
    "증빙확인여부",
    "원본분석표_시트명",
    "원본분석표_No",
    "원본매칭키",
    "데이터검증상태",
    "출력키",
    "진행현황",
    "관리구분",
]
MASTER_IMPORT_PATH = Path(
    r"Z:\4. 안전보건관리실\104 사고 조사 및 이력 (산재요양승인 내역)\BESMA_사고MASTER_2026.xlsx"
)
MASTER_IMPORT_SHEET = "MASTER_사고한장표"


def normalize_site_name(value: str | None) -> str | None:
    text = (value or "").strip()
    return text or None


def normalize_text(value: str | None) -> str | None:
    text = (value or "").strip()
    return text or None


def compute_is_complete(*, site_standard_name: str | None, accident_datetime: datetime | None, injured_person_name: str | None) -> bool:
    return bool(site_standard_name and accident_datetime and injured_person_name)


def default_verification_status() -> str:
    return "검증완료"


def accident_storage_root(accident_datetime: datetime | None = None) -> Path:
    base = Path(settings.accident_nas_root) if settings.accident_nas_root else Path(settings.storage_root) / "accidents"
    root = base / str((accident_datetime or datetime.now()).year)
    root.mkdir(parents=True, exist_ok=True)
    return root


def _sanitize_folder_part(value: str | None, fallback: str) -> str:
    text = re.sub(r'[<>:"/\\|?*]+', " ", (value or "").strip())
    text = re.sub(r"\s+", " ", text).strip()
    return text or fallback


def build_nas_folder_key(
    *,
    accident_datetime: datetime | None,
    injured_person_name: str | None,
    site_name: str | None,
    fallback_code: str,
) -> str:
    if not accident_datetime:
        return fallback_code
    mmdd = accident_datetime.strftime("%m%d")
    injured = _sanitize_folder_part(injured_person_name, "미상")
    site = _sanitize_folder_part(site_name, "현장미상")
    return f"{mmdd} {injured} {site}"


def accident_nas_dir(folder_key: str, accident_datetime: datetime | None = None) -> Path:
    root = accident_storage_root(accident_datetime)
    folder = root / folder_key
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def resolve_accident_folder(row: Accident) -> Path:
    if row.nas_folder_path:
        folder = Path(row.nas_folder_path)
        folder.mkdir(parents=True, exist_ok=True)
        return folder
    folder_key = row.nas_folder_key or row.accident_id
    folder = accident_nas_dir(folder_key, row.accident_datetime)
    row.nas_folder_path = str(folder)
    return folder


def ensure_accident_folder_persisted(db: Session, row: Accident) -> None:
    """탐색기/업로드와 맞추기 위해 서버 측 사고 폴더를 생성하고 `nas_folder_path`를 DB에 반영한다."""
    resolve_accident_folder(row)
    db.add(row)
    db.commit()
    db.refresh(row)


def _create_accident_row(
    db: Session,
    *,
    created_by_user_id: int | None,
    source_type: str,
    message_raw: str,
    fields: dict,
    parse_status: str,
    parse_note: str | None,
    status: str,
    management_category: str,
    verification_status: str,
    notes: str | None,
) -> Accident:
    year = datetime.now().year
    display_code = repository.next_display_code(db, year=year)
    accident_id = repository.next_accident_id(db, year=year)
    site_standard_name = normalize_site_name(fields.get("site_name"))
    folder_key = build_nas_folder_key(
        accident_datetime=fields.get("accident_datetime"),
        injured_person_name=fields.get("injured_person_name"),
        site_name=site_standard_name or fields.get("site_name"),
        fallback_code=accident_id,
    )
    nas_dir = accident_nas_dir(folder_key, fields.get("accident_datetime"))
    initial_report_template = compose_initial_report_view(fields, message_raw)["composed_line"]
    return repository.create_accident(
        db,
        display_code=display_code,
        accident_id=accident_id,
        source_type=source_type,
        message_raw=message_raw,
        parse_status=parse_status,
        parse_note=parse_note,
        site_name=fields.get("site_name"),
        reporter_name=fields.get("reporter_name"),
        accident_datetime_text=fields.get("accident_datetime_text"),
        accident_datetime=fields.get("accident_datetime"),
        accident_place=fields.get("accident_place"),
        work_content=fields.get("work_content"),
        injured_person_name=fields.get("injured_person_name"),
        accident_circumstance=fields.get("accident_circumstance"),
        accident_reason=fields.get("accident_reason"),
        injured_part=fields.get("injured_part"),
        diagnosis_name=fields.get("diagnosis_name"),
        action_taken=fields.get("action_taken"),
        status=status,
        management_category=management_category,
        verification_status=verification_status,
        site_standard_name=site_standard_name,
        initial_report_template=initial_report_template,
        is_complete=compute_is_complete(
            site_standard_name=site_standard_name,
            accident_datetime=fields.get("accident_datetime"),
            injured_person_name=fields.get("injured_person_name"),
        ),
        nas_folder_path=str(nas_dir),
        nas_folder_key=folder_key,
        notes=notes,
        created_by_user_id=created_by_user_id,
        updated_by_user_id=created_by_user_id,
    )


def parse_and_create(db: Session, *, message_raw: str, created_by_user_id: int | None) -> Accident:
    parsed = parse_initial_report_message(message_raw)
    return _create_accident_row(
        db,
        created_by_user_id=created_by_user_id,
        source_type="naverworks_message",
        message_raw=message_raw,
        fields=parsed["fields"],
        parse_status=parsed["parse_status"],
        parse_note=parsed.get("parse_note"),
        status="신규",
        management_category="일반",
        verification_status=default_verification_status(),
        notes=None,
    )


def preview_parse(message_raw: str) -> dict:
    parsed = parse_initial_report_message(message_raw)
    view = compose_initial_report_view(parsed["fields"], message_raw)
    return {
        "parse_status": parsed["parse_status"],
        "parse_note": parsed.get("parse_note"),
        "composed_line": view["composed_line"],
        "fields": view["fields"],
        "message_raw": message_raw,
    }


def build_manual_supplement_note(
    *,
    fields: dict,
    original_parse_status: str | None,
    original_parse_note: str | None,
) -> str:
    base_note: dict = {"patterns": {"manual_input": True}, "manual_supplemented": True}
    if original_parse_status:
        base_note["original_parse_status"] = original_parse_status
    if original_parse_note:
        try:
            original_note = json.loads(original_parse_note)
            base_note["original_parse_note"] = original_note
            if original_note.get("template_marker_present"):
                base_note["template_marker_present"] = True
                base_note["patterns"]["template_marker"] = True
        except json.JSONDecodeError:
            base_note["original_parse_note_raw"] = original_parse_note
    _, merged_note = evaluate_parse_fields(fields, base_note)
    return merged_note


def manual_create(
    db: Session,
    *,
    created_by_user_id: int | None,
    site_standard_name: str,
    reporter_name: str,
    accident_datetime_text: str | None,
    accident_datetime: datetime | None,
    accident_place: str | None,
    work_content: str | None,
    injured_person_name: str | None,
    accident_circumstance: str | None,
    accident_reason: str | None,
    injured_part: str | None,
    diagnosis_name: str | None,
    action_taken: str | None,
    status: str,
    management_category: str,
    notes: str | None,
    message_raw: str | None,
    parse_status_override: str | None = None,
    parse_note_override: str | None = None,
) -> Accident:
    fields = {
        "site_name": normalize_site_name(site_standard_name),
        "reporter_name": normalize_text(reporter_name),
        "accident_datetime_text": normalize_text(accident_datetime_text)
        or (accident_datetime.strftime("%Y-%m-%d %H:%M") if accident_datetime else None),
        "accident_datetime": accident_datetime,
        "accident_place": normalize_text(accident_place),
        "work_content": normalize_text(work_content),
        "injured_person_name": normalize_text(injured_person_name),
        "accident_circumstance": normalize_text(accident_circumstance),
        "accident_reason": normalize_text(accident_reason),
        "injured_part": normalize_text(injured_part),
        "diagnosis_name": normalize_text(diagnosis_name),
        "action_taken": normalize_text(action_taken),
    }
    evaluated_status, evaluated_note = evaluate_parse_fields(
        fields,
        note={"patterns": {"manual_input": True}},
    )
    parse_status = parse_status_override or evaluated_status
    parse_note = (
        build_manual_supplement_note(
            fields=fields,
            original_parse_status=parse_status_override,
            original_parse_note=parse_note_override,
        )
        if parse_status_override
        else evaluated_note
    )
    generated_message = compose_initial_report_view(fields, message_raw or "")["composed_line"]
    return _create_accident_row(
        db,
        created_by_user_id=created_by_user_id,
        source_type="manual_input",
        message_raw=normalize_text(message_raw) or generated_message,
        fields=fields,
        parse_status=parse_status,
        parse_note=parse_note,
        status=status,
        management_category=management_category,
        verification_status=default_verification_status(),
        notes=notes,
    )


def create_from_request(db: Session, *, body, created_by_user_id: int | None) -> Accident:
    if getattr(body, "input_mode", "auto") == "manual":
        status = body.status or "신규"
        management_category = body.management_category or "일반"
        if status not in STATUS_OPTIONS:
            raise HTTPException(status_code=400, detail="허용되지 않는 상태값입니다.")
        if management_category not in MANAGEMENT_CATEGORY_OPTIONS:
            raise HTTPException(status_code=400, detail="허용되지 않는 관리구분입니다.")
        return manual_create(
            db,
            created_by_user_id=created_by_user_id,
            site_standard_name=body.site_standard_name or "",
            reporter_name=body.reporter_name or "",
            accident_datetime_text=body.accident_datetime_text,
            accident_datetime=body.accident_datetime,
            accident_place=body.accident_place,
            work_content=body.work_content,
            injured_person_name=body.injured_person_name,
            accident_circumstance=body.accident_circumstance,
            accident_reason=body.accident_reason,
            injured_part=body.injured_part,
            diagnosis_name=body.diagnosis_name,
            action_taken=body.action_taken,
            status=status,
            management_category=management_category,
            notes=body.notes,
            message_raw=body.message_raw,
            parse_status_override=body.parse_status_override,
            parse_note_override=body.parse_note_override,
        )
    return parse_and_create(db, message_raw=body.message_raw or "", created_by_user_id=created_by_user_id)


def list_accidents(
    db: Session,
    *,
    queue_keys: list[str] | None = None,
    statuses: list[str] | None = None,
    management_categories: list[str] | None = None,
    only_incomplete: bool = False,
    default_queue_only: bool = True,
    limit: int = 500,
) -> list[Accident]:
    return repository.list_accidents(
        db,
        queue_keys=queue_keys,
        statuses=statuses,
        management_categories=management_categories,
        only_incomplete=only_incomplete,
        default_queue_only=default_queue_only,
        limit=limit,
    )


def get_accident_or_404(db: Session, accident_id: int) -> Accident:
    row = repository.get_accident(db, accident_id)
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="사고를 찾을 수 없습니다.")
    return row


def initial_report_view(row: Accident) -> dict:
    fields = {
        "site_name": row.site_standard_name or row.site_name,
        "reporter_name": row.reporter_name,
        "accident_datetime_text": row.accident_datetime_text,
        "accident_datetime": row.accident_datetime,
        "accident_place": row.accident_place,
        "work_content": row.work_content,
        "injured_person_name": row.injured_person_name,
        "accident_circumstance": row.accident_circumstance,
        "accident_reason": row.accident_reason,
        "injured_part": row.injured_part,
        "diagnosis_name": row.diagnosis_name,
        "action_taken": row.action_taken,
    }
    return compose_initial_report_view(fields, row.message_raw)


def update_accident(
    db: Session,
    *,
    accident_id: int,
    actor_user_id: int | None,
    site_standard_name: str,
    reporter_name: str | None,
    status: str,
    management_category: str,
    accident_datetime_text: str | None,
    accident_datetime: datetime | None,
    accident_place: str | None,
    work_content: str | None,
    injured_person_name: str | None,
    accident_circumstance: str | None,
    accident_reason: str | None,
    injured_part: str | None,
    diagnosis_name: str | None,
    action_taken: str | None,
    notes: str | None,
    initial_report_template: str | None,
) -> Accident:
    row = get_accident_or_404(db, accident_id)
    standardized = normalize_site_name(site_standard_name)
    if status not in STATUS_OPTIONS:
        raise HTTPException(status_code=400, detail="허용되지 않는 상태값입니다.")
    if management_category not in MANAGEMENT_CATEGORY_OPTIONS:
        raise HTTPException(status_code=400, detail="허용되지 않는 관리구분입니다.")
    if standardized:
        repository.ensure_site_standard(db, standardized)
    return repository.update_accident(
        db,
        row,
        site_name=standardized,
        site_standard_name=standardized,
        reporter_name=normalize_text(reporter_name),
        status=status,
        management_category=management_category,
        verification_status=default_verification_status(),
        accident_datetime_text=accident_datetime_text,
        accident_datetime=accident_datetime,
        accident_place=(accident_place or "").strip() or None,
        work_content=(work_content or "").strip() or None,
        injured_person_name=(injured_person_name or "").strip() or None,
        accident_circumstance=(accident_circumstance or "").strip() or None,
        accident_reason=(accident_reason or "").strip() or None,
        injured_part=(injured_part or "").strip() or None,
        diagnosis_name=(diagnosis_name or "").strip() or None,
        action_taken=(action_taken or "").strip() or None,
        notes=(notes or "").strip() or None,
        initial_report_template=(initial_report_template or "").strip() or None,
        is_complete=compute_is_complete(
            site_standard_name=standardized,
            accident_datetime=accident_datetime,
            injured_person_name=(injured_person_name or "").strip() or None,
        ),
        updated_by_user_id=actor_user_id,
    )


async def save_attachment(
    db: Session,
    *,
    accident_id: int,
    upload,
) -> AccidentAttachment:
    row = get_accident_or_404(db, accident_id)
    folder = resolve_accident_folder(row)
    safe_name = Path(upload.filename or "attachment.bin").name
    target = folder / f"{uuid.uuid4().hex}_{safe_name}"
    with target.open("wb") as fh:
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            fh.write(chunk)
    await upload.close()
    base = Path(settings.accident_nas_root) if settings.accident_nas_root else Path(settings.storage_root) / "accidents"
    relative = target.relative_to(base)
    return repository.add_attachment(
        db,
        accident_id_fk=row.id,
        file_name=safe_name,
        stored_path=str(relative).replace("\\", "/"),
        content_type=getattr(upload, "content_type", None),
        file_size=target.stat().st_size if target.exists() else None,
    )


def get_attachment_or_404(db: Session, attachment_id: int) -> AccidentAttachment:
    row = repository.get_attachment(db, attachment_id)
    if row:
        return row
    raise HTTPException(status_code=404, detail="첨부 파일을 찾을 수 없습니다.")


def attachment_absolute_path(row: AccidentAttachment) -> Path:
    base = Path(settings.accident_nas_root) if settings.accident_nas_root else Path(settings.storage_root) / "accidents"
    return base / row.stored_path


def _candidate_accident_folder(row: Accident) -> Path | None:
    if row.nas_folder_path:
        return Path(row.nas_folder_path)
    folder_key = row.nas_folder_key or row.accident_id
    if not folder_key:
        return None
    return accident_storage_root(row.accident_datetime) / folder_key


def _is_under_root(path: Path, root: Path) -> bool:
    try:
        path.resolve(strict=False).relative_to(root.resolve(strict=False))
        return True
    except ValueError:
        return False


def delete_accident(db: Session, *, accident_id: int) -> None:
    row = get_accident_or_404(db, accident_id)
    folder = _candidate_accident_folder(row)

    for attachment in list(row.attachments):
        path = attachment_absolute_path(attachment)
        if path.exists():
            path.unlink(missing_ok=True)

    if folder and folder.exists():
        root = accident_storage_root(row.accident_datetime)
        if _is_under_root(folder, root):
            shutil.rmtree(folder, ignore_errors=True)

    repository.delete_accident(db, row)


def lookups(db: Session) -> dict[str, list[str]]:
    sites = [row.site_name for row in repository.list_site_standards(db) if row.site_name]
    return {
        "statuses": STATUS_OPTIONS,
        "management_categories": MANAGEMENT_CATEGORY_OPTIONS,
        "site_names": sites,
    }


def _accident_id_sort_key(value: str | None) -> tuple[int, int, str]:
    raw = (value or "").strip()
    try:
        head, tail = raw.split("-", 1)
        return (int(head), int(tail), raw)
    except (ValueError, AttributeError):
        return (0, 0, raw)


def _master_headers_from_sheet(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
        if str(ws.cell(1, c).value or "").strip()
    }


def _copy_template_row_style(ws, *, template_row: int, target_row: int, max_col: int) -> None:
    if template_row == target_row:
        return
    source_height = ws.row_dimensions[template_row].height
    if source_height is not None:
        ws.row_dimensions[target_row].height = source_height
    for col in range(1, max_col + 1):
        ws.cell(target_row, col)._style = copy(ws.cell(template_row, col)._style)


def _clear_master_row_values(ws, *, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row, col).value = None


def _formula_ref(headers: dict[str, int], name: str, row: int) -> str:
    return f"{get_column_letter(headers[name])}{row}"


def apply_row_formulas(ws, row: int, headers: dict[str, int]) -> None:
    formula_map: dict[str, str] = {}
    if "발생연도" in headers and "사고일시" in headers:
        formula_map["발생연도"] = f"=YEAR({_formula_ref(headers, '사고일시', row)})"
    if "연령" in headers and "생년월일" in headers:
        formula_map["연령"] = f'=DATEDIF({_formula_ref(headers, "생년월일", row)},TODAY(),"Y")'
    if "고용형태" in headers:
        formula_map["고용형태"] = '="일용직"'
    if {"증빙폴더링크", "NAS폴더경로"}.issubset(headers):
        formula_map["증빙폴더링크"] = f'=HYPERLINK({_formula_ref(headers, "NAS폴더경로", row)},"폴더열기")'
    if {"원본매칭키", "사고일시", "성명", "현장명"}.issubset(headers):
        formula_map["원본매칭키"] = (
            f'=TEXT({_formula_ref(headers, "사고일시", row)},"yyyymmdd")'
            f'&"_"&{_formula_ref(headers, "성명", row)}'
            f'&"_"&{_formula_ref(headers, "현장명", row)}'
        )
    if {"출력키", "사고ID"}.issubset(headers):
        formula_map["출력키"] = f"={_formula_ref(headers, '사고ID', row)}"

    for name, formula in formula_map.items():
        ws.cell(row, headers[name], formula)


def _accident_to_master_values(row: Accident) -> dict[str, object | None]:
    accident_dt = row.accident_datetime or None
    site_name = row.site_standard_name or row.site_name
    injured_name = row.injured_person_name
    year = accident_dt.year if accident_dt else datetime.now().year
    folder_key = None
    if accident_dt and site_name and injured_name:
        folder_key = f"{accident_dt.strftime('%m%d')}_{injured_name}_{site_name}"
    local_folder_path = (
        f"D:\\104 사고 조사 및 이력 (산재요양승인 내역)\\{year}\\{folder_key}"
        if folder_key
        else None
    )
    nas_folder_path = (
        f"Z:\\4. 안전보건관리실\\104 사고 조사 및 이력 (산재요양승인 내역)\\{year}\\{folder_key}"
        if folder_key
        else row.nas_folder_path
    )
    return {
        "등록일": row.created_at,
        "수정일": row.updated_at,
        "사고ID": row.accident_id,
        "사고일시": accident_dt,
        "발생연도": accident_dt.year if accident_dt else None,
        "현장명": site_name,
        "작업명": row.work_content,
        "성명": injured_name,
        "사고장소": row.accident_place,
        "사고유형": None,
        "재해부위": row.injured_part,
        "사고개요": row.accident_circumstance or row.message_raw,
        "직접원인": row.accident_reason,
        "간접원인": row.parse_note,
        "사고상태": row.status or "신규",
        "치료현황": row.action_taken,
        "보고자": row.reporter_name,
        "비고": row.notes,
        "폴더키": folder_key,
        "로컬폴더경로": local_folder_path,
        "NAS폴더경로": nas_folder_path,
        "증빙폴더링크": nas_folder_path,
        "증빙확인여부": "Y" if row.attachments else "N",
        "원본분석표_시트명": "ACCIDENT_WEB",
        "원본분석표_No": row.id,
        "원본매칭키": row.display_code,
        "데이터검증상태": row.parse_status,
        "출력키": row.display_code,
        "진행현황": "웹운영",
        "관리구분": row.management_category,
        "폴더존재여부": "Y" if nas_folder_path else "N",
        "파일개수": len(row.attachments),
        "증빙상태": (
            "정상"
            if nas_folder_path and row.attachments
            else "파일없음"
            if nas_folder_path
            else "폴더없음"
        ),
    }


def get_worklist(db: Session) -> dict:
    rows = repository.list_accidents(db, default_queue_only=False, limit=5000)

    def section(items: list[Accident]) -> dict:
        return {"count": len(items), "items": items[:20]}

    unverified: list[Accident] = []
    parse_review = [row for row in rows if row.parse_status != "success"]
    missing_attachments = [row for row in rows if not row.has_attachments]
    recent = sorted(rows, key=lambda row: row.created_at, reverse=True)

    return {
        "unverified": section(unverified),
        "parse_review": section(parse_review),
        "missing_attachments": section(missing_attachments),
        "recent": section(recent),
    }


def export_master_workbook(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = MASTER_IMPORT_SHEET
    ws.append(MASTER_HEADERS)
    rows = repository.list_accidents(db, default_queue_only=False, limit=5000)
    for row in rows:
        accident_dt = row.accident_datetime or None
        status = row.status or "신규"
        ws.append(
            [
                row.created_at,
                row.updated_at,
                row.accident_id,
                accident_dt,
                accident_dt.year if accident_dt else None,
                row.site_standard_name or row.site_name,
                row.work_content,
                row.injured_person_name,
                None,
                None,
                None,
                None,
                None,
                row.accident_place,
                None,
                row.injured_part,
                row.accident_circumstance or row.message_raw,
                row.accident_reason,
                row.parse_note,
                status,
                row.action_taken,
                None,
                None,
                None,
                None,
                None,
                row.reporter_name,
                row.notes,
                row.nas_folder_key,
                None,
                row.nas_folder_path,
                row.nas_folder_path,
                "Y" if row.attachments else "N",
                "ACCIDENT_WEB",
                row.id,
                row.display_code,
                row.parse_status,
                row.display_code,
                "웹운영",
                row.management_category,
            ]
        )
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_verified_accidents_to_master_excel(
    db: Session,
    *,
    workbook_path: Path | None = None,
) -> dict[str, object]:
    workbook_path = workbook_path or MASTER_IMPORT_PATH
    if not workbook_path.is_file():
        raise FileNotFoundError(f"MASTER 파일이 없습니다: {workbook_path}")

    rows = repository.list_accidents(db, default_queue_only=False, limit=5000)
    export_rows = sorted(rows, key=lambda item: _accident_id_sort_key(item.accident_id))
    excluded_count = 0

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = workbook_path.with_name(f"{workbook_path.stem}.bak_{timestamp}{workbook_path.suffix}")
    temp_path = workbook_path.with_name(f"{workbook_path.stem}.tmp_{timestamp}{workbook_path.suffix}")

    shutil.copy2(workbook_path, backup_path)
    wb = None
    try:
        wb = load_workbook(workbook_path, data_only=False, read_only=False)
        if MASTER_IMPORT_SHEET not in wb.sheetnames:
            raise KeyError(f"MASTER 시트가 없습니다: {MASTER_IMPORT_SHEET}")
        ws = wb[MASTER_IMPORT_SHEET]
        headers = _master_headers_from_sheet(ws)
        if "사고ID" not in headers:
            raise KeyError("MASTER 헤더에 사고ID가 없습니다.")

        max_col = ws.max_column
        max_row = ws.max_row
        template_row = 2 if max_row >= 2 else 1

        required_rows = max(2, len(export_rows) + 1)
        while ws.max_row < required_rows:
            next_row = ws.max_row + 1
            _copy_template_row_style(ws, template_row=template_row, target_row=next_row, max_col=max_col)

        for row_idx in range(2, ws.max_row + 1):
            if row_idx > template_row:
                _copy_template_row_style(ws, template_row=template_row, target_row=row_idx, max_col=max_col)
            _clear_master_row_values(ws, row=row_idx, max_col=max_col)

        for idx, row in enumerate(export_rows, start=2):
            values = _accident_to_master_values(row)
            for header, col_idx in headers.items():
                if header in values:
                    ws.cell(idx, col_idx, values[header])
            apply_row_formulas(ws, idx, headers)

        wb.save(temp_path)
        wb.close()
        wb = None
        shutil.copy2(temp_path, workbook_path)
    except Exception:
        if wb is not None:
            wb.close()
        if backup_path.is_file():
            shutil.copy2(backup_path, workbook_path)
        raise
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)

    return {
        "target_path": str(workbook_path),
        "backup_path": str(backup_path),
        "exported_count": len(export_rows),
        "excluded_count": excluded_count,
        "backup_created": backup_path.is_file(),
    }


def import_master_rows(db: Session, *, workbook_path: Path = MASTER_IMPORT_PATH) -> dict[str, int]:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"MASTER 파일이 없습니다: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws = wb[MASTER_IMPORT_SHEET]
        headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
        required = ["사고ID", "사고일시", "현장명", "성명", "사고상태", "관리구분", "NAS폴더경로"]
        for name in required:
            if name not in headers:
                raise KeyError(f"MASTER 헤더 없음: {name}")

        imported = 0
        skipped = 0
        for r in range(2, ws.max_row + 1):
            accident_code = str(ws.cell(r, headers["사고ID"]).value or "").strip()
            if not accident_code:
                continue
            if repository.get_accident_by_accident_code(db, accident_code):
                skipped += 1
                continue
            accident_dt_value = ws.cell(r, headers["사고일시"]).value
            accident_dt = accident_dt_value if isinstance(accident_dt_value, datetime) else None
            site_name = normalize_site_name(ws.cell(r, headers["현장명"]).value)
            name = str(ws.cell(r, headers["성명"]).value or "").strip() or None
            status = str(ws.cell(r, headers["사고상태"]).value or "").strip() or "접수"
            if status not in STATUS_OPTIONS:
                status = "접수"
            management_category = str(ws.cell(r, headers["관리구분"]).value or "").strip() or "일반"
            if management_category not in MANAGEMENT_CATEGORY_OPTIONS:
                management_category = "일반"
            nas_folder_path = str(ws.cell(r, headers["NAS폴더경로"]).value or "").strip() or None
            notes = str(ws.cell(r, headers.get("비고", 0)).value or "").strip() if headers.get("비고") else None
            work_content = str(ws.cell(r, headers.get("작업명", 0)).value or "").strip() if headers.get("작업명") else None
            accident_place = str(ws.cell(r, headers.get("사고장소", 0)).value or "").strip() if headers.get("사고장소") else None
            injured_part = str(ws.cell(r, headers.get("재해부위", 0)).value or "").strip() if headers.get("재해부위") else None
            diagnosis_name = str(ws.cell(r, headers.get("사고유형", 0)).value or "").strip() if headers.get("사고유형") else None
            reason = str(ws.cell(r, headers.get("직접원인", 0)).value or "").strip() if headers.get("직접원인") else None
            message_raw = str(ws.cell(r, headers.get("사고개요", 0)).value or "").strip() if headers.get("사고개요") else ""
            repository.ensure_site_standard(db, site_name or "")
            repository.create_accident(
                db,
                display_code=f"ACC-{accident_code}",
                accident_id=accident_code,
                source_type="master_import",
                message_raw=message_raw or f"{site_name or '(현장미상)'} / {name or '(성명미상)'}",
                parse_status="imported",
                parse_note=None,
                site_name=site_name,
                reporter_name=None,
                accident_datetime_text=str(accident_dt_value or "").strip() or None,
                accident_datetime=accident_dt,
                accident_place=accident_place or None,
                work_content=work_content or None,
                injured_person_name=name,
                accident_circumstance=message_raw or None,
                accident_reason=reason or None,
                injured_part=injured_part or None,
                diagnosis_name=diagnosis_name or None,
                action_taken=str(ws.cell(r, headers.get("치료현황", 0)).value or "").strip() if headers.get("치료현황") else None,
                status=status,
                management_category=management_category,
                verification_status=default_verification_status(),
                site_standard_name=site_name,
                initial_report_template=None,
                is_complete=compute_is_complete(
                    site_standard_name=site_name,
                    accident_datetime=accident_dt,
                    injured_person_name=name,
                ),
                nas_folder_path=nas_folder_path,
                nas_folder_key=accident_code,
                notes=notes or None,
                created_by_user_id=None,
                updated_by_user_id=None,
            )
            imported += 1
        return {"imported": imported, "skipped": skipped}
    finally:
        wb.close()
