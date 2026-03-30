from __future__ import annotations

from dataclasses import dataclass, field
import csv
import io
import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
import xlrd

from app.config.settings import settings

logger = logging.getLogger(__name__)

OLE_XLS_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def open_xlrd_workbook(path: Path | str, **kwargs: Any) -> Any:
    """xlrd 버전에 따라 ignore_workbook_corruption 미지원일 수 있어 폴백한다."""
    p = str(path)
    try:
        return xlrd.open_workbook(p, ignore_workbook_corruption=True, **kwargs)
    except TypeError:
        return xlrd.open_workbook(p, **kwargs)
ZIP_SIGNATURE = b"PK\x03\x04"


@dataclass
class ParseAttempt:
    stage: str
    parser: str
    success: bool
    message: str


@dataclass
class ParsedSheet:
    headers: list[str]
    rows: list[list[Any]]
    parser_used: str
    fallback_used: bool
    file_type: str
    attempts: list[ParseAttempt] = field(default_factory=list)
    standardized_xlsx_path: str | None = None

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def columns(self) -> list[str]:
        return self.headers

    def to_diagnostics(self, required_columns: list[str] | None = None) -> dict[str, Any]:
        required = required_columns or []
        missing = [col for col in required if col not in self.headers]
        return {
            "file_type": self.file_type,
            "parser_used": self.parser_used,
            "fallback_used": self.fallback_used,
            "row_count": self.row_count,
            "columns": self.columns,
            "required_columns": required,
            "required_columns_ok": len(missing) == 0,
            "missing_required_columns": missing,
            "attempts": [
                {
                    "stage": at.stage,
                    "parser": at.parser,
                    "success": at.success,
                    "message": at.message,
                }
                for at in self.attempts
            ],
            "standardized_xlsx_path": self.standardized_xlsx_path,
        }

    def to_public_diagnostics(
        self,
        *,
        header_valid: bool,
        warnings: list[str] | None = None,
    ) -> dict[str, Any]:
        failures: list[dict[str, str]] = []
        for at in self.attempts:
            if at.success:
                continue
            failures.append(
                {
                    "stage": at.stage,
                    "parser": at.parser,
                    "message": _sanitize_failure_message(at.message),
                }
            )
        normalized_created = bool(self.standardized_xlsx_path)
        normalized_filename = Path(self.standardized_xlsx_path).name if normalized_created else None
        return {
            "detected_type": self.file_type,
            "selected_parser": self.parser_used,
            "normalized_file_created": normalized_created,
            "normalized_filename": normalized_filename,
            "header_valid": header_valid,
            "row_count": self.row_count,
            "warnings": warnings or [],
            "failures": failures,
        }


class IngestionPipelineError(ValueError):
    def __init__(self, message: str, diagnostics: dict[str, Any] | None = None):
        super().__init__(message)
        self.diagnostics = diagnostics or {}


def detect_file_type(path: Path) -> str:
    ext = path.suffix.lower()
    with path.open("rb") as fp:
        sig = fp.read(8)
    if sig.startswith(OLE_XLS_SIGNATURE):
        return "xls"
    if sig.startswith(ZIP_SIGNATURE):
        return "xlsx"
    if ext == ".csv":
        return "csv"
    # 마지막 fallback: 확장자 기반
    if ext in {".xls", ".xlsx"}:
        return ext.lstrip(".")
    return "unknown"


def _sanitize_failure_message(raw_message: str) -> str:
    text = raw_message or ""
    lower = text.lower()
    if "module not found" in lower or "no module named" in lower:
        return "Parser not available"
    if "workbook corruption" in lower or "compdocerror" in lower:
        return "Workbook corruption detected"
    if "csv" in lower and "not" in lower and "target" in lower:
        return "File format could not be parsed"
    if "invalidfileexception" in lower:
        return "File format could not be parsed"
    # traceback/경로/내부 문자열 제거
    text = re.sub(r"File \".*?\"", "", text)
    text = re.sub(r"[A-Za-z]:\\\\[^\\s]+", "", text)
    text = re.sub(r"[A-Za-z]:/[^\\s]+", "", text)
    text = re.sub(r"\\s+", " ", text).strip()
    if not text:
        return "File format could not be parsed"
    return "File format could not be parsed"


def _parse_with_openpyxl(path: Path) -> tuple[list[str], list[list[Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    first_row = next(iterator, None)
    if first_row is None:
        return [], []
    headers = [str(v).strip() if v is not None else "" for v in first_row]
    rows: list[list[Any]] = []
    for row in iterator:
        values = list(row)
        if any(v not in (None, "") for v in values):
            rows.append(values)
    return headers, rows


def _parse_with_xlrd(path: Path, *, alt_mode: bool = False) -> tuple[list[str], list[list[Any]]]:
    wb = open_xlrd_workbook(path, on_demand=alt_mode)
    sheet = wb.sheet_by_index(0)
    if sheet.nrows <= 0 or sheet.ncols <= 0:
        return [], []
    headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    rows: list[list[Any]] = []
    for r in range(1, sheet.nrows):
        values = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        if any(str(v).strip() for v in values):
            rows.append(values)
    return headers, rows


def _parse_with_pandas(path: Path) -> tuple[list[str], list[list[Any]]]:
    import pandas as pd  # optional dependency

    df = pd.read_excel(path, sheet_name=0, dtype=object)
    headers = [str(c).strip() for c in list(df.columns)]
    rows = df.fillna("").values.tolist()
    return headers, rows


def _parse_with_csv_fallback(path: Path) -> tuple[list[str], list[list[Any]]]:
    raw = path.read_bytes()
    # 바이너리 OLE/ZIP 파일은 CSV 강제 읽기를 시도하지 않는다.
    if raw.startswith(OLE_XLS_SIGNATURE) or raw.startswith(ZIP_SIGNATURE):
        raise ValueError("CSV 강제 읽기 대상이 아닙니다(바이너리 엑셀 시그니처 감지)")
    for enc in ("utf-8-sig", "cp949", "euc-kr", "latin-1"):
        try:
            text = raw.decode(enc)
            reader = csv.reader(io.StringIO(text))
            data = list(reader)
            if not data:
                continue
            headers = [str(c).strip() for c in data[0]]
            rows = [row for row in data[1:] if any(str(v).strip() for v in row)]
            return headers, rows
        except Exception:
            continue
    raise ValueError("CSV 강제 읽기에 실패했습니다")


def _write_standardized_xlsx(path: Path, headers: list[str], rows: list[list[Any]]) -> Path:
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "normalized"
    sh.append(headers)
    for row in rows:
        sh.append(row)
    out_dir = settings.storage_root / "ingestion_normalized"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"{path.stem}.normalized.xlsx"
    wb.save(out)
    return out


def parse_excel_with_fallback(path: Path) -> ParsedSheet:
    if not path.exists():
        raise FileNotFoundError(f"입력 파일이 존재하지 않습니다: {path}")

    file_type = detect_file_type(path)
    attempts: list[ParseAttempt] = []
    headers: list[str] = []
    rows: list[list[Any]] = []
    parser_used = ""
    fallback_used = False
    standardized_path: Path | None = None

    def _try(stage: str, parser: str, fn) -> bool:
        nonlocal headers, rows, parser_used, fallback_used
        try:
            h, r = fn()
            headers, rows = h, r
            parser_used = parser
            fallback_used = stage != "primary"
            attempts.append(ParseAttempt(stage=stage, parser=parser, success=True, message="ok"))
            logger.info("file_ingestion success: stage=%s parser=%s rows=%s", stage, parser, len(rows))
            return True
        except Exception as exc:  # noqa: BLE001
            msg = f"{type(exc).__name__}: {exc}"
            attempts.append(ParseAttempt(stage=stage, parser=parser, success=False, message=msg))
            logger.warning("file_ingestion failed: stage=%s parser=%s error=%s", stage, parser, msg)
            return False

    if file_type == "xlsx":
        if not _try("primary", "openpyxl", lambda: _parse_with_openpyxl(path)):
            _try("fallback", "pandas.read_excel", lambda: _parse_with_pandas(path))
    elif file_type == "xls":
        if not _try("primary", "xlrd", lambda: _parse_with_xlrd(path, alt_mode=False)):
            _try("fallback", "pandas.read_excel", lambda: _parse_with_pandas(path))
        if not parser_used:
            _try("fallback", "xlrd(on_demand=True)", lambda: _parse_with_xlrd(path, alt_mode=True))
        if not parser_used:
            _try("fallback", "csv-force-read", lambda: _parse_with_csv_fallback(path))
    elif file_type == "csv":
        _try("primary", "csv", lambda: _parse_with_csv_fallback(path))
    else:
        # 유형이 불명확하면 안전하게 시도
        if not _try("fallback", "openpyxl", lambda: _parse_with_openpyxl(path)):
            if not _try("fallback", "xlrd", lambda: _parse_with_xlrd(path, alt_mode=False)):
                _try("fallback", "csv-force-read", lambda: _parse_with_csv_fallback(path))

    if not parser_used:
        diagnostics = {
            "detected_type": file_type,
            "selected_parser": None,
            "normalized_file_created": False,
            "normalized_filename": None,
            "header_valid": False,
            "row_count": 0,
            "warnings": [],
            "failures": [
                {
                    "stage": at.stage,
                    "parser": at.parser,
                    "message": _sanitize_failure_message(at.message),
                }
                for at in attempts
                if not at.success
            ],
        }
        raise IngestionPipelineError("파일 파싱에 실패했습니다.", diagnostics=diagnostics)

    if headers and rows:
        try:
            standardized_path = _write_standardized_xlsx(path, headers, rows)
            logger.info("file_ingestion standardized xlsx: %s", standardized_path)
        except Exception as exc:  # noqa: BLE001
            logger.warning("file_ingestion standardize failed: %s", exc)

    return ParsedSheet(
        headers=headers,
        rows=rows,
        parser_used=parser_used,
        fallback_used=fallback_used,
        file_type=file_type,
        attempts=attempts,
        standardized_xlsx_path=str(standardized_path) if standardized_path is not None else None,
    )


def validate_worker_file_structure(headers: list[str], required_columns: list[str]) -> dict[str, Any]:
    normalized_headers = [str(h).strip() for h in headers]
    missing = [c for c in required_columns if c not in normalized_headers]
    return {
        "has_sheet": len(headers) > 0,
        "header_detected": len(headers) > 0,
        "columns": normalized_headers,
        "required_columns": required_columns,
        "required_columns_ok": len(missing) == 0,
        "missing_required_columns": missing,
    }

