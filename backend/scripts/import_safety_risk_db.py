from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl

# Ensure "app" package import works when script runs directly.
BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.core.database import SessionLocal  # noqa: E402
from app.core.datetime_utils import utc_now  # noqa: E402
from app.modules.risk_library.models import (  # noqa: E402
    DailyWorkPlanItemRiskRef,
    RiskLibraryItem,
    RiskLibraryItemRevision,
    RiskLibraryKeyword,
)


TOKEN_STOPWORDS = {
    "process",
    "risk",
    "factor",
    "counterplan",
    "note",
    "유해",
    "위험",
    "요인",
    "세부",
    "작업",
    "감소",
    "대책",
}

PHRASE_SUFFIXES = {"배관", "배선", "설치", "검토", "작업"}
FORCED_DOMAIN_TOKENS = {
    "배관",
    "벽체 배관",
    "천장슬라브 배관",
    "전선관 배관",
    "배선",
    "등기구",
    "전주",
    "고소작업대",
    "이동식비계",
    "사다리",
    "감전",
    "추락",
    "낙하",
    "협착",
}


@dataclass
class ParsedRiskRow:
    unit_work: str
    work_category: str
    trade_type: str
    process: str
    risk_factor: str
    risk_cause: str
    counterplan: str
    note: str
    risk_f: int
    risk_s: int
    risk_r: int
    keywords: list[str]
    source_file: str
    source_sheet: str
    source_row: int
    source_page_or_section: str


def _norm_text(value: object) -> str:
    if value is None:
        return ""
    s = str(value).replace("\n", " ").strip()
    return " ".join(s.split())


def _to_int(value: object) -> int | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def _clean_counterplan(text: str) -> str:
    return text.lstrip("▶").strip()


def _extract_keywords(process: str, risk_factor: str, counterplan: str) -> list[str]:
    text = f"{process} {risk_factor} {counterplan}".lower()
    raw_tokens = [t for t in re.split(r"[\s,./()\-:;|]+", text) if t]
    tokens: list[str] = []
    seen: set[str] = set()

    for token in raw_tokens:
        token = token.strip()
        if len(token) < 2:
            continue
        if token in TOKEN_STOPWORDS:
            continue
        if token in seen:
            continue
        seen.add(token)
        tokens.append(token)

    for idx in range(len(raw_tokens) - 1):
        left = raw_tokens[idx].strip()
        right = raw_tokens[idx + 1].strip()
        if len(left) < 2 or len(right) < 2:
            continue
        if right not in PHRASE_SUFFIXES:
            continue
        phrase = f"{left} {right}"
        if phrase in seen:
            continue
        seen.add(phrase)
        tokens.append(phrase)

    for domain_token in sorted(FORCED_DOMAIN_TOKENS, key=len, reverse=True):
        if domain_token in text and domain_token not in seen:
            seen.add(domain_token)
            tokens.append(domain_token)

    return tokens


def _parse_effective_from(rows: list[list[object]]) -> date:
    raw = None
    if len(rows) >= 3 and len(rows[2]) >= 15:
        raw = rows[2][14]
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return date.today()


def _is_process_header_row(row: list[object]) -> bool:
    c1 = _norm_text(row[0] if len(row) > 0 else "").lower()
    c4 = _norm_text(row[3] if len(row) > 3 else "").lower()
    return "process" in c1 and "risk factor" in c4


def _find_process_header_rows(rows: list[list[object]]) -> list[int]:
    return [idx for idx, row in enumerate(rows, start=1) if _is_process_header_row(row)]


def _is_risk_data_row(row: list[object]) -> bool:
    c4 = _norm_text(row[3] if len(row) > 3 else "")
    c16 = _to_int(row[15] if len(row) > 15 else None)
    c17 = _to_int(row[16] if len(row) > 16 else None)
    c18 = _to_int(row[17] if len(row) > 17 else None)
    c19 = _norm_text(row[18] if len(row) > 18 else "")
    return bool(c4 and c19 and c16 is not None and c17 is not None and c18 is not None)


def _is_non_data_break_row(row: list[object]) -> bool:
    c1 = _norm_text(row[0] if len(row) > 0 else "").lower()
    c4 = _norm_text(row[3] if len(row) > 3 else "").lower()
    if "process" in c1 or "risk factor" in c4:
        return True
    if "작성일" in c1 or "특이사항" in c1 or "comments" in c1:
        return True
    return False


def _resolve_source_workbook() -> Path:
    docs_dir = BACKEND_ROOT.parent / "docs"
    preferred = docs_dir / "최초위험성평가 표준 모델(배포용).xlsx"
    if preferred.exists():
        return preferred

    deployment_candidates = sorted(
        [p for p in docs_dir.glob("*.xlsx") if "배포용" in p.name],
        key=lambda p: p.name,
    )
    if deployment_candidates:
        return deployment_candidates[0]

    fallback = docs_dir / "safetyRisk_DB.xlsx"
    if fallback.exists():
        return fallback
    raise FileNotFoundError(f"Source workbook not found under: {docs_dir}")


def parse_risk_rows(xlsx_path: Path) -> tuple[str, list[ParsedRiskRow], date]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    effective_from = _parse_effective_from(rows)

    header_rows = _find_process_header_rows(rows)
    if not header_rows:
        raise RuntimeError("Could not find safety risk table headers ('process' / 'risk factor').")

    parsed: list[ParsedRiskRow] = []
    seen_keys: set[tuple[str, str, str, int, int, int]] = set()

    for section_idx, header_row in enumerate(header_rows, start=1):
        section_name = f"section_{section_idx}"
        end_row = header_rows[section_idx] - 1 if section_idx < len(header_rows) else len(rows)
        current_process = ""

        for r in range(header_row + 1, end_row + 1):
            row = rows[r - 1]
            if _is_non_data_break_row(row):
                continue

            c1 = _norm_text(row[0] if len(row) > 0 else "")
            c4 = _norm_text(row[3] if len(row) > 3 else "")
            c16 = _to_int(row[15] if len(row) > 15 else None)
            c17 = _to_int(row[16] if len(row) > 16 else None)
            c18 = _to_int(row[17] if len(row) > 17 else None)
            c19 = _clean_counterplan(_norm_text(row[18] if len(row) > 18 else ""))
            c31_note = _norm_text(row[30] if len(row) > 30 else "")

            if not _is_risk_data_row(row):
                if c1 and "process" not in c1.lower():
                    current_process = c1
                continue

            if c1 and "process" not in c1.lower():
                current_process = c1

            process = current_process or c1 or "미분류"
            unit_work = process
            work_category = process
            trade_type = process
            risk_factor = c4
            counterplan = c19
            note = c31_note

            dedupe_key = (process, risk_factor, counterplan, int(c16), int(c17), int(c18))
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)

            keywords = _extract_keywords(process=process, risk_factor=risk_factor, counterplan=counterplan)
            parsed.append(
                ParsedRiskRow(
                    unit_work=unit_work,
                    work_category=work_category,
                    trade_type=trade_type,
                    process=process,
                    risk_factor=risk_factor,
                    risk_cause="미기재",
                    counterplan=counterplan,
                    note=note,
                    risk_f=int(c16),
                    risk_s=int(c17),
                    risk_r=int(c18),
                    keywords=keywords,
                    source_file=xlsx_path.name,
                    source_sheet=sheet.title,
                    source_row=r,
                    source_page_or_section=section_name,
                )
            )

    return sheet.title, parsed, effective_from


def import_to_db(xlsx_path: Path) -> dict[str, int | str]:
    sheet_name, rows, effective_from = parse_risk_rows(xlsx_path)
    print(
        f"[import] parsed rows={len(rows)} sheet={sheet_name} source={xlsx_path.name}",
        flush=True,
    )
    db = SessionLocal()
    try:
        # Reset risk library with idempotent full refresh.
        db.query(DailyWorkPlanItemRiskRef).delete(synchronize_session=False)
        db.query(RiskLibraryKeyword).delete(synchronize_session=False)
        db.query(RiskLibraryItemRevision).delete(synchronize_session=False)
        db.query(RiskLibraryItem).delete(synchronize_session=False)
        db.commit()
        print("[import] cleared existing risk tables", flush=True)

        items: list[RiskLibraryItem] = [
            RiskLibraryItem(source_scope="HQ_STANDARD", owner_site_id=None, is_active=True)
            for _ in rows
        ]
        db.add_all(items)
        db.flush()
        item_count = len(items)
        print(f"[import] inserted items={item_count}", flush=True)

        revisions: list[RiskLibraryItemRevision] = []
        for item, row in zip(items, rows):
            revisions.append(
                RiskLibraryItemRevision(
                    item_id=item.id,
                    revision_no=1,
                    is_current=True,
                    effective_from=effective_from,
                    effective_to=None,
                    unit_work=row.unit_work,
                    work_category=row.work_category,
                    trade_type=row.trade_type,
                    process=row.process,
                    risk_factor=row.risk_factor,
                    risk_cause=row.risk_cause,
                    countermeasure=row.counterplan,
                    note=row.note,
                    source_file=row.source_file,
                    source_sheet=row.source_sheet,
                    source_row=row.source_row,
                    source_page_or_section=row.source_page_or_section,
                    risk_f=row.risk_f,
                    risk_s=row.risk_s,
                    risk_r=row.risk_r,
                    revised_by_user_id=None,
                    revised_at=utc_now(),
                    revision_note=f"imported from {row.source_file}:{row.source_sheet}:{row.source_row}",
                )
            )
        db.add_all(revisions)
        db.flush()
        revision_count = len(revisions)
        print(f"[import] inserted revisions={revision_count}", flush=True)

        keywords: list[RiskLibraryKeyword] = []
        for rev, row in zip(revisions, rows):
            for kw in row.keywords:
                keywords.append(
                    RiskLibraryKeyword(
                        risk_revision_id=rev.id,
                        keyword=kw,
                        weight=1.0,
                    )
                )
        db.add_all(keywords)
        keyword_count = len(keywords)
        print(f"[import] prepared keywords={keyword_count}", flush=True)

        db.commit()
        print("[import] commit done", flush=True)
        return {
            "sheet_name": sheet_name,
            "item_count": item_count,
            "revision_count": revision_count,
            "keyword_count": keyword_count,
            "effective_from": str(effective_from),
        }
    finally:
        db.close()


def main() -> None:
    # Register FK target tables in metadata.
    from app.modules.sites import models as site_models  # noqa: F401
    from app.modules.users import models as user_models  # noqa: F401
    from app.modules.workers import models as worker_models  # noqa: F401

    xlsx_path = _resolve_source_workbook()
    result = import_to_db(xlsx_path)
    print(result)


if __name__ == "__main__":
    main()

