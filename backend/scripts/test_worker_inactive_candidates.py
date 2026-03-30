from pathlib import Path

from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.core.database import Base
from app.modules.workers.models import Employment, Person, WorkerInactiveCandidate
from app.modules.workers.service import diff_employees_from_path


def main() -> None:
    # in-memory DB
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)

    # create two baseline persons/employments
    p1 = Person(name="A", rrn_hash="h1")
    p2 = Person(name="B", rrn_hash="h2")
    session.add_all([p1, p2])
    session.flush()
    e1 = Employment(person_id=p1.id, source_type="employee", is_active=True)
    e2 = Employment(person_id=p2.id, source_type="employee", is_active=True)
    session.add_all([e1, e2])
    session.commit()

    # create temp xlsx with only p1 (so p2 becomes missing)
    import openpyxl
    from tempfile import NamedTemporaryFile

    wb = openpyxl.Workbook()
    sh = wb.active
    # minimal header row + data rows consistent with _load_employees_rows_from_xlsx expectations
    # name at col 1, rrn_front at col 7, rrn_back at col 8
    sh.cell(row=1, column=1, value="name")
    sh.cell(row=1, column=7, value="rrn_front")
    sh.cell(row=1, column=8, value="rrn_back")
    # row for p1
    sh.cell(row=2, column=1, value="A")
    sh.cell(row=2, column=7, value="h1f")
    sh.cell(row=2, column=8, value="h1b")

    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = Path(tmp.name)

    # monkeypatch hash so that "h1f"+"h1b" -> "h1"
    from app.modules.workers import service as svc

    orig_hash = svc._hash_rrn

    def fake_hash(rrn_raw: str) -> str:
        if rrn_raw == "h1fh1b":
            return "h1"
        return orig_hash(rrn_raw)

    svc._hash_rrn = fake_hash

    try:
        diff = diff_employees_from_path(session, tmp_path)
        print("missing_count", diff.missing_count)
        print("missing_sample", diff.missing_sample)

        cands = session.query(WorkerInactiveCandidate).all()
        print("candidates", [(c.rrn_hash, c.status, c.missing_streak) for c in cands])
    finally:
        svc._hash_rrn = orig_hash
        tmp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    main()

