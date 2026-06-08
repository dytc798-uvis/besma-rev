"""수동 작성된 기능인등급 템플릿 xlsx → legacy_site_grade_data.json (1회 추출).

Usage:
  cd backend && PYTHONPATH=. python scripts/extract_legacy_site_grade_workbook.py
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

BACKEND_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_ROOT))

from app.modules.functional_eval.legacy_site_grade import parse_eval_marks_from_row  # noqa: E402
from app.modules.functional_eval.site_grade_workbook import (  # noqa: E402
    COL_AGE,
    COL_GRADE_FUNCTIONAL,
    COL_GRADE_SAFETY,
    COL_JOB,
    COL_NAME,
    COL_PHONE,
    COL_POSITION,
    COL_RRN,
    COL_SEQ,
    COL_SITE_CODE,
    COL_SITE_NAME,
    SHEET1_DATA_START_ROW,
    SHEET_EVAL_START_ROW,
    _find_sheet_name,
    resolve_site_grade_template_path,
)

OUT = BACKEND_ROOT / "app" / "modules" / "functional_eval" / "legacy_site_grade_data.json"


def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


def extract() -> dict:
    template = resolve_site_grade_template_path()
    wb = openpyxl.load_workbook(template, data_only=True)
    ws1 = wb[_find_sheet_name(wb, "1.")]
    ws_fn = wb[_find_sheet_name(wb, "2-1")]
    ws_sf = wb[_find_sheet_name(wb, "2-2")]

    eval_offset = SHEET_EVAL_START_ROW - SHEET1_DATA_START_ROW
    workers: list[dict] = []
    for row1 in range(SHEET1_DATA_START_ROW, (ws1.max_row or SHEET1_DATA_START_ROW) + 1):
        name = _cell_str(ws1, row1, COL_NAME)
        if not name or name.startswith("="):
            continue
        eval_row = row1 + eval_offset
        functional = parse_eval_marks_from_row(ws_fn, eval_row, "FUNCTIONAL")
        safety = parse_eval_marks_from_row(ws_sf, eval_row, "SAFETY")
        if not functional and not safety:
            continue
        workers.append(
            {
                "legacy_row": row1,
                "site_code": _cell_str(ws1, row1, COL_SITE_CODE),
                "seq": ws1.cell(row1, COL_SEQ).value,
                "site_name": _cell_str(ws1, row1, COL_SITE_NAME),
                "name": name,
                "age_label": _cell_str(ws1, row1, COL_AGE),
                "rrn_masked": _cell_str(ws1, row1, COL_RRN),
                "position_name": _cell_str(ws1, row1, COL_POSITION),
                "job_name": _cell_str(ws1, row1, COL_JOB),
                "phone_mobile": _cell_str(ws1, row1, COL_PHONE),
                "sheet1_functional_grade": _cell_str(ws1, row1, COL_GRADE_FUNCTIONAL),
                "sheet1_safety_grade": _cell_str(ws1, row1, COL_GRADE_SAFETY),
                "functional_assessment": functional,
                "safety_assessment": safety,
            }
        )
    wb.close()
    return {
        "source_template": template.name,
        "extracted_at": datetime.now(timezone.utc).isoformat(),
        "worker_count": len(workers),
        "workers": workers,
    }


def main() -> None:
    payload = extract()
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    complete = sum(
        1
        for w in payload["workers"]
        if (w.get("functional_assessment") or {}).get("is_complete")
        and (w.get("safety_assessment") or {}).get("is_complete")
    )
    print(f"wrote {OUT}")
    print(f"workers={payload['worker_count']} fully_complete={complete}")


if __name__ == "__main__":
    main()
