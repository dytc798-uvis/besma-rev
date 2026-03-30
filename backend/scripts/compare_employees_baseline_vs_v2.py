import hashlib
from pathlib import Path

import openpyxl


def load_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb[wb.sheetnames[0]]
    out: list[dict] = []
    max_row = sh.max_row or 0
    max_col = sh.max_column or 0
    for r in range(2, max_row + 1):
        vals = [sh.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if not any(vals):
            continue
        rrn_front = vals[6]
        rrn_back = vals[7]
        if not rrn_front or not rrn_back:
            continue
        rrn_raw = f"{rrn_front}{rrn_back}"
        rrn_hash = hashlib.sha256(str(rrn_raw).encode("utf-8")).hexdigest()
        phone1, phone2, phone3 = vals[18], vals[19], vals[20]
        digits = "".join(
            ch
            for part in (phone1, phone2, phone3)
            if part is not None
            for ch in str(part)
            if ch.isdigit()
        )
        phone_mobile = digits if len(digits) >= 9 else None
        out.append(
            {
                "rrn_hash": rrn_hash,
                "name": vals[0],
                "phone_mobile": phone_mobile,
                "nationality": vals[13],
                "department_code": vals[4],
                "position_code": vals[5],
            }
        )
    return out


def main() -> None:
    base = Path("d:/besma-rev/docs/sample/site_import/raw")
    baseline = base / "employees_raw.xlsx"
    v2 = base / "employees_raw_v2.xlsx"

    rows_base = load_rows(baseline)
    rows_v2 = load_rows(v2)
    map_base = {r["rrn_hash"]: r for r in rows_base}

    changed: list[tuple[str, dict]] = []
    for r in rows_v2:
        b = map_base.get(r["rrn_hash"])
        if not b:
            continue
        diffs = {
            k: (b.get(k), r.get(k))
            for k in [
                "name",
                "phone_mobile",
                "nationality",
                "department_code",
                "position_code",
            ]
            if b.get(k) != r.get(k)
        }
        if diffs:
            changed.append((r["rrn_hash"], diffs))

    print("BASE", len(rows_base))
    print("V2", len(rows_v2))
    print("CHANGED", len(changed))
    if changed:
        print("EXAMPLE", changed[0][1])


if __name__ == "__main__":
    main()

