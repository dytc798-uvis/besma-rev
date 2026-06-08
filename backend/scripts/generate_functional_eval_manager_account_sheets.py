"""기능인제 현장소장 로그인 계정표 xlsx 생성 (관리용·배포용).

규칙 ([DECISION-087]): 직종=1 인 소장 후보 중 연장자 1명 / ID=소속현장코드 / PW=주민번호 앞 6자리(YYMMDD)

Usage (from backend/):
  PYTHONPATH=. python scripts/generate_functional_eval_manager_account_sheets.py
  PYTHONPATH=. python scripts/generate_functional_eval_manager_account_sheets.py ../docs/일용직사원리스트_20260529135528.xlsx
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

BACKEND_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_ROOT.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.modules.functional_eval.roster import parse_daily_roster_xlsx  # noqa: E402
from app.modules.functional_eval.service import _birth_sort_key, _rrn_front_password  # noqa: E402


@dataclass
class SiteManagerAccount:
    site_code: str
    site_label: str
    manager_name: str
    login_id: str
    password: str


def _default_roster_path() -> Path:
    docs = REPO_ROOT / "docs"
    candidates = sorted(docs.glob("일용직사원리스트_*.xlsx"), reverse=True)
    for path in candidates:
        if not path.name.startswith("~$"):
            return path
    raise FileNotFoundError(f"일용직 명부 xlsx not found under {docs}")


def _load_site_name_map() -> dict[str, str]:
    names: dict[str, str] = {}
    try:
        from app.config.settings import settings
        from app.core.database import SessionLocal, init_db
        from app.modules.functional_eval import models as fe_models  # noqa: F401
        from app.modules.sites.models import Site
        from app.modules.users import models as user_models  # noqa: F401

        init_db()
        db = SessionLocal()
        try:
            for site in db.query(Site).all():
                if site.site_code:
                    names[str(site.site_code).strip()] = (site.site_name or "").strip()
        finally:
            db.close()
    except Exception:
        pass
    return names


def _site_label(site_code: str, site_names: dict[str, str]) -> str:
    name = (site_names.get(site_code) or "").strip()
    if name and name != f"현장 {site_code}":
        return name
    return site_code


def collect_site_managers(roster_path: Path) -> list[SiteManagerAccount]:
    parsed = parse_daily_roster_xlsx(roster_path)
    site_names = _load_site_name_map()
    managers_by_site: dict[str, list[tuple[tuple[int, int, int], str, str]]] = {}

    for row in parsed:
        if not row.is_site_manager:
            continue
        managers_by_site.setdefault(row.site_code, []).append(
            (_birth_sort_key(row.rrn_raw), row.name, row.rrn_raw)
        )

    accounts: list[SiteManagerAccount] = []
    for site_code in sorted(managers_by_site.keys(), key=lambda c: (len(c), c)):
        managers = managers_by_site[site_code]
        managers.sort(key=lambda x: x[0])
        manager_name = managers[0][1]
        rrn_raw = managers[0][2]
        password = _rrn_front_password(rrn_raw) or ""
        if not password:
            continue
        accounts.append(
            SiteManagerAccount(
                site_code=site_code,
                site_label=_site_label(site_code, site_names),
                manager_name=manager_name,
                login_id=site_code,
                password=password,
            )
        )
    return accounts


def _write_sheet(ws, headers: list[str], rows: list[list[str]], *, password_col: bool) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for account in rows:
        ws.append(account)
    if password_col and ws.max_row > 1:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=4).number_format = "@"


def write_workbooks(accounts: list[SiteManagerAccount], out_dir: Path) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    admin_path = out_dir / f"기능인제_소장계정_관리용_{stamp}.xlsx"
    deploy_path = out_dir / f"기능인제_소장계정_배포용_{stamp}.xlsx"

    wb_admin = openpyxl.Workbook()
    ws_admin = wb_admin.active
    ws_admin.title = "소장계정_관리용"
    _write_sheet(
        ws_admin,
        ["현장", "소장이름", "ID", "PW"],
        [[a.site_label, a.manager_name, a.login_id, a.password] for a in accounts],
        password_col=True,
    )
    for col in ("A", "B", "C", "D"):
        ws_admin.column_dimensions[col].width = 22
    wb_admin.save(admin_path)

    wb_deploy = openpyxl.Workbook()
    ws_deploy = wb_deploy.active
    ws_deploy.title = "소장계정_배포용"
    _write_sheet(
        ws_deploy,
        ["현장", "소장", "ID"],
        [[a.site_label, a.manager_name, a.login_id] for a in accounts],
        password_col=False,
    )
    for col in ("A", "B", "C"):
        ws_deploy.column_dimensions[col].width = 24
    wb_deploy.save(deploy_path)

    return admin_path, deploy_path


def main() -> None:
    roster = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else _default_roster_path()
    out_dir = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else (REPO_ROOT / "docs")
    if not roster.is_file():
        print(f"roster not found: {roster}", file=sys.stderr)
        raise SystemExit(1)

    accounts = collect_site_managers(roster)
    if not accounts:
        print("no site managers (직종=1) found in roster", file=sys.stderr)
        raise SystemExit(1)

    admin_path, deploy_path = write_workbooks(accounts, out_dir)
    print(f"roster: {roster}")
    print(f"sites: {len(accounts)}")
    print(f"admin:  {admin_path}")
    print(f"deploy: {deploy_path}")


if __name__ == "__main__":
    main()
