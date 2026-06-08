"""신규현장 배포 — 현장소장 로그인 계정표 xlsx 생성 (관리용·배포용).

규칙: ID = {현장별칭}-{소장이름}, 초기 PW = 1111 (must_change_password=True)

Usage (from backend/):
  PYTHONPATH=. python scripts/generate_new_site_deployment_manager_account_sheets.py
  PYTHONPATH=. python scripts/generate_new_site_deployment_manager_account_sheets.py ../docs/new-site-deployment/현장소장계정
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

BACKEND_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_ROOT.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.core.database import SessionLocal, init_db  # noqa: E402
from app.modules.functional_eval.site_alias import build_eval_login_id  # noqa: E402
from app.modules.new_site_deployment import models as nsd_models  # noqa: F401, E402
from app.modules.new_site_deployment.deployment_alias import derive_deployment_site_alias  # noqa: E402
from app.modules.new_site_deployment.models import (  # noqa: E402
    NewSiteDeployment,
    NewSiteDeploymentAdministrator,
)
from app.modules.new_site_deployment.service import INITIAL_SITE_PASSWORD  # noqa: E402
from app.modules.sites.models import Site  # noqa: E402
from app.modules.users import models as user_models  # noqa: F401, E402

INITIAL_PASSWORD = INITIAL_SITE_PASSWORD


@dataclass
class SiteManagerAccount:
    site_code: str
    site_name: str
    contractor: str
    site_alias: str
    manager_name: str
    login_id: str
    password: str


def collect_site_managers() -> list[SiteManagerAccount]:
    init_db()
    db = SessionLocal()
    accounts: list[SiteManagerAccount] = []
    try:
        rows = (
            db.query(NewSiteDeploymentAdministrator, NewSiteDeployment)
            .join(NewSiteDeployment, NewSiteDeploymentAdministrator.deployment_id == NewSiteDeployment.id)
            .filter(NewSiteDeploymentAdministrator.role == "SITE_MANAGER")
            .order_by(NewSiteDeployment.site_name.asc(), NewSiteDeploymentAdministrator.sort_order.asc())
            .all()
        )
        for admin, dep in rows:
            login_id = (admin.login_id or "").strip()
            if not login_id:
                continue
            accounts.append(
                SiteManagerAccount(
                    site_code=(dep.site_code or "").strip(),
                    site_name=(dep.site_name or "").strip(),
                    contractor=(dep.contractor or "").strip(),
                    site_alias=(dep.site_alias or "").strip(),
                    manager_name=(admin.name or "").strip(),
                    login_id=login_id,
                    password=INITIAL_PASSWORD,
                )
            )

        if accounts:
            return accounts

        # administrators 테이블 이전·미동기화 대비 — legacy mirror 필드
        legacy_rows = (
            db.query(NewSiteDeployment)
            .filter(NewSiteDeployment.site_manager_name.isnot(None))
            .filter(NewSiteDeployment.site_manager_login_id.isnot(None))
            .order_by(NewSiteDeployment.site_name.asc())
            .all()
        )
        for dep in legacy_rows:
            login_id = (dep.site_manager_login_id or "").strip()
            name = (dep.site_manager_name or "").strip()
            if not login_id or not name:
                continue
            accounts.append(
                SiteManagerAccount(
                    site_code=(dep.site_code or "").strip(),
                    site_name=(dep.site_name or "").strip(),
                    contractor=(dep.contractor or "").strip(),
                    site_alias=(dep.site_alias or "").strip(),
                    manager_name=name,
                    login_id=login_id,
                    password=INITIAL_PASSWORD,
                )
            )

        if accounts:
            return accounts

        # 배포 테이블 비어 있을 때 — sites + 별칭 계산으로 표 생성
        sites = (
            db.query(Site)
            .filter(Site.site_manager.isnot(None))
            .filter(Site.site_manager != "")
            .order_by(Site.site_name.asc())
            .all()
        )
        for site in sites:
            site_name = (site.site_name or "").strip()
            manager_name = (site.site_manager or "").strip()
            if not site_name or not manager_name:
                continue
            contractor = (site.contractor_name or "").strip()
            alias = derive_deployment_site_alias(contractor or None, site_name)
            login_id = build_eval_login_id(alias, manager_name)
            if not login_id:
                continue
            accounts.append(
                SiteManagerAccount(
                    site_code=(site.site_code or "").strip(),
                    site_name=site_name,
                    contractor=contractor,
                    site_alias=alias,
                    manager_name=manager_name,
                    login_id=login_id,
                    password=INITIAL_PASSWORD,
                )
            )
    finally:
        db.close()
    return accounts


def _write_sheet(ws, headers: list[str], rows: list[list[str]], *, password_col: bool) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    if password_col and ws.max_row > 1:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=7).number_format = "@"


def write_workbooks(accounts: list[SiteManagerAccount], out_dir: Path) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    admin_path = out_dir / f"신규현장_현장소장계정_관리용_{stamp}.xlsx"
    deploy_path = out_dir / f"신규현장_현장소장계정_배포용_{stamp}.xlsx"

    admin_headers = ["현장코드", "현장명", "도급사", "별칭", "소장이름", "아이디", "비밀번호"]
    deploy_headers = ["현장코드", "현장명", "도급사", "별칭", "소장이름", "아이디"]

    wb_admin = openpyxl.Workbook()
    ws_admin = wb_admin.active
    ws_admin.title = "현장소장_관리용"
    _write_sheet(
        ws_admin,
        admin_headers,
        [
            [
                a.site_code,
                a.site_name,
                a.contractor,
                a.site_alias,
                a.manager_name,
                a.login_id,
                a.password,
            ]
            for a in accounts
        ],
        password_col=True,
    )
    for col, width in zip("ABCDEFG", [10, 36, 16, 14, 12, 22, 10], strict=False):
        ws_admin.column_dimensions[col].width = width
    wb_admin.save(admin_path)

    wb_deploy = openpyxl.Workbook()
    ws_deploy = wb_deploy.active
    ws_deploy.title = "현장소장_배포용"
    _write_sheet(
        ws_deploy,
        deploy_headers,
        [
            [a.site_code, a.site_name, a.contractor, a.site_alias, a.manager_name, a.login_id]
            for a in accounts
        ],
        password_col=False,
    )
    for col, width in zip("ABCDEF", [10, 36, 16, 14, 12, 22], strict=False):
        ws_deploy.column_dimensions[col].width = width
    wb_deploy.save(deploy_path)

    return admin_path, deploy_path


def main() -> None:
    out_dir = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else (REPO_ROOT / "docs" / "new-site-deployment" / "현장소장계정")
    accounts = collect_site_managers()
    admin_path, deploy_path = write_workbooks(accounts, out_dir)
    print(f"sites: {len(accounts)}")
    print(f"admin:  {admin_path}")
    print(f"deploy: {deploy_path}")
    if not accounts:
        print("note: 등록된 신규현장 소장 계정이 없습니다. 현장 등록 후 다시 실행하세요.")


if __name__ == "__main__":
    main()
