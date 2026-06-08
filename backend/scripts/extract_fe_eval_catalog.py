"""Extract 2-1/2-2 criteria from docs xlsx → JSON. Run once: python scripts/extract_fe_eval_catalog.py"""
from __future__ import annotations

import json
import sys
from pathlib import Path

BACKEND_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_ROOT))

import openpyxl

DOCS = BACKEND_ROOT.parent / "docs" / "01. 현장별 기능인등급-25년12월최초.xlsx"
OUT = BACKEND_ROOT / "app" / "modules" / "functional_eval" / "eval_catalog_data.json"

GRADE_KEYS = ("TOP", "MID", "LOW", "BOTTOM")
POINT_SET = {3, 4, 6, 7, 8, 9, 10, 11, 12, 15}


def _is_points(v) -> bool:
    return isinstance(v, (int, float)) and int(v) in POINT_SET


def parse_sheet(ws) -> list[dict]:
    rows = list(ws.iter_rows(min_row=3, max_row=5, values_only=True))
    if len(rows) < 3:
        return []
    r_title, r_grade_label, r_points = rows[0], rows[1], rows[2]
    n = max(len(r_title or []), len(r_grade_label or []), len(r_points or []))
    criteria: list[dict] = []
    i = 7
    while i < n:
        title = r_title[i] if r_title and i < len(r_title) else None
        if not title or not str(title).strip():
            i += 1
            continue
        grades: list[dict] = []
        j = i
        while j < n and _is_points(r_points[j] if r_points and j < len(r_points) else None):
            gl = r_grade_label[j] if r_grade_label and j < len(r_grade_label) else ""
            gl = str(gl).strip() if gl is not None else ""
            pt = int(r_points[j])
            key = GRADE_KEYS[len(grades)] if len(grades) < 4 else f"G{len(grades)}"
            grades.append({"key": key, "label": gl or key, "points": pt})
            j += 1
            if len(grades) >= 4:
                break
        if len(grades) == 4:
            criteria.append({"id": f"c{len(criteria) + 1}", "title": str(title).strip(), "grades": grades})
        i = j if j > i else i + 1
    return criteria


def main() -> None:
    wb = openpyxl.load_workbook(DOCS, read_only=True, data_only=True)
    functional_name = next(n for n in wb.sheetnames if "2-1" in n)
    safety_name = next(n for n in wb.sheetnames if "2-2" in n)
    payload = {
        "FUNCTIONAL": parse_sheet(wb[functional_name]),
        "SAFETY": parse_sheet(wb[safety_name]),
    }
    wb.close()
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", OUT, "functional", len(payload["FUNCTIONAL"]), "safety", len(payload["SAFETY"]))


if __name__ == "__main__":
    main()
