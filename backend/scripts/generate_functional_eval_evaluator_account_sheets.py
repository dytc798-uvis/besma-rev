"""기능인제 평가자(소장·팀장) 아이디/비밀번호 표 xlsx 생성.

데이터:
  - docs/월별현장별집계_*.xls (현장명·별칭·대표)
  - docs/출역일보_*.xls (팀장·당일 출역)
  - docs/sample/site_import/raw/* (일용직 사원리스트 — 소장·주민번호 우선)

Usage (from backend/):
  PYTHONPATH=. python scripts/generate_functional_eval_evaluator_account_sheets.py
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

BACKEND_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_ROOT.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.core.database import SessionLocal, init_db  # noqa: E402
from app.modules.functional_eval import models as fe_models  # noqa: F401, E402
from app.modules.functional_eval.attendance import parse_attendance_report  # noqa: E402
from app.modules.functional_eval.constants import TEAM_LEADER_SPLIT_THRESHOLD  # noqa: E402
from app.modules.functional_eval.eval_provisioning import (  # noqa: E402
    _ensure_unique_aliases,
    _rrn_front_password,
    normalize_erp_site_label,
)
from app.modules.functional_eval.roster import (  # noqa: E402
    ParsedRosterRow,
    parse_daily_roster,
    parse_employee_master,
)
from app.modules.functional_eval.site_aggregate import parse_monthly_site_aggregate  # noqa: E402
from app.modules.functional_eval.site_alias import build_eval_login_id  # noqa: E402
from app.modules.functional_eval.service import _birth_sort_key  # noqa: E402
from app.modules.sites.models import Site  # noqa: E402
from app.modules.users import models as user_models  # noqa: F401, E402

NON_PERSON_REP_LABELS = frozenset(
    {
        "직영",
        "외주",
        "합계",
        "소계",
        "미배정",
        "없음",
        "공무",
        "소장",
        "팀장",
        "대표",
        "미지정",
        "해당없음",
        "없",
    }
)


@dataclass
class EvaluatorAccountRow:
    site_code: str
    site_name: str
    site_alias: str
    role: str
    name: str
    login_id: str
    password: str
    note: str = ""


def _latest_doc(pattern: str) -> Path:
    docs = REPO_ROOT / "docs"
    candidates = sorted(docs.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    for path in candidates:
        if not path.name.startswith("~$"):
            return path
    raise FileNotFoundError(f"{pattern} not found under {docs}")


def _find_roster_sources() -> tuple[Path | None, Path | None]:
    raw = REPO_ROOT / "docs" / "sample" / "site_import" / "raw"
    employee_candidates = sorted(
        {
            p
            for pattern in ("*사원리스트*.xls", "*사원리스트*.xlsx", "사원리스트*.xls")
            for p in raw.glob(pattern)
            if p.is_file() and not p.name.startswith("~$")
        },
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    employee_path = employee_candidates[0] if employee_candidates else None

    daily_candidates = sorted(
        {
            p
            for pattern in (
                "daily_workers_raw.xls.normalized.xlsx",
                "daily_workers_raw.xls",
                "일용직사원리스트*.xls",
                "일용직*.xls",
            )
            for p in ([raw / pattern] if "*" not in pattern else raw.glob(pattern))
            if p.is_file() and not p.name.startswith("~$")
        },
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    daily_path = daily_candidates[0] if daily_candidates else None
    return daily_path, employee_path


def _load_site_name_map() -> dict[str, str]:
    names: dict[str, str] = {}
    try:
        init_db()
        db = SessionLocal()
        try:
            for site in db.query(Site).all():
                if site.site_code:
                    names[str(site.site_code).strip()] = (site.site_name or "").strip()
        finally:
            db.close()
    except Exception:
        pass
    return names


def _roster_managers_by_site(rows: list[ParsedRosterRow]) -> dict[str, tuple[str, str]]:
    """직종=1 소장 후보 중 연장자 1명 → (name, rrn_raw)."""
    managers: dict[str, list[tuple[tuple[int, int, int], str, str]]] = defaultdict(list)
    for row in rows:
        if not row.is_site_manager:
            continue
        managers[row.site_code].append((_birth_sort_key(row.rrn_raw), row.name.strip(), row.rrn_raw))
    out: dict[str, tuple[str, str]] = {}
    for site_code, items in managers.items():
        items.sort(key=lambda x: x[0])
        out[site_code] = (items[0][1], items[0][2])
    return out


def _roster_name_rrn_by_site(rows: list[ParsedRosterRow]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = defaultdict(dict)
    for row in rows:
        name = row.name.strip()
        if name:
            out[row.site_code][name] = row.rrn_raw
    return out


def is_person_name(name: str) -> bool:
    text = (name or "").strip().replace(" ", "")
    if not text or text in NON_PERSON_REP_LABELS:
        return False
    if not re.fullmatch(r"[가-힣]{2,4}", text):
        return False
    return True


def _global_name_rrn(employee_path: Path | None) -> dict[str, str]:
    if employee_path is None or not employee_path.is_file():
        return {}
    out: dict[str, str] = {}
    for row in parse_employee_master(employee_path):
        out[row.name.strip()] = row.rrn_raw
    return out


def collect_evaluator_accounts(
    aggregate_path: Path,
    attendance_path: Path,
    daily_roster_path: Path | None,
    employee_master_path: Path | None,
) -> list[EvaluatorAccountRow]:
    agg_rows = parse_monthly_site_aggregate(aggregate_path)
    aliases = _ensure_unique_aliases(agg_rows)
    att_rows = parse_attendance_report(attendance_path)
    roster_rows: list[ParsedRosterRow] = []
    if daily_roster_path and daily_roster_path.is_file():
        roster_rows = parse_daily_roster(daily_roster_path)
    global_rrn = _global_name_rrn(employee_master_path)
    site_names = _load_site_name_map()
    roster_mgr = _roster_managers_by_site(roster_rows)
    roster_rrn = _roster_name_rrn_by_site(roster_rows)

    code_to_agg = {r.site_code: r for r in agg_rows}
    label_to_code: dict[str, str] = {}
    for row in agg_rows:
        label_to_code[normalize_erp_site_label(row.erp_site_name)] = row.site_code
        label_to_code[row.erp_site_name.strip()] = row.site_code

    site_att: dict[str, list] = defaultdict(list)
    for row in att_rows:
        key = normalize_erp_site_label(row.erp_site_label)
        code = label_to_code.get(key) or label_to_code.get((row.erp_site_label or "").strip())
        if code:
            site_att[code].append(row)

    out: list[EvaluatorAccountRow] = []
    seen_logins: set[tuple[str, str]] = set()

    def _append(row: EvaluatorAccountRow) -> None:
        key = (row.site_code, row.login_id)
        if row.login_id and key in seen_logins:
            return
        if row.login_id:
            seen_logins.add(key)
        out.append(row)

    for site_code in sorted(code_to_agg.keys(), key=lambda c: (len(c), c)):
        agg = code_to_agg[site_code]
        site_name = agg.erp_site_name.strip() or site_names.get(site_code, site_code)
        site_alias = aliases[site_code]
        rows = site_att.get(site_code, [])

        if site_code in roster_mgr:
            manager_name, manager_rrn = roster_mgr[site_code]
        else:
            manager_name = agg.manager_name.strip()
            manager_rrn = ""

        name_rrn: dict[str, str] = {}
        for r in rows:
            name_rrn[r.name.strip()] = r.rrn_raw
        for name, rrn in roster_rrn.get(site_code, {}).items():
            name_rrn.setdefault(name, rrn)
        for name, rrn in global_rrn.items():
            name_rrn.setdefault(name, rrn)

        if not manager_rrn:
            manager_rrn = name_rrn.get(manager_name, "")
        if not manager_rrn:
            for r in rows:
                if (r.job_name or "").strip() == "소장":
                    manager_rrn = r.rrn_raw
                    break

        manager_login = build_eval_login_id(site_alias, manager_name)
        manager_pw = _rrn_front_password(manager_rrn or "") or ""
        _append(
            EvaluatorAccountRow(
                site_code=site_code,
                site_name=site_name,
                site_alias=site_alias,
                role="소장",
                name=manager_name,
                login_id=manager_login if manager_login and manager_pw else "",
                password=manager_pw,
                note="" if manager_pw else "사원리스트·출역일보에 주민번호 없음",
            )
        )

        workers_today = [
            r
            for r in rows
            if (r.job_name or "").strip() != "소장" and r.name.strip() != manager_name
        ]
        if len(workers_today) <= TEAM_LEADER_SPLIT_THRESHOLD:
            continue

        by_rep: dict[str, list] = defaultdict(list)
        for row in workers_today:
            rep = (row.rep_name or "").strip() or manager_name
            by_rep[rep].append(row)

        for rep_name, team_rows in sorted(by_rep.items()):
            if rep_name == manager_name:
                continue
            if not is_person_name(rep_name):
                continue

            attendance_rrn = {r.name.strip(): r.rrn_raw for r in rows}
            rep_attended = rep_name in attendance_rrn
            rep_login = build_eval_login_id(site_alias, rep_name)

            if rep_attended:
                rep_rrn = attendance_rrn[rep_name]
                rep_pw = _rrn_front_password(rep_rrn) or ""
                login_id = rep_login if rep_login and rep_pw else ""
                password = rep_pw
                note = "" if login_id else "출역일보 주민번호 없음"
            else:
                login_id = ""
                password = ""
                note = f"출역하지 않는 팀장(팀원 {len(team_rows)}명 → 소장 평가)"

            _append(
                EvaluatorAccountRow(
                    site_code=site_code,
                    site_name=site_name,
                    site_alias=site_alias,
                    role="팀장",
                    name=rep_name,
                    login_id=login_id,
                    password=password,
                    note=note,
                )
            )

    return out


def _write_sheet(ws, headers: list[str], rows: list[list[str]], *, password_col: int | None) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    if password_col and ws.max_row > 1:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=password_col).number_format = "@"


def write_workbooks(accounts: list[EvaluatorAccountRow], out_dir: Path) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    admin_path = out_dir / f"기능인제_평가자계정_관리용_{stamp}.xlsx"
    deploy_path = out_dir / f"기능인제_평가자계정_배포용_{stamp}.xlsx"

    headers_admin = ["현장코드", "현장명", "별칭", "구분", "이름", "아이디", "비밀번호", "비고"]
    headers_deploy = ["현장코드", "현장명", "별칭", "구분", "이름", "아이디", "비고"]

    def _row(a: EvaluatorAccountRow, *, with_password: bool) -> list[str]:
        base = [a.site_code, a.site_name, a.site_alias, a.role, a.name, a.login_id]
        if with_password:
            return base + [a.password, a.note]
        return base + [a.note]

    wb_admin = openpyxl.Workbook()
    ws_admin = wb_admin.active
    ws_admin.title = "평가자_관리용"
    _write_sheet(
        ws_admin,
        headers_admin,
        [_row(a, with_password=True) for a in accounts],
        password_col=7,
    )
    for col, width in zip("ABCDEFGH", [10, 38, 14, 8, 12, 22, 10, 18], strict=False):
        ws_admin.column_dimensions[col].width = width
    wb_admin.save(admin_path)

    wb_deploy = openpyxl.Workbook()
    ws_deploy = wb_deploy.active
    ws_deploy.title = "평가자_배포용"
    _write_sheet(
        ws_deploy,
        headers_deploy,
        [_row(a, with_password=False) for a in accounts],
        password_col=None,
    )
    for col, width in zip("ABCDEFG", [10, 38, 14, 8, 12, 22, 18], strict=False):
        ws_deploy.column_dimensions[col].width = width
    wb_deploy.save(deploy_path)

    return admin_path, deploy_path


def main() -> None:
    out_dir = (
        Path(sys.argv[1]).resolve()
        if len(sys.argv) > 1
        else (REPO_ROOT / "docs" / "기능인제_평가자계정")
    )
    agg = _latest_doc("월별현장별집계_*.xls")
    att = _latest_doc("출역일보_*.xls")
    daily_roster, employee_master = _find_roster_sources()
    if daily_roster is None and employee_master is None:
        raise FileNotFoundError("일용직/사원리스트 not found under docs/sample/site_import/raw")
    accounts = collect_evaluator_accounts(agg, att, daily_roster, employee_master)
    if not accounts:
        print("no evaluator accounts", file=sys.stderr)
        raise SystemExit(1)

    admin_path, deploy_path = write_workbooks(accounts, out_dir)
    mgr = sum(1 for a in accounts if a.role == "소장")
    lead = sum(1 for a in accounts if a.role == "팀장")
    filled = sum(1 for a in accounts if a.login_id)
    print(f"aggregate: {agg.name}")
    print(f"attendance: {att.name}")
    print(f"daily_roster: {daily_roster.name if daily_roster else '-'}")
    print(f"employee_master: {employee_master.name if employee_master else '-'}")
    print(f"rows: {len(accounts)} (소장 {mgr}, 팀장 {lead}, ID부여 {filled})")
    print(f"admin:  {admin_path}")
    print(f"deploy: {deploy_path}")


if __name__ == "__main__":
    main()
