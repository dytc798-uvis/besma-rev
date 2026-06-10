"""기능인제·삼성인정제 본사(한글 ID) 계정 xlsx 생성.

데이터: create_hq_safe_accounts.HQ_SAFE_ACCOUNTS, create_ceo_eval_account (대표 승인)

Usage (from backend/):
  PYTHONPATH=. python scripts/generate_hq_account_sheets.py
  PYTHONPATH=. python scripts/generate_hq_account_sheets.py ../docs/기능인제_평가자계정
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

BACKEND_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_ROOT.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.modules.functional_eval.constants import CEO_EVAL_LOGIN_IDS  # noqa: E402
from app.modules.users.hq_safe_accounts import HQ_SAFE_ACCOUNT_SPECS  # noqa: E402
from scripts.create_ceo_eval_account import CEO_NAME, CEO_PASSWORD  # noqa: E402


@dataclass
class HqAccountRow:
    department: str
    role_label: str
    title: str
    name: str
    login_id: str
    password: str
    note: str = ""


def collect_hq_web_accounts() -> list[HqAccountRow]:
    """웹 로그인 본사 계정(한글 ID) — 기능인제 HQ·삼성인정제·대표 승인."""
    rows: list[HqAccountRow] = []
    for spec in HQ_SAFE_ACCOUNT_SPECS:
        note = "기능인제 본사 / 삼성인정제·문서"
        if not spec.fe_samsung_web:
            note = "본사(기능인제·삼성인정제 미사용)"
        rows.append(
            HqAccountRow(
                department="안전보건실",
                role_label="본사",
                title=spec.title,
                name=spec.name,
                login_id=spec.login_id,
                password=spec.password,
                note=note,
            )
        )
    ceo_login = next(iter(CEO_EVAL_LOGIN_IDS))
    rows.append(
        HqAccountRow(
            department="대표이사",
            role_label="최종승인",
            title="대표이사",
            name=CEO_NAME,
            login_id=ceo_login,
            password=CEO_PASSWORD,
            note="기능인제 대표 최종승인",
        )
    )
    return rows


def _write_sheet(ws, headers: list[str], rows: list[list[str]], *, password_col: int | None) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    if password_col and ws.max_row > 1:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=password_col).number_format = "@"


def write_hq_workbooks(accounts: list[HqAccountRow], out_dir: Path) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    admin_path = out_dir / f"기능인제_본사계정_관리용_{stamp}.xlsx"
    deploy_path = out_dir / f"기능인제_본사계정_배포용_{stamp}.xlsx"

    headers_admin = ["부서", "구분", "직급", "이름", "아이디", "비밀번호", "비고"]
    headers_deploy = ["부서", "구분", "직급", "이름", "아이디", "비고"]

    def _row(a: HqAccountRow, *, with_password: bool) -> list[str]:
        base = [a.department, a.role_label, a.title, a.name, a.login_id]
        if with_password:
            return base + [a.password, a.note]
        return base + [a.note]

    wb_admin = openpyxl.Workbook()
    ws_admin = wb_admin.active
    ws_admin.title = "본사_관리용"
    _write_sheet(
        ws_admin,
        headers_admin,
        [_row(a, with_password=True) for a in accounts],
        password_col=6,
    )
    for col, width in zip("ABCDEFG", [14, 10, 8, 12, 22, 10, 28], strict=False):
        ws_admin.column_dimensions[col].width = width
    wb_admin.save(admin_path)

    wb_deploy = openpyxl.Workbook()
    ws_deploy = wb_deploy.active
    ws_deploy.title = "본사_배포용"
    _write_sheet(
        ws_deploy,
        headers_deploy,
        [_row(a, with_password=False) for a in accounts],
        password_col=None,
    )
    for col, width in zip("ABCDEF", [14, 10, 8, 12, 22, 28], strict=False):
        ws_deploy.column_dimensions[col].width = width
    wb_deploy.save(deploy_path)

    return admin_path, deploy_path


def main() -> None:
    out_dir = (
        Path(sys.argv[1]).resolve()
        if len(sys.argv) > 1
        else (REPO_ROOT / "docs" / "기능인제_평가자계정")
    )
    accounts = collect_hq_web_accounts()
    admin_path, deploy_path = write_hq_workbooks(accounts, out_dir)
    print(f"rows: {len(accounts)} (안전보건 {len(HQ_SAFE_ACCOUNT_SPECS)}, 대표 1)")
    print(f"admin:  {admin_path}")
    print(f"deploy: {deploy_path}")


if __name__ == "__main__":
    main()
