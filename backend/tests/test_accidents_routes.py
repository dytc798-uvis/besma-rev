# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
import tempfile

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.auth import get_current_user_with_bypass, get_db
from app.core.database import Base
from app.core.enums import Role, UIType
from app.modules.accidents.routes import router as accidents_router
from app.modules.accidents import service as accident_service

TEMPLATE_MESSAGE = """[사고최초보고]
현  장  명: 테스트현장
보  고  자: 보고자
사고시각: 2026-04-19 10:00
사고장소: A구역
작업내용: 용접 작업
재  해  자: 김철수
사고경위: 용접 작업 중 넘어짐
사고원인: 바닥 정리 미흡
상해부위: 무릎
조치사항: 현장 응급조치 후 보고
위와 같이 보고드립니다."""


def _setup_db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    from app.modules.sites import models as site_models  # noqa: F401
    from app.modules.workers import models as worker_models  # noqa: F401
    from app.modules.users import models as user_models  # noqa: F401
    from app.modules.accidents import models as accident_models  # noqa: F401

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    db.add(
        user_models.User(
            name="HQ",
            login_id="hq_test",
            password_hash="x",
            role=Role.HQ_SAFE,
            ui_type=UIType.HQ_SAFE,
            must_change_password=False,
            password_changed_at=datetime.utcnow(),
        )
    )
    db.commit()
    return db


def _build_client(db):
    app = FastAPI()
    app.include_router(accidents_router)
    app.dependency_overrides[get_current_user_with_bypass] = lambda: SimpleNamespace(
        id=1,
        role=Role.HQ_SAFE,
        ui_type="HQ_SAFE",
    )

    def override_get_db():
        yield db

    app.dependency_overrides[get_db] = override_get_db
    return TestClient(app)


def test_parse_and_create_and_list():
    db = _setup_db()
    client = _build_client(db)
    msg = TEMPLATE_MESSAGE
    r = client.post("/accidents/initial-report/parse-and-create", json={"message_raw": msg})
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["parse_status"] == "success"
    assert body["display_code"].startswith("ACC-")
    assert body["nas_folder_key"] == "0419 김철수 테스트현장"
    nas_parts = Path(body["nas_folder_path"]).parts
    assert nas_parts[-2] == "2026"
    assert nas_parts[-1] == "0419 김철수 테스트현장"
    aid = body["id"]

    r2 = client.get("/accidents", params={"show_all": True})
    assert r2.status_code == 200
    assert len(r2.json()) == 1

    r3 = client.get(f"/accidents/{aid}")
    assert r3.status_code == 200
    assert r3.json()["message_raw"] == msg

    r4 = client.get(f"/accidents/{aid}/initial-report")
    assert r4.status_code == 200
    assert "composed_line" in r4.json()

    bat = client.get(f"/accidents/{aid}/nas-folder-launcher")
    assert bat.status_code == 200
    assert b"explorer" in bat.content
    assert b"mkdir" in bat.content.lower()
    assert "attachment" in (bat.headers.get("content-disposition") or "").lower()


def test_manual_create_and_worklist():
    db = _setup_db()
    client = _build_client(db)
    body = {
        "input_mode": "manual",
        "site_standard_name": "수동입력현장",
        "reporter_name": "홍길동",
        "accident_datetime_text": "2026-04-19 14:00",
        "accident_place": "지상 2층",
        "work_content": "자재 정리",
        "injured_person_name": "김철수",
        "accident_circumstance": "자재 이동 중",
        "injured_part": "우측 손목",
        "status": "신규",
        "management_category": "일반",
    }
    r = client.post("/accidents/initial-report/parse-and-create", json=body)
    assert r.status_code == 201, r.text
    created = r.json()
    assert "verification_status" not in created
    assert created["parse_status"] == "partial"

    worklist = client.get("/accidents/worklist")
    assert worklist.status_code == 200
    payload = worklist.json()
    assert payload["unverified"]["count"] == 0
    assert payload["recent"]["count"] >= 1

    updated = client.put(
        f"/accidents/{created['id']}",
        json={
            "site_standard_name": "수동입력현장",
            "reporter_name": "홍길동",
            "status": "접수",
            "management_category": "일반",
            "accident_datetime_text": "2026-04-19 14:00",
            "accident_datetime": None,
            "accident_place": "지상 2층",
            "work_content": "자재 정리 보완",
            "injured_person_name": "김철수",
            "accident_circumstance": "자재 이동 중",
            "accident_reason": None,
            "injured_part": "우측 손목",
            "diagnosis_name": None,
            "notes": "수정",
            "initial_report_template": None,
        },
    )
    assert updated.status_code == 200
    assert "verification_status" not in updated.json()


def test_parse_preview_and_manual_supplement_keeps_partial_trace():
    db = _setup_db()
    client = _build_client(db)
    msg = """[사고최초보고]
현  장  명: 테스트현장
보  고  자: 보고자
사고시각: 2026-04-19 10:00
사고장소: A구역
작업내용: 용접 작업
재  해  자: 김철수
사고경위: 용접 작업 중 넘어짐
상해부위: 무릎
조치사항: 현장 응급조치 후 보고
위와 같이 보고드립니다."""

    preview = client.post("/accidents/initial-report/parse-preview", json={"message_raw": msg})
    assert preview.status_code == 200, preview.text
    preview_body = preview.json()
    assert preview_body["parse_status"] == "partial"

    created = client.post(
        "/accidents/initial-report/parse-and-create",
        json={
            "input_mode": "manual",
            "message_raw": msg,
            "site_standard_name": "테스트현장",
            "reporter_name": "보고자",
            "accident_datetime_text": "2026-04-19 10:00",
            "accident_place": "A구역",
            "work_content": "용접 작업",
            "injured_person_name": "김철수",
            "accident_circumstance": "이동 중 넘어짐",
            "accident_reason": "바닥 정리 미흡",
            "injured_part": "무릎",
            "action_taken": "현장 응급조치 후 보고",
            "status": "신규",
            "management_category": "일반",
            "parse_status_override": preview_body["parse_status"],
            "parse_note_override": preview_body["parse_note"],
        },
    )
    assert created.status_code == 201, created.text
    created_body = created.json()
    assert created_body["parse_status"] == "partial"
    assert created_body["action_taken"] == "현장 응급조치 후 보고"

    note = created_body["parse_note"]
    assert note is not None
    assert "manual_supplemented" in note


def test_export_to_master_excel_writes_all_rows():
    db = _setup_db()
    client = _build_client(db)

    verified = client.post(
        "/accidents/initial-report/parse-and-create",
        json={
            "message_raw": """[사고최초보고]
현  장  명: 검증현장
보  고  자: 홍길동
사고시각: 2026-04-19 09:00
사고장소: 지상 1층
작업내용: 자재 이동
재  해  자: 김철수
사고경위: 이동 중 전도
사고원인: 바닥 정리 미흡
상해부위: 우측 손목
조치사항: 응급조치
위와 같이 보고드립니다."""
        },
    )
    assert verified.status_code == 201, verified.text
    verified_id = verified.json()["id"]

    unverified = client.post(
        "/accidents/initial-report/parse-and-create",
        json={
            "input_mode": "manual",
            "site_standard_name": "미검증현장",
            "reporter_name": "이몽룡",
            "accident_datetime_text": "2026-04-20 09:00",
            "accident_datetime": "2026-04-20T09:00:00",
            "accident_place": "지상 2층",
            "work_content": "정리 작업",
            "injured_person_name": "성춘향",
            "accident_circumstance": "정리 중 충돌",
            "accident_reason": "주의 미흡",
            "injured_part": "좌측 팔",
            "status": "신규",
            "management_category": "일반",
        },
    )
    assert unverified.status_code == 201, unverified.text

    complete = client.put(
        f"/accidents/{verified_id}",
        json={
            "site_standard_name": "검증현장",
            "reporter_name": "홍길동",
            "status": "접수",
            "management_category": "일반",
            "accident_datetime_text": "2026-04-19 09:00",
            "accident_datetime": "2026-04-19T09:00:00",
            "accident_place": "지상 1층",
            "work_content": "자재 이동",
            "injured_person_name": "김철수",
            "accident_circumstance": "이동 중 전도",
            "accident_reason": "바닥 정리 미흡",
            "injured_part": "우측 손목",
            "diagnosis_name": None,
            "action_taken": "응급조치",
            "notes": "검증완료",
            "initial_report_template": None,
        },
    )
    assert complete.status_code == 200, complete.text

    with tempfile.TemporaryDirectory() as tmpdir:
        workbook_path = Path(tmpdir) / "BESMA_사고MASTER_2026.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "MASTER_사고한장표"
        headers = accident_service.MASTER_HEADERS + ["폴더존재여부", "파일개수", "증빙상태"]
        ws.append(headers)
        ws.append(["" for _ in headers])
        wb.save(workbook_path)
        wb.close()

        original_path = accident_service.MASTER_IMPORT_PATH
        accident_service.MASTER_IMPORT_PATH = workbook_path
        try:
            result = client.post("/accidents/export-to-master-excel")
            assert result.status_code == 200, result.text
            body = result.json()
            assert body["exported_count"] == 2
            assert body["excluded_count"] == 0
            assert body["backup_created"] is True
            assert Path(body["backup_path"]).is_file()

            saved = load_workbook(workbook_path, data_only=False, read_only=False)
            saved_ws = saved["MASTER_사고한장표"]
            accident_ids = {complete.json()["accident_id"], unverified.json()["accident_id"]}
            assert str(saved_ws.cell(2, 3).value) in accident_ids
            assert str(saved_ws.cell(3, 3).value) in accident_ids
            assert str(saved_ws.cell(2, 3).value) != str(saved_ws.cell(3, 3).value)
            saved.close()
        finally:
            accident_service.MASTER_IMPORT_PATH = original_path


def test_delete_accident_removes_row_and_folder():
    db = _setup_db()
    client = _build_client(db)

    with tempfile.TemporaryDirectory() as tmpdir:
        original_storage_root = accident_service.settings.storage_root
        original_accident_nas_root = accident_service.settings.accident_nas_root
        accident_service.settings.storage_root = Path(tmpdir) / "storage"
        accident_service.settings.accident_nas_root = None
        try:
            created = client.post("/accidents/initial-report/parse-and-create", json={"message_raw": TEMPLATE_MESSAGE})
            assert created.status_code == 201, created.text
            body = created.json()
            folder_path = Path(body["nas_folder_path"])
            assert folder_path.exists()

            uploaded = client.post(
                f"/accidents/{body['id']}/attachments",
                files={"file": ("sample.txt", b"hello", "text/plain")},
            )
            assert uploaded.status_code == 200, uploaded.text
            assert folder_path.exists()
            assert any(folder_path.iterdir())

            deleted = client.delete(f"/accidents/{body['id']}")
            assert deleted.status_code == 204, deleted.text

            missing = client.get(f"/accidents/{body['id']}")
            assert missing.status_code == 404
            assert not folder_path.exists()
        finally:
            accident_service.settings.storage_root = original_storage_root
            accident_service.settings.accident_nas_root = original_accident_nas_root
