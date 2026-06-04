from __future__ import annotations

import openpyxl

from app.modules.work_plan_excel.equipment_lookup import lookup_forklift_equipment_specs
from app.modules.work_plan_excel.service import SHEET_NAME, generate_forklift_work_plan, resolve_forklift_template_path
from app.schemas.work_plan_forklift import ForkliftWorkPlanInput


def test_lookup_catalog_50dn():
    spec = lookup_forklift_equipment_specs("50DN-9VB", allow_web=False)
    assert spec.source == "catalog"
    assert spec.rated_capacity == "5ton"
    assert spec.length_mm == 4510
    assert spec.max_lifting_kg == 11480


def test_lookup_unknown_without_web():
    spec = lookup_forklift_equipment_specs("UNKNOWN-MODEL-XYZ", allow_web=False)
    assert spec.source == "none"


def test_generate_forklift_work_plan_preserves_template_and_fills_cells(tmp_path, monkeypatch):
    template = resolve_forklift_template_path()
    assert template.is_file()

    from app.modules.work_plan_excel import service

    monkeypatch.setattr(service, "_work_plan_output_dir", lambda: tmp_path)

    payload = ForkliftWorkPlanInput(
        site_name="테스트현장",
        company_name="(주)부현전기",
        document_date_year="2026년",
        document_date_month="6월",
        document_date_day="4일",
        work_name="단위테스트 작업",
        work_location="지상2층",
        supervisor_name="홍길동",
        equipment_model="UNIT-TEST",
        rated_capacity="3ton",
        length_mm=4000,
        width_mm=1700,
        height_mm=3000,
    )
    out_path, filename = generate_forklift_work_plan(payload)
    assert out_path.is_file()
    assert filename.endswith("_작업계획서.xlsx")

    wb = openpyxl.load_workbook(out_path, data_only=True)
    assert SHEET_NAME in wb.sheetnames
    ws = wb[SHEET_NAME]
    assert ws["A1"].value == "테스트현장"
    assert ws["K8"].value == "(주)부현전기"
    assert ws["AD6"].value == "2026년 6월 4일"
    assert ws["G6"].value == "단위테스트 작업"
    assert ws["AD7"].value == "지상2층"
    assert ws["K11"].value == "홍길동"
    assert ws["U21"].value == "UNIT-TEST"
    assert ws["U22"].value == "3ton"
    assert ws["K25"].value == 4000
    assert ws["AC25"].value == 3000
    wb.close()
