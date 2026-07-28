from __future__ import annotations

from datetime import date
from pathlib import Path
from types import SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.config.settings import settings
from app.core.auth import get_current_user_with_bypass, get_db
from app.core.database import Base
from app.core.enums import Role
from app.main import app
from app.modules.safety_ledgers.models import (
    SafetyCardExpense,
    SafetyVehicle,
    SafetyVehicleDriver,
    SafetyVehicleLog,
)
from app.modules.safety_ledgers.workbook_export import build_card_workbook, build_vehicle_workbook


def _client(tmp_path: Path, monkeypatch, *, user_name: str = "정상익") -> TestClient:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    testing_session = sessionmaker(bind=engine, autocommit=False, autoflush=False)

    def override_db():
        db = testing_session()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user_with_bypass] = lambda: SimpleNamespace(
        id=777,
        name=user_name,
        role=Role.HQ_SAFE,
        must_change_password=False,
    )
    monkeypatch.setattr(settings, "storage_root", tmp_path / "storage")
    monkeypatch.setattr(settings, "safety_ledger_nas_root", None)
    monkeypatch.setattr(settings, "openai_api_key", None)
    return TestClient(app)


def test_mobile_session_default_is_seven_days():
    assert settings.access_token_expire_minutes == 60 * 24 * 7


def test_pilot_account_vehicle_and_card_scopes(tmp_path: Path, monkeypatch):
    shared_client = _client(tmp_path / "shared", monkeypatch, user_name="정상익")
    shared = shared_client.get("/safety-ledgers/bootstrap")
    assert shared.status_code == 200
    assert shared.json()["vehicle"]["plate_number"] == "181하8339"
    assert shared.json()["card_account"]["scope"] == "SAFETY_SHARED"
    assert shared.json()["card_account"]["card_number_masked"] == "5585-03**-****-6925"

    jo_client = _client(tmp_path / "jo", monkeypatch, user_name="조동문")
    jo = jo_client.get("/safety-ledgers/bootstrap")
    assert jo.status_code == 200
    assert jo.json()["vehicle"]["vehicle_name"] == "그랜저"
    assert jo.json()["vehicle"]["plate_number"] == "160하3180"
    assert jo.json()["vehicle"]["drivers"] == ["조동문"]
    assert jo.json()["card_account"]["scope"] == "JO_DONGMUN"
    assert jo.json()["card_account"]["card_last4"] == "3946"

    blocked_client = _client(tmp_path / "blocked", monkeypatch, user_name="김복수")
    assert blocked_client.get("/safety-ledgers/bootstrap").status_code == 403


def test_upload_review_and_export_flow(tmp_path: Path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    boot = client.get("/safety-ledgers/bootstrap")
    assert boot.status_code == 200
    assert boot.json()["vehicle"]["plate_number"] == "181하8339"
    assert boot.json()["vehicle"]["drivers"] == ["정상익", "박영선"]

    vehicle = client.post(
        "/safety-ledgers/vehicle-logs",
        data={
            "driven_on": "2026-07-28",
            "driver_name": "정상익",
            "odometer_km": "12345",
            "trip_km": "42",
            "purpose": "본사 안전점검",
        },
        files={"photo": ("dashboard.jpg", b"\xff\xd8\xff\xd9", "image/jpeg")},
    )
    assert vehicle.status_code == 200, vehicle.text
    vehicle_row = vehicle.json()
    assert vehicle_row["trip_km"] == 42
    assert vehicle_row["extraction_status"] == "NEEDS_REVIEW"

    confirmed_vehicle = client.patch(
        f"/safety-ledgers/vehicle-logs/{vehicle_row['id']}",
        json={
            "driven_on": "2026-07-28",
            "driver_name": "정상익",
            "odometer_km": 12345,
            "trip_km": 42,
            "use_type": "6.업무용(왕복)",
            "purpose": "본사 안전점검",
            "confirm": True,
        },
    )
    assert confirmed_vehicle.status_code == 200
    assert confirmed_vehicle.json()["extraction_status"] == "CONFIRMED"

    expense = client.post(
        "/safety-ledgers/card-expenses",
        data={
            "used_at": "2026-07-28T12:30",
            "merchant": "안전식당",
            "amount": "22000",
            "description": "중식비",
            "card_last4": "8339",
        },
        files={"receipt": ("receipt.jpg", b"\xff\xd8\xff\xd9", "image/jpeg")},
    )
    assert expense.status_code == 200, expense.text
    expense_row = expense.json()
    assert expense_row["merchant"] == "안전식당"
    assert expense_row["card_last4"] == "6925"

    confirmed_expense = client.patch(
        f"/safety-ledgers/card-expenses/{expense_row['id']}",
        json={
            "used_at": "2026-07-28T12:30:00",
            "site_name": None,
            "merchant": "안전식당",
            "amount": 22000,
            "description": "중식비",
            "card_last4": "8339",
            "note": None,
            "confirm": True,
        },
    )
    assert confirmed_expense.status_code == 200
    assert confirmed_expense.json()["extraction_status"] == "CONFIRMED"
    assert confirmed_expense.json()["card_last4"] == "6925"

    updated_card = client.put("/safety-ledgers/card-account", json={"card_number": "1234"})
    assert updated_card.status_code == 200
    assert updated_card.json()["card_number_masked"] == "****-****-****-1234"

    for kind in ("vehicle", "card"):
        response = client.get(f"/safety-ledgers/exports/{kind}")
        assert response.status_code == 200
        assert response.content[:2] == b"PK"

    app.dependency_overrides.clear()


def test_workbooks_preserve_expected_columns(tmp_path: Path):
    vehicle = SimpleNamespace(
        vehicle_name="투싼",
        plate_number="181하8339",
        department="안전보건실",
        ownership_type="0.회사",
        drivers=[
            SimpleNamespace(driver_name="정상익", is_active=True),
            SimpleNamespace(driver_name="박영선", is_active=True),
        ],
    )
    log = SimpleNamespace(
        id=1,
        driven_on=date(2026, 7, 28),
        created_at=date(2026, 7, 28),
        driver_name="박영선",
        use_type="6.업무용(왕복)",
        odometer_km=12345,
        trip_km=35,
        purpose="현장 점검",
    )
    expense = SimpleNamespace(
        id=1,
        used_at=None,
        created_at=date(2026, 7, 28),
        site_name="본사",
        merchant="안전식당",
        amount=22000,
        description="중식비",
        card_last4="8339",
        note=None,
    )
    vehicle_path = build_vehicle_workbook(vehicle, [log], tmp_path / "vehicle.xlsx")
    card_path = build_card_workbook([expense], tmp_path / "card.xlsx")
    vehicle_sheet = load_workbook(vehicle_path, data_only=False).active
    card_sheet = load_workbook(card_path, data_only=False).active
    assert vehicle_sheet["E3"].value == "181하8339"
    assert vehicle_sheet["E7"].value == "박영선"
    assert vehicle_sheet["G7"].value == 35
    assert card_sheet["D4"].value == "안전식당"
    assert card_sheet["E4"].value == 22000
