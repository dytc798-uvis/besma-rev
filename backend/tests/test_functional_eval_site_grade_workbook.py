from __future__ import annotations

import io
from datetime import date

import openpyxl
import pytest

from app.modules.functional_eval.eval_catalog import get_criteria
from app.modules.functional_eval.site_grade_workbook import (
    generate_site_grade_workbook_bytes,
    resolve_site_grade_template_path,
    site_grade_export_filename,
)


def test_template_exists():
    path = resolve_site_grade_template_path()
    assert path.is_file()
    assert "기능인등급" in path.name


def test_export_filename_format():
    assert site_grade_export_filename(date(2026, 6, 8)) == "현장별 기능인등급-20260608.xlsx"


def test_generate_workbook_fills_three_sheets():
    workers = [
        {
            "id": 1,
            "site_code": "24025",
            "site_name": "[1.대우건설] 청라C18BL 테스트",
            "row_no": 1,
            "name": "홍길동",
            "age_label": "40세",
            "rrn_masked": "800101-*******",
            "position_name": "반장",
            "job_name": "철근",
            "phone_mobile": "010-1234-5678",
            "functional_assessment": {
                "is_complete": True,
                "grade_code": "A",
                "scores": {c["id"]: "TOP" for c in get_criteria("FUNCTIONAL")},
            },
            "safety_assessment": {
                "is_complete": True,
                "grade_code": "S",
                "scores": {c["id"]: "MID" for c in get_criteria("SAFETY")},
            },
        },
    ]
    raw = generate_site_grade_workbook_bytes(workers)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sheet1 = next(n for n in wb.sheetnames if n.startswith("1."))
    sheet_fn = next(n for n in wb.sheetnames if "2-1" in n)
    sheet_sf = next(n for n in wb.sheetnames if "2-2" in n)
    ws1 = wb[sheet1]
    assert ws1.cell(6, 3).value == 1
    assert ws1.cell(6, 5).value == "홍길동"
    assert ws1.cell(6, 17).value == "A"
    assert ws1.cell(6, 18).value == "S"
    ws_fn = wb[sheet_fn]
    assert ws_fn.cell(7, 8).value == "O"
    ws_sf = wb[sheet_sf]
    assert ws_sf.cell(7, 9).value == "O"
    wb.close()


def test_workbook_multi_row_vlookup_column_index_stable():
    """행 복사 시 VLOOKUP 열 인덱스(7,8)가 행번호로 오염되지 않아야 한다."""
    criteria = get_criteria("SAFETY")
    workers = []
    for i in range(12):
        workers.append(
            {
                "id": i + 1,
                "site_code": "24025",
                "site_name": "C18",
                "row_no": i + 1,
                "name": f"근로자{i + 1}",
                "rrn_masked": f"90010{i + 1}-*******",
                "position_name": "반장",
                "job_name": "철근",
                "functional_assessment": {
                    "is_complete": True,
                    "grade_code": "S",
                    "scores": {c["id"]: "TOP" for c in get_criteria("FUNCTIONAL")},
                },
                "safety_assessment": {
                    "is_complete": True,
                    "grade_code": "S",
                    "scores": {c["id"]: "TOP" for c in criteria},
                },
            }
        )
    raw = generate_site_grade_workbook_bytes(workers)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
    sheet_sf = next(n for n in wb.sheetnames if "2-2" in n)
    ws = wb[sheet_sf]
    for row in range(8, 19):
        f6 = str(ws.cell(row, 6).value or "")
        f7 = str(ws.cell(row, 7).value or "")
        assert ",7,0)" in f6, f"row {row} col6 corrupted: {f6}"
        assert ",8,0)" in f7, f"row {row} col7 corrupted: {f7}"
        assert f",7,0)" not in f6.replace(",7,0)", "", 1) or f6.count(",7,0)") == 2
    wb.close()


def test_unevaluated_worker_clears_template_sample_marks():
    """미평가·미사용 행에 템플릿 샘플 O가 남지 않아야 한다."""
    workers = [
        {
            "id": 1,
            "site_code": "24025",
            "site_name": "C18",
            "row_no": 1,
            "name": "미평가",
            "rrn_masked": "900101-*******",
            "functional_assessment": None,
            "safety_assessment": None,
        }
    ]
    raw = generate_site_grade_workbook_bytes(workers)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    for sheet_key in ("2-1", "2-2"):
        ws = wb[next(n for n in wb.sheetnames if sheet_key in n)]
        for row in range(7, 20):
            marks = [ws.cell(row, c).value for c in range(8, 40)]
            assert not any(m == "O" for m in marks), f"{sheet_key} row {row} has leftover O marks"
    wb.close()


def test_sheet1_birth_from_rrn_when_age_missing():
    workers = [
        {
            "id": 1,
            "site_code": "24025",
            "site_name": "C18",
            "row_no": 1,
            "name": "홍길동",
            "rrn_masked": "900101-*******",
            "position_name": "반장",
            "job_name": "철근",
            "functional_assessment": None,
            "safety_assessment": None,
        }
    ]
    raw = generate_site_grade_workbook_bytes(workers)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sheet1 = next(n for n in wb.sheetnames if n.startswith("1."))
    assert wb[sheet1].cell(6, 6).value == "1990.01.01"
    wb.close()
