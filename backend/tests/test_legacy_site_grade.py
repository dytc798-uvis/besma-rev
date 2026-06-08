from __future__ import annotations

import io

import openpyxl

from app.modules.functional_eval.legacy_site_grade import apply_legacy_assessments, legacy_worker_record
from app.modules.functional_eval.site_grade_workbook import generate_site_grade_workbook_bytes


def test_legacy_data_has_gangmin():
    record = legacy_worker_record("24018", "강민", "880409-1170112")
    assert record is not None
    assert record["sheet1_functional_grade"] == "S"


def test_apply_legacy_fills_unevaluated_worker_for_export():
    legacy = legacy_worker_record("24018", "강민", "880409-1170112")
    assert legacy is not None
    worker = {
        "id": 1,
        "site_code": "24018",
        "site_name": legacy["site_name"],
        "row_no": 1,
        "name": legacy["name"],
        "age_label": legacy["age_label"],
        "rrn_masked": legacy["rrn_masked"],
        "position_name": legacy["position_name"],
        "job_name": legacy["job_name"],
        "phone_mobile": legacy["phone_mobile"],
        "functional_assessment": None,
        "safety_assessment": None,
    }
    apply_legacy_assessments(worker)
    assert worker["functional_assessment"]["grade_code"] == "S"
    assert worker["safety_assessment"]["grade_code"] == "S"

    raw = generate_site_grade_workbook_bytes([worker])
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws1 = wb[next(n for n in wb.sheetnames if n.startswith("1."))]
    ws_sf = wb[next(n for n in wb.sheetnames if "2-2" in n)]
    assert ws1.cell(6, 17).value == "S"
    assert ws1.cell(6, 18).value == "S"
    assert ws_sf.cell(7, 8).value == "O"
    wb.close()


def test_db_complete_overrides_legacy():
    legacy = legacy_worker_record("24018", "강민", "880409-1170112")
    assert legacy is not None
    worker = {
        "site_code": "24018",
        "name": legacy["name"],
        "rrn_masked": legacy["rrn_masked"],
        "functional_assessment": {"is_complete": True, "grade_code": "B", "scores": {}},
        "safety_assessment": None,
    }
    apply_legacy_assessments(worker)
    assert worker["functional_assessment"]["grade_code"] == "B"
    assert worker["safety_assessment"]["grade_code"] == "S"
